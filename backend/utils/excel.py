from __future__ import annotations

import io
import uuid
from datetime import date, datetime, timedelta, timezone
from typing import Dict

import pandas as pd
from fastapi import APIRouter, BackgroundTasks, Depends, File, HTTPException, UploadFile
from fastapi.responses import StreamingResponse
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from pydantic import BaseModel
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession

from deps import AsyncSessionLocal, get_db
from models.products import Product as ProductORM, StockBatch as BatchORM, StockMovement as MovementORM
from routers.auth_helpers import User, get_current_user

router = APIRouter(prefix="/api", tags=["excel"])

# ── In-memory job store ───────────────────────────────────────────────────────
bulk_upload_jobs: Dict[str, Dict] = {}

# ── Column auto-detection keywords ────────────────────────────────────────────
COLUMN_KEYWORDS = {
    "sku": ["sku", "code", "product_code", "item_code", "product code", "item code"],
    "name": ["name", "product_name", "item_name", "medicine", "product name", "medicine name", "item name", "product", "item", "description"],
    "price": ["price", "mrp", "rate", "selling_price", "selling price", "mrp_per_unit", "mrp per unit", "unit_price", "unit price"],
    "quantity": ["quantity", "qty", "stock", "qty_on_hand", "qty on hand", "packs", "units", "opening_qty", "opening qty"],
    "expiry_date": ["expiry", "expiry_date", "exp_date", "exp", "expiry date", "exp date", "expires", "expire"],
    "batch_number": ["batch", "batch_number", "batch_no", "batch no", "lot", "lot_number", "lot number", "batch number"],
    "brand": ["brand", "manufacturer", "company", "mfr"],
    "category": ["category", "type", "group", "class"],
    "cost_price": ["cost", "cost_price", "purchase_price", "purchase price", "cost price", "buy_price", "buy price", "ptr"],
    "gst_percent": ["gst", "gst_percent", "tax", "gst percent", "tax_rate", "tax rate"],
    "hsn_code": ["hsn", "hsn_code", "hsn code"],
    "units_per_pack": ["units_per_pack", "pack_size", "units per pack", "pack size", "strip_qty", "strip qty"],
}

REQUIRED_FIELDS = ["sku", "name", "price", "quantity", "expiry_date", "batch_number"]
OPTIONAL_FIELDS = ["brand", "category", "cost_price", "gst_percent", "hsn_code", "units_per_pack"]


def _rupees_to_paise(rupees: float) -> int:
    return int(round(rupees * 100))


# ── Pydantic models ──────────────────────────────────────────────────────────

class BulkUploadValidateRequest(BaseModel):
    job_id: str
    column_mapping: Dict[str, str]


class BulkUploadImportRequest(BaseModel):
    job_id: str
    import_valid_only: bool = True


# ── Routes ────────────────────────────────────────────────────────────────────

@router.get("/inventory/bulk-upload/template")
async def download_bulk_upload_template(current_user: User = Depends(get_current_user)):
    wb = Workbook()
    ws = wb.active
    ws.title = "Inventory Template"

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    required_fill = PatternFill(start_color="FF6B6B", end_color="FF6B6B", fill_type="solid")
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )

    headers = [
        ("SKU *", True), ("Name *", True), ("Brand", False), ("Category", False),
        ("Batch Number *", True), ("Expiry Date *", True), ("Quantity (Packs) *", True),
        ("MRP per Unit *", True), ("Cost Price per Unit", False), ("GST %", False),
        ("HSN Code", False), ("Units per Pack", False),
    ]
    for col, (header, is_required) in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = required_fill if is_required else header_fill
        cell.alignment = Alignment(horizontal="center")
        cell.border = thin_border
        ws.column_dimensions[ws.cell(row=1, column=col).column_letter].width = 18

    sample_data = [
        ["MED001", "Paracetamol 500mg", "Cipla", "Tablets", "BTN2024001", "2025-12-31", 100, 2.50, 1.80, 12, "30049099", 10],
        ["MED002", "Amoxicillin 250mg", "Sun Pharma", "Antibiotics", "BTN2024002", "2025-06-30", 50, 8.00, 6.50, 12, "30042011", 10],
        ["MED003", "Vitamin D3 Capsules", "Abbott", "Vitamins", "BTN2024003", "2026-03-15", 200, 5.50, 4.20, 5, "21069099", 15],
    ]
    for row_num, row_data in enumerate(sample_data, 2):
        for col_num, value in enumerate(row_data, 1):
            ws.cell(row=row_num, column=col_num, value=value).border = thin_border

    ws_instr = wb.create_sheet("Instructions")
    instructions = [
        ["Excel Bulk Upload Instructions"], [""],
        ["Required Fields (marked with *)"],
        ["- SKU: Unique product identifier"],
        ["- Name: Product/medicine name"],
        ["- Batch Number: Batch identifier"],
        ["- Expiry Date: Format YYYY-MM-DD"],
        ["- Quantity: Number of packs"],
        ["- MRP per Unit: Maximum retail price per unit"],
        [""], ["Optional Fields"],
        ["- Brand: Manufacturer or brand name"],
        ["- Category: Product category"],
        ["- Cost Price: Purchase price per unit"],
        ["- GST %: Tax percentage (default 5%)"],
        ["- HSN Code: Harmonized System code"],
        ["- Units per Pack: Number of units in a pack (default 1)"],
        [""], ["Important Notes:"],
        ["- Dates must be in YYYY-MM-DD format (e.g., 2025-12-31)"],
        ["- Expiry date must be in the future"],
        ["- MRP must be greater than Cost Price"],
        ["- Duplicate SKU+Batch combinations will update existing batches"],
        ["- Maximum 5000 rows per upload"],
    ]
    for row_num, row in enumerate(instructions, 1):
        cell = ws_instr.cell(row=row_num, column=1, value=row[0] if row else "")
        if row_num == 1:
            cell.font = Font(bold=True, size=14)
    ws_instr.column_dimensions["A"].width = 60

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=inventory_upload_template.xlsx"},
    )


@router.post("/inventory/bulk-upload/parse")
async def parse_bulk_upload_file(file: UploadFile = File(...), current_user: User = Depends(get_current_user)):
    if not file.filename.endswith((".xlsx", ".xls")):
        raise HTTPException(status_code=400, detail="Only Excel files (.xlsx, .xls) are supported")

    try:
        contents = await file.read()
        engine = "openpyxl" if file.filename.endswith(".xlsx") else "xlrd"
        df = pd.read_excel(io.BytesIO(contents), engine=engine)

        if len(df) > 5000:
            raise HTTPException(status_code=400, detail=f"File has {len(df)} rows. Maximum allowed is 5000 rows.")
        if len(df) == 0:
            raise HTTPException(status_code=400, detail="File is empty or has no data rows")

        file_columns = list(df.columns.astype(str))
        auto_mappings: dict = {}
        for sys_field, keywords in COLUMN_KEYWORDS.items():
            for col in file_columns:
                col_lower = str(col).lower().strip()
                if col_lower in keywords or any(kw in col_lower for kw in keywords):
                    if sys_field not in auto_mappings:
                        auto_mappings[sys_field] = col
                    break

        job_id = str(uuid.uuid4())
        bulk_upload_jobs[job_id] = {
            "status": "parsed",
            "filename": file.filename,
            "total_rows": len(df),
            "columns": file_columns,
            "data": df.to_dict(orient="records"),
            "created_at": datetime.now(timezone.utc).isoformat(),
            "created_by": current_user.id,
            "pharmacy_id": str(current_user.pharmacy_id),
        }

        return {
            "job_id": job_id,
            "filename": file.filename,
            "total_rows": len(df),
            "columns": file_columns,
            "auto_mappings": auto_mappings,
            "sample_data": df.head(5).fillna("").to_dict(orient="records"),
            "required_fields": REQUIRED_FIELDS,
            "optional_fields": OPTIONAL_FIELDS,
        }
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Failed to parse Excel file: {str(e)}")


@router.post("/inventory/bulk-upload/validate")
async def validate_bulk_upload(
    request: BulkUploadValidateRequest,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    job_id = request.job_id
    column_mapping = request.column_mapping

    if job_id not in bulk_upload_jobs:
        raise HTTPException(status_code=404, detail="Upload job not found. Please re-upload the file.")

    job = bulk_upload_jobs[job_id]
    data = job["data"]

    missing_required = [f for f in REQUIRED_FIELDS if f not in column_mapping or not column_mapping[f]]
    if missing_required:
        raise HTTPException(status_code=400, detail=f"Missing required field mappings: {', '.join(missing_required)}")

    pid = current_user.pharmacy_id

    # Load existing products and batches for duplicate detection
    products = (await db.execute(
        select(ProductORM.id, ProductORM.sku).where(ProductORM.pharmacy_id == pid, ProductORM.deleted_at.is_(None))
    )).all()
    existing_products = {p.sku: str(p.id) for p in products}

    batches = (await db.execute(
        select(BatchORM.id, BatchORM.product_id, BatchORM.batch_number)
        .join(ProductORM, ProductORM.id == BatchORM.product_id)
        .where(BatchORM.pharmacy_id == pid)
    )).all()
    product_id_to_sku = {str(p.id): p.sku for p in products}
    existing_batches = set()
    for b in batches:
        sku = product_id_to_sku.get(str(b.product_id))
        if sku:
            existing_batches.add(f"{sku}_{b.batch_number}")

    validation_results = []
    valid_count = error_count = warning_count = 0
    today = date.today()

    for row_idx, row in enumerate(data, start=2):
        row_errors: list = []
        row_warnings: list = []
        row_data: dict = {}

        for sys_field, file_col in column_mapping.items():
            value = row.get(file_col)
            if pd.isna(value) if hasattr(pd, "isna") else (value == "" or value is None):
                value = None
            else:
                value = str(value).strip() if not isinstance(value, (int, float)) else value
            row_data[sys_field] = value

        for field in REQUIRED_FIELDS:
            if row_data.get(field) is None or row_data.get(field) == "":
                row_errors.append(f"Missing required field: {field}")

        sku = row_data.get("sku")
        if sku:
            sku = str(sku).strip().upper()
            row_data["sku"] = sku

        name = row_data.get("name")
        if name and len(str(name)) < 2:
            row_errors.append("Product name must be at least 2 characters")

        try:
            price = float(row_data.get("price", 0) or 0)
            row_data["price"] = price
            if price <= 0:
                row_errors.append("MRP must be greater than 0")
        except (ValueError, TypeError):
            row_errors.append("Invalid MRP value")
            price = 0

        try:
            cost_price = float(row_data.get("cost_price", 0) or 0)
            row_data["cost_price"] = cost_price
            if cost_price > 0 and price > 0 and cost_price >= price:
                row_warnings.append("Cost price should be less than MRP")
        except (ValueError, TypeError):
            row_warnings.append("Invalid cost price, will use 0")
            row_data["cost_price"] = 0

        try:
            quantity = int(float(row_data.get("quantity", 0) or 0))
            row_data["quantity"] = quantity
            if quantity < 0:
                row_errors.append("Quantity cannot be negative")
        except (ValueError, TypeError):
            row_errors.append("Invalid quantity value")

        expiry_str = row_data.get("expiry_date")
        if expiry_str:
            try:
                if isinstance(expiry_str, datetime):
                    expiry_date = expiry_str.date()
                elif isinstance(expiry_str, str):
                    for fmt in ["%Y-%m-%d", "%d-%m-%Y", "%d/%m/%Y", "%Y/%m/%d", "%m/%d/%Y"]:
                        try:
                            expiry_date = datetime.strptime(str(expiry_str).strip(), fmt).date()
                            break
                        except ValueError:
                            continue
                    else:
                        raise ValueError("Unknown date format")
                else:
                    expiry_date = pd.to_datetime(expiry_str).date()

                row_data["expiry_date"] = expiry_date.isoformat()
                if expiry_date <= today:
                    row_warnings.append("Expiry date is in the past or today")
                elif expiry_date <= today + timedelta(days=30):
                    row_warnings.append("Expiry date is within 30 days")
            except Exception:
                row_errors.append("Invalid expiry date format. Use YYYY-MM-DD")

        batch_no = row_data.get("batch_number")
        if batch_no:
            batch_no = str(batch_no).strip().upper()
            row_data["batch_number"] = batch_no

        if sku and batch_no and f"{sku}_{batch_no}" in existing_batches:
            row_warnings.append("Batch already exists - will update existing batch")

        try:
            gst = float(row_data.get("gst_percent", 5) or 5)
            row_data["gst_percent"] = gst
            if gst < 0 or gst > 100:
                row_warnings.append("GST % should be between 0 and 100")
        except (ValueError, TypeError):
            row_data["gst_percent"] = 5

        try:
            row_data["units_per_pack"] = max(1, int(float(row_data.get("units_per_pack", 1) or 1)))
        except (ValueError, TypeError):
            row_data["units_per_pack"] = 1

        if row_errors:
            status = "error"
            error_count += 1
        elif row_warnings:
            status = "warning"
            warning_count += 1
            valid_count += 1
        else:
            status = "valid"
            valid_count += 1

        validation_results.append({"row_number": row_idx, "status": status, "errors": row_errors, "warnings": row_warnings, "data": row_data})

    bulk_upload_jobs[job_id].update({
        "status": "validated",
        "validation_results": validation_results,
        "column_mapping": column_mapping,
        "valid_count": valid_count,
        "error_count": error_count,
        "warning_count": warning_count,
    })

    return {
        "job_id": job_id,
        "total_rows": len(data),
        "valid_count": valid_count,
        "error_count": error_count,
        "warning_count": warning_count,
        "preview_results": validation_results[:10],
        "can_import": valid_count > 0,
    }


@router.post("/inventory/bulk-upload/import")
async def import_bulk_upload(
    request: BulkUploadImportRequest,
    background_tasks: BackgroundTasks,
    current_user: User = Depends(get_current_user),
):
    job_id = request.job_id
    if job_id not in bulk_upload_jobs:
        raise HTTPException(status_code=404, detail="Upload job not found. Please re-upload the file.")

    job = bulk_upload_jobs[job_id]
    if job.get("status") != "validated":
        raise HTTPException(status_code=400, detail="Data must be validated before import")

    validation_results = job.get("validation_results", [])
    rows_to_import = [r for r in validation_results if r["status"] in ["valid", "warning"]]
    if not rows_to_import:
        raise HTTPException(status_code=400, detail="No valid rows to import")

    bulk_upload_jobs[job_id]["status"] = "importing"
    bulk_upload_jobs[job_id]["import_progress"] = {"total": len(rows_to_import), "processed": 0, "success": 0, "failed": 0, "errors": []}

    pharmacy_id = uuid.UUID(job["pharmacy_id"])
    user_id = current_user.id if isinstance(current_user.id, uuid.UUID) else uuid.UUID(current_user.id)

    async def process_import() -> None:
        progress = bulk_upload_jobs[job_id]["import_progress"]
        async with AsyncSessionLocal() as db:
            for idx, result in enumerate(rows_to_import):
                row_data = result["data"]
                row_number = result["row_number"]
                try:
                    sku = row_data.get("sku")
                    name = row_data.get("name")
                    batch_no = row_data.get("batch_number")
                    qty = row_data.get("quantity", 0)
                    mrp_paise = _rupees_to_paise(row_data.get("price", 0))
                    cost_paise = _rupees_to_paise(row_data.get("cost_price", 0))

                    # Find or create product
                    existing = (await db.execute(
                        select(ProductORM).where(ProductORM.pharmacy_id == pharmacy_id, ProductORM.sku == sku)
                    )).scalars().first()

                    if not existing:
                        product = ProductORM(
                            pharmacy_id=pharmacy_id, sku=sku, name=name,
                            brand=row_data.get("brand"), category=row_data.get("category"),
                            units_per_pack=row_data.get("units_per_pack", 1),
                            gst_rate=row_data.get("gst_percent", 5),
                            hsn_code=row_data.get("hsn_code") or "3004",
                            reorder_level=10,
                        )
                        db.add(product)
                        await db.flush()
                        product_id = product.id
                    else:
                        product_id = existing.id
                        if row_data.get("brand") and not existing.brand:
                            existing.brand = row_data["brand"]
                        if row_data.get("category") and not existing.category:
                            existing.category = row_data["category"]

                    # Find or create batch
                    existing_batch = (await db.execute(
                        select(BatchORM).where(
                            BatchORM.pharmacy_id == pharmacy_id,
                            BatchORM.product_id == product_id,
                            BatchORM.batch_number == batch_no,
                        )
                    )).scalars().first()

                    if existing_batch:
                        qty_before = existing_batch.quantity_on_hand
                        existing_batch.quantity_on_hand += qty
                        existing_batch.quantity_received += qty
                        existing_batch.mrp_paise = mrp_paise
                        if cost_paise > 0:
                            existing_batch.cost_price_paise = cost_paise
                        batch_id = existing_batch.id
                        qty_after = existing_batch.quantity_on_hand
                    else:
                        expiry = date.fromisoformat(row_data["expiry_date"]) if row_data.get("expiry_date") else date.today() + timedelta(days=365)
                        batch = BatchORM(
                            pharmacy_id=pharmacy_id, product_id=product_id,
                            batch_number=batch_no, expiry_date=expiry,
                            mrp_paise=mrp_paise, cost_price_paise=cost_paise or mrp_paise,
                            quantity_received=qty, quantity_on_hand=qty,
                        )
                        db.add(batch)
                        await db.flush()
                        batch_id = batch.id
                        qty_before = 0
                        qty_after = qty

                    # Record stock movement
                    db.add(MovementORM(
                        pharmacy_id=pharmacy_id, batch_id=batch_id, product_id=product_id,
                        movement_type="opening_stock", quantity=qty,
                        quantity_before=qty_before, quantity_after=qty_after,
                        reference_type="bulk_upload", user_id=user_id,
                        notes=f"Bulk upload from {job.get('filename', 'Excel file')}",
                    ))

                    await db.flush()
                    progress["success"] += 1
                except Exception as e:
                    await db.rollback()
                    progress["failed"] += 1
                    progress["errors"].append({"row_number": row_number, "error": str(e)})

                progress["processed"] = idx + 1

            await db.commit()

        bulk_upload_jobs[job_id]["status"] = "completed"
        bulk_upload_jobs[job_id]["completed_at"] = datetime.now(timezone.utc).isoformat()

    background_tasks.add_task(process_import)
    return {"job_id": job_id, "message": "Import started", "total_rows": len(rows_to_import)}


@router.get("/inventory/bulk-upload/progress/{job_id}")
async def get_bulk_upload_progress(job_id: str, current_user: User = Depends(get_current_user)):
    if job_id not in bulk_upload_jobs:
        raise HTTPException(status_code=404, detail="Upload job not found")
    job = bulk_upload_jobs[job_id]
    return {
        "job_id": job_id,
        "status": job.get("status"),
        "filename": job.get("filename"),
        "total_rows": job.get("total_rows"),
        "valid_count": job.get("valid_count", 0),
        "error_count": job.get("error_count", 0),
        "warning_count": job.get("warning_count", 0),
        "import_progress": job.get("import_progress"),
        "completed_at": job.get("completed_at"),
    }


@router.get("/inventory/bulk-upload/error-report/{job_id}")
async def download_error_report(job_id: str, current_user: User = Depends(get_current_user)):
    if job_id not in bulk_upload_jobs:
        raise HTTPException(status_code=404, detail="Upload job not found")

    job = bulk_upload_jobs[job_id]
    validation_results = job.get("validation_results", [])
    if not validation_results:
        raise HTTPException(status_code=400, detail="No validation results available")

    wb = Workbook()
    ws = wb.active
    ws.title = "Validation Results"

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    error_fill = PatternFill(start_color="FFCDD2", end_color="FFCDD2", fill_type="solid")
    warning_fill = PatternFill(start_color="FFF9C4", end_color="FFF9C4", fill_type="solid")
    success_fill = PatternFill(start_color="C8E6C9", end_color="C8E6C9", fill_type="solid")
    thin_border = Border(left=Side(style="thin"), right=Side(style="thin"), top=Side(style="thin"), bottom=Side(style="thin"))

    headers = ["Row #", "Status", "SKU", "Name", "Batch #", "Expiry", "Qty", "MRP", "Errors", "Warnings"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border
        cell.alignment = Alignment(horizontal="center")

    for row_idx, result in enumerate(validation_results, 2):
        data = result.get("data", {})
        status = result.get("status", "unknown")
        row_values = [
            result.get("row_number"), status.upper(),
            data.get("sku", ""), data.get("name", ""), data.get("batch_number", ""),
            data.get("expiry_date", ""), data.get("quantity", ""), data.get("price", ""),
            "; ".join(result.get("errors", [])), "; ".join(result.get("warnings", [])),
        ]
        row_fill = error_fill if status == "error" else (warning_fill if status == "warning" else success_fill)
        for col, value in enumerate(row_values, 1):
            cell = ws.cell(row=row_idx, column=col, value=value)
            cell.border = thin_border
            cell.fill = row_fill

    for col, width in enumerate([8, 10, 15, 25, 15, 12, 8, 10, 40, 40], 1):
        ws.column_dimensions[ws.cell(row=1, column=col).column_letter].width = width

    ws_summary = wb.create_sheet("Summary")
    summary_data = [
        ["Validation Summary"], [""],
        ["Total Rows", job.get("total_rows", 0)],
        ["Valid Rows", job.get("valid_count", 0)],
        ["Rows with Warnings", job.get("warning_count", 0)],
        ["Rows with Errors", job.get("error_count", 0)],
        [""], ["File Name", job.get("filename", "")],
        ["Uploaded At", job.get("created_at", "")],
    ]
    for row_num, row in enumerate(summary_data, 1):
        for col, value in enumerate(row, 1):
            cell = ws_summary.cell(row=row_num, column=col, value=value)
            if row_num == 1:
                cell.font = Font(bold=True, size=14)
    ws_summary.column_dimensions["A"].width = 25
    ws_summary.column_dimensions["B"].width = 30

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=validation_report_{job_id[:8]}.xlsx"},
    )
