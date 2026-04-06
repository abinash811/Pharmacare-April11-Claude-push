from fastapi import FastAPI, APIRouter, HTTPException, Depends, Response, Request, Cookie, UploadFile, File, Form, BackgroundTasks
from fastapi.responses import JSONResponse, StreamingResponse
from dotenv import load_dotenv
from starlette.middleware.cors import CORSMiddleware
from motor.motor_asyncio import AsyncIOMotorClient
import os
import logging
from pathlib import Path
from pydantic import BaseModel, Field, ConfigDict, EmailStr
from typing import List, Optional, Dict, Any
import uuid
from datetime import datetime, timezone, timedelta
from passlib.context import CryptContext
import jwt
import httpx
from decimal import Decimal
import json
import io
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows

ROOT_DIR = Path(__file__).parent
load_dotenv(ROOT_DIR / '.env')

# MongoDB connection
mongo_url = os.environ['MONGO_URL']
client = AsyncIOMotorClient(mongo_url)
db = client[os.environ['DB_NAME']]

# Security
pwd_context = CryptContext(schemes=["bcrypt"], deprecated="auto")
SECRET_KEY = os.environ.get('JWT_SECRET', 'your-secret-key-change-in-production')
ALGORITHM = "HS256"
ACCESS_TOKEN_EXPIRE_MINUTES = 60 * 24 * 7  # 7 days

# Configure logging
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

# Create the main app
app = FastAPI()
api_router = APIRouter(prefix="/api")

# ==================== BILL NUMBER GENERATION ====================

async def generate_bill_number(invoice_type: str = "SALE", branch_id: str = None) -> str:
    """
    Generate a unique, sequential bill number using atomic MongoDB operations.
    Uses findOneAndUpdate with upsert for concurrency safety.
    
    Args:
        invoice_type: "SALE" or "SALES_RETURN"
        branch_id: Optional branch ID for future multi-branch support
    
    Returns:
        Formatted bill number like "INV-000001" or "RTN-000001"
    """
    # Determine prefix based on invoice type
    default_prefix = "RTN" if invoice_type == "SALES_RETURN" else "INV"
    
    # Get or create sequence for this prefix
    # Use findOneAndUpdate with upsert for atomic operation
    sequence_doc = await db.bill_number_sequences.find_one_and_update(
        {
            "prefix": default_prefix,
            "branch_id": branch_id  # None for V1
        },
        {
            "$inc": {"current_sequence": 1},
            "$setOnInsert": {
                "id": str(uuid.uuid4()),
                "prefix": default_prefix,
                "branch_id": branch_id,
                "sequence_length": 6,
                "allow_prefix_change": True,
                "created_at": datetime.now(timezone.utc).isoformat()
            },
            "$set": {"updated_at": datetime.now(timezone.utc).isoformat()}
        },
        upsert=True,
        return_document=True,  # Return the updated document
        projection={"_id": 0}
    )
    
    # Extract values
    prefix = sequence_doc.get("prefix", default_prefix)
    sequence = sequence_doc.get("current_sequence", 1)
    length = sequence_doc.get("sequence_length", 6)
    
    # Format bill number with zero-padded sequence
    bill_number = f"{prefix}-{str(sequence).zfill(length)}"
    
    return bill_number


async def get_bill_sequence_settings(prefix: str = "INV", branch_id: str = None) -> dict:
    """Get current sequence settings for a prefix"""
    doc = await db.bill_number_sequences.find_one(
        {"prefix": prefix, "branch_id": branch_id},
        {"_id": 0}
    )
    if not doc:
        return {
            "prefix": prefix,
            "current_sequence": 0,
            "sequence_length": 6,
            "allow_prefix_change": True,
            "next_number": 1
        }
    
    return {
        **doc,
        "next_number": doc.get("current_sequence", 0) + 1
    }


async def validate_and_update_sequence_settings(
    prefix: str,
    starting_number: int,
    sequence_length: int,
    branch_id: str = None
) -> dict:
    """
    Validate and update sequence settings.
    Ensures starting_number is greater than any existing bill number for this prefix.
    """
    # Check current sequence for this prefix
    existing = await db.bill_number_sequences.find_one(
        {"prefix": prefix, "branch_id": branch_id},
        {"_id": 0}
    )
    
    if existing and existing.get("current_sequence", 0) >= starting_number:
        raise HTTPException(
            status_code=400,
            detail=f"Starting number must be greater than last used number ({existing['current_sequence']}) for prefix '{prefix}'"
        )
    
    # Also check if any bills exist with higher numbers for this prefix
    # Find highest bill number for this prefix
    highest_bill = await db.bills.find_one(
        {"bill_number": {"$regex": f"^{prefix}-"}},
        {"_id": 0, "bill_number": 1},
        sort=[("bill_number", -1)]
    )
    
    if highest_bill:
        # Extract sequence number from bill_number like "INV-000123"
        try:
            parts = highest_bill["bill_number"].split("-")
            if len(parts) >= 2:
                existing_seq = int(parts[-1])
                if existing_seq >= starting_number:
                    raise HTTPException(
                        status_code=400,
                        detail=f"Starting number must be greater than highest existing bill number ({existing_seq}) for prefix '{prefix}'"
                    )
        except (ValueError, IndexError):
            pass  # Invalid bill number format, skip check
    
    # Update or create sequence
    result = await db.bill_number_sequences.find_one_and_update(
        {"prefix": prefix, "branch_id": branch_id},
        {
            "$set": {
                "current_sequence": starting_number - 1,  # Will be incremented on next bill
                "sequence_length": sequence_length,
                "updated_at": datetime.now(timezone.utc).isoformat()
            },
            "$setOnInsert": {
                "id": str(uuid.uuid4()),
                "prefix": prefix,
                "branch_id": branch_id,
                "allow_prefix_change": True,
                "created_at": datetime.now(timezone.utc).isoformat()
            }
        },
        upsert=True,
        return_document=True,
        projection={"_id": 0}
    )
    
    return {
        **result,
        "next_number": result.get("current_sequence", 0) + 1
    }

# ==================== MODELS ====================

# User Models
class User(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    email: EmailStr
    name: str
    role: str  # "admin", "manager", "cashier", "inventory_staff"
    password_hash: Optional[str] = None
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))
    updated_at: Optional[datetime] = None
    is_active: bool = True
    created_by: Optional[str] = None
    updated_by: Optional[str] = None

class UserCreate(BaseModel):
    email: EmailStr
    name: str
    password: str
    role: str

class UserLogin(BaseModel):
    email: EmailStr
    password: str

class SessionCreate(BaseModel):
    user_id: str
    session_token: str
    email: str
    name: str
    expires_at: datetime

class ChangePassword(BaseModel):
    current_password: str
    new_password: str

class UserUpdate(BaseModel):
    name: Optional[str] = None
    email: Optional[EmailStr] = None
    role: Optional[str] = None
    is_active: Optional[bool] = None

class Role(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    name: str
    display_name: str
    permissions: List[str]
    is_default: bool = False
    is_super_admin: bool = False
    created_by: Optional[str] = None
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))
    updated_at: Optional[datetime] = None

class RoleCreate(BaseModel):
    name: str
    display_name: str
    permissions: List[str]

class RoleUpdate(BaseModel):
    display_name: Optional[str] = None
    permissions: Optional[List[str]] = None

# Granular Permission Structure
ALL_PERMISSIONS = {
    "dashboard": {
        "display_name": "Dashboard",
        "permissions": [
            {"id": "dashboard:view", "name": "View Dashboard"}
        ]
    },
    "billing": {
        "display_name": "Billing",
        "permissions": [
            {"id": "billing:create", "name": "Create Bills"},
            {"id": "billing:view", "name": "View Bills"},
            {"id": "billing:edit", "name": "Edit Bills"},
            {"id": "billing:delete", "name": "Delete Bills"}
        ]
    },
    "inventory": {
        "display_name": "Inventory",
        "permissions": [
            {"id": "inventory:view", "name": "View Inventory"},
            {"id": "inventory:create", "name": "Add Products"},
            {"id": "inventory:edit", "name": "Edit Products"},
            {"id": "inventory:delete", "name": "Delete Products"},
            {"id": "inventory:batches_view", "name": "View Batches"},
            {"id": "inventory:batches_create", "name": "Add Batches"},
            {"id": "inventory:stock_adjust", "name": "Adjust Stock"}
        ]
    },
    "purchases": {
        "display_name": "Purchases",
        "permissions": [
            {"id": "purchases:create", "name": "Create Purchases"},
            {"id": "purchases:view", "name": "View Purchases"},
            {"id": "purchases:edit", "name": "Edit Purchases"},
            {"id": "purchases:delete", "name": "Delete Purchases"}
        ]
    },
    "purchase_returns": {
        "display_name": "Purchase Returns",
        "permissions": [
            {"id": "purchase_returns:create", "name": "Create Returns"},
            {"id": "purchase_returns:view", "name": "View Returns"},
            {"id": "purchase_returns:confirm", "name": "Confirm Returns"}
        ]
    },
    "sales_returns": {
        "display_name": "Sales Returns",
        "permissions": [
            {"id": "sales_returns:create", "name": "Create Returns"},
            {"id": "sales_returns:view", "name": "View Returns"},
            {"id": "sales_returns:process", "name": "Process Returns"}
        ]
    },
    "customers": {
        "display_name": "Customers",
        "permissions": [
            {"id": "customers:view", "name": "View Customers"},
            {"id": "customers:create", "name": "Add Customers"},
            {"id": "customers:edit", "name": "Edit Customers"},
            {"id": "customers:delete", "name": "Delete Customers"}
        ]
    },
    "reports": {
        "display_name": "Reports",
        "permissions": [
            {"id": "reports:view", "name": "View Reports"},
            {"id": "reports:export", "name": "Export Reports"}
        ]
    },
    "settings": {
        "display_name": "Settings",
        "permissions": [
            {"id": "settings:view", "name": "View Settings"},
            {"id": "settings:edit", "name": "Edit Settings"}
        ]
    },
    "users": {
        "display_name": "User Management",
        "permissions": [
            {"id": "users:view", "name": "View Users"},
            {"id": "users:create", "name": "Create Users"},
            {"id": "users:edit", "name": "Edit Users"},
            {"id": "users:delete", "name": "Deactivate Users"}
        ]
    },
    "roles": {
        "display_name": "Roles & Permissions",
        "permissions": [
            {"id": "roles:view", "name": "View Roles"},
            {"id": "roles:create", "name": "Create Roles"},
            {"id": "roles:edit", "name": "Edit Roles"},
            {"id": "roles:delete", "name": "Delete Roles"}
        ]
    },
    "suppliers": {
        "display_name": "Suppliers",
        "permissions": [
            {"id": "suppliers:view", "name": "View Suppliers"},
            {"id": "suppliers:create", "name": "Create Suppliers"},
            {"id": "suppliers:edit", "name": "Edit Suppliers"},
            {"id": "suppliers:deactivate", "name": "Deactivate Suppliers"}
        ]
    }
}

# Default Roles Configuration
DEFAULT_ROLES = [
    {
        "name": "admin",
        "display_name": "Administrator",
        "permissions": ["*"],
        "is_default": True,
        "is_super_admin": True
    },
    {
        "name": "manager",
        "display_name": "Manager",
        "permissions": [
            "dashboard:view", "billing:create", "billing:view", "billing:edit",
            "inventory:view", "inventory:edit", "inventory:create", "inventory:batches_view",
            "inventory:batches_create", "inventory:stock_adjust",
            "purchases:create", "purchases:view", "purchase_returns:create", 
            "purchase_returns:view", "sales_returns:create", "sales_returns:view",
            "customers:view", "customers:edit", "customers:create", "reports:view"
        ],
        "is_default": True,
        "is_super_admin": False
    },
    {
        "name": "cashier",
        "display_name": "Cashier",
        "permissions": [
            "dashboard:view", "billing:create", "billing:view", "inventory:view",
            "sales_returns:create", "sales_returns:view", "customers:view", 
            "customers:edit", "customers:create"
        ],
        "is_default": True,
        "is_super_admin": False
    },
    {
        "name": "inventory_staff",
        "display_name": "Inventory Staff",
        "permissions": [
            "dashboard:view", "inventory:view", "inventory:edit", "inventory:create",
            "inventory:batches_view", "inventory:batches_create", "inventory:stock_adjust",
            "purchases:create", "purchases:view", "purchase_returns:create", 
            "purchase_returns:view"
        ],
        "is_default": True,
        "is_super_admin": False
    }
]

async def has_permission(user_role: str, permission: str) -> bool:
    """Check if a user role has a specific permission"""
    # Fetch role from database
    role = await db.roles.find_one({"name": user_role}, {"_id": 0})
    
    if not role:
        return False
    
    user_permissions = role.get("permissions", [])
    
    # Super Admin or wildcard has all permissions
    if "*" in user_permissions or role.get("is_super_admin", False):
        return True
    
    return permission in user_permissions

def require_permission(permission: str):
    """Decorator to check if user has required permission"""
    def decorator(func):
        async def wrapper(*args, **kwargs):
            # Get current_user from kwargs
            current_user = kwargs.get('current_user')
            if not current_user:
                raise HTTPException(status_code=401, detail="Not authenticated")
            
            if not has_permission(current_user.role, permission):
                raise HTTPException(
                    status_code=403, 
                    detail=f"Permission denied. Required permission: {permission}"
                )
            
            return await func(*args, **kwargs)
        return wrapper
    return decorator

# Product Models (Master Data) - Phase 0
class Product(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    sku: str  # Unique product code (primary business identifier)
    name: str
    manufacturer: Optional[str] = None  # Phase 0: manufacturer field
    brand: Optional[str] = None
    pack_size: Optional[str] = None  # e.g., "Strip", "Box" - Display label
    units_per_pack: int = 1  # Numeric: how many units in one pack (e.g., 10 tablets per strip)
    uom: Optional[str] = "units"  # Unit of measure (units, ml, gm)
    category: Optional[str] = None
    barcode: Optional[str] = None  # EAN-13/UPC barcode for scanner
    # Backward compatibility: support both old and new field names
    default_mrp: Optional[float] = None  # Legacy field
    default_mrp_per_unit: float = 0  # Phase 0: MRP per unit (not per pack)
    default_ptr_per_unit: Optional[float] = None  # Phase 0: PTR (Price to Retailer) per unit
    landing_price_per_unit: Optional[float] = None  # LP (Landing Price) = PTR after schemes, updated on purchase
    gst_percent: float = 5.0
    hsn_code: Optional[str] = None
    description: Optional[str] = None
    low_stock_threshold: Optional[int] = None  # Legacy field
    low_stock_threshold_units: int = 10  # Phase 0: Alert threshold in units
    # Drug Schedule: OTC = over the counter, H = prescription required, H1 = prescription + 3yr register, X = narcotic
    schedule: Optional[str] = "OTC"
    status: str = "active"  # active, inactive
    created_by: Optional[str] = None
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))
    updated_by: Optional[str] = None
    updated_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))
    
    @classmethod
    def model_validate(cls, obj):
        # Backward compatibility: normalize old field names
        if isinstance(obj, dict):
            if 'default_mrp' in obj and 'default_mrp_per_unit' not in obj:
                obj['default_mrp_per_unit'] = obj['default_mrp']
            if 'low_stock_threshold' in obj and 'low_stock_threshold_units' not in obj:
                obj['low_stock_threshold_units'] = obj['low_stock_threshold']
        return super().model_validate(obj)

class ProductCreate(BaseModel):
    sku: str
    name: str
    manufacturer: Optional[str] = None
    brand: Optional[str] = None
    pack_size: Optional[str] = None
    units_per_pack: int = 1  # How many units (tablets) in one pack (strip)
    uom: Optional[str] = "units"
    category: Optional[str] = None
    default_mrp_per_unit: Optional[float] = None  # Phase 0 field
    default_mrp: Optional[float] = None  # Legacy support
    default_ptr_per_unit: Optional[float] = None
    gst_percent: float = 5.0
    hsn_code: Optional[str] = None
    description: Optional[str] = None
    low_stock_threshold_units: Optional[int] = 10
    low_stock_threshold: Optional[int] = None  # Legacy support
    # Drug Schedule: OTC = over the counter, H = prescription required, H1 = prescription + 3yr register, X = narcotic
    schedule: Optional[str] = "OTC"
    status: str = "active"

class ProductUpdate(BaseModel):
    name: Optional[str] = None
    manufacturer: Optional[str] = None
    brand: Optional[str] = None
    units_per_pack: Optional[int] = None
    pack_size: Optional[str] = None
    uom: Optional[str] = None
    category: Optional[str] = None
    default_mrp_per_unit: Optional[float] = None
    default_ptr_per_unit: Optional[float] = None
    gst_percent: Optional[float] = None
    hsn_code: Optional[str] = None
    description: Optional[str] = None
    low_stock_threshold_units: Optional[int] = None
    # Drug Schedule: OTC = over the counter, H = prescription required, H1 = prescription + 3yr register, X = narcotic
    schedule: Optional[str] = None
    status: Optional[str] = None

# Stock Batch Models (Inventory) - Phase 0
class StockBatch(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    product_sku: str  # Phase 0: FK to product SKU
    batch_no: str
    manufacture_date: Optional[datetime] = None  # Phase 0: manufacture date
    expiry_date: datetime
    qty_on_hand: int  # Quantity in PACKS (strips). Total units = qty_on_hand × product.units_per_pack
    cost_price_per_unit: float  # Phase 0: Cost price per individual unit
    mrp_per_unit: float  # Phase 0: MRP per individual unit
    ptr_per_unit: Optional[float] = None  # PTR (Price to Retailer) per unit
    lp_per_unit: Optional[float] = None  # LP (Landing Price) per unit = PTR after schemes
    supplier_name: Optional[str] = None
    supplier_invoice_no: Optional[str] = None  # Phase 0: supplier invoice number
    received_date: Optional[datetime] = None  # Phase 0: date received
    location: Optional[str] = "default"  # Phase 0: location field (not location_id)
    free_qty_units: Optional[int] = 0  # Phase 0: free quantity in units
    batch_priority: str = "LIFA"  # LIFA = Last In First Available, LILA = Last In Last Available (FIFO)
    notes: Optional[str] = None  # Phase 0: notes field
    purchase_id: Optional[str] = None  # Reference to purchase that created this batch
    created_by: Optional[str] = None
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))
    updated_by: Optional[str] = None
    updated_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

class StockBatchCreate(BaseModel):
    product_sku: str
    batch_no: str
    manufacture_date: Optional[str] = None
    expiry_date: str
    qty_on_hand: int  # In packs
    cost_price_per_unit: float
    mrp_per_unit: float
    supplier_name: Optional[str] = None
    supplier_invoice_no: Optional[str] = None
    received_date: Optional[str] = None
    location: Optional[str] = "default"
    free_qty_units: Optional[int] = 0
    notes: Optional[str] = None

class StockBatchUpdate(BaseModel):
    batch_no: Optional[str] = None
    manufacture_date: Optional[str] = None
    expiry_date: Optional[str] = None
    qty_on_hand: Optional[int] = None
    cost_price_per_unit: Optional[float] = None
    mrp_per_unit: Optional[float] = None
    supplier_name: Optional[str] = None
    supplier_invoice_no: Optional[str] = None
    received_date: Optional[str] = None
    location: Optional[str] = None
    free_qty_units: Optional[int] = None
    notes: Optional[str] = None

# Legacy Medicine Models (for backward compatibility during migration)
class Medicine(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    name: str
    batch_number: str
    expiry_date: datetime
    mrp: float
    quantity: int
    supplier_name: str
    purchase_rate: float
    selling_price: float
    hsn_code: Optional[str] = None
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))
    updated_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

class MedicineCreate(BaseModel):
    name: str
    batch_number: str
    expiry_date: str
    mrp: float
    quantity: int
    supplier_name: str
    purchase_rate: float
    selling_price: float
    hsn_code: Optional[str] = None

class MedicineUpdate(BaseModel):
    name: Optional[str] = None
    batch_number: Optional[str] = None
    expiry_date: Optional[str] = None
    mrp: Optional[float] = None
    quantity: Optional[int] = None
    supplier_name: Optional[str] = None
    purchase_rate: Optional[float] = None
    selling_price: Optional[float] = None
    hsn_code: Optional[str] = None

# Bill Models
class BillItem(BaseModel):
    product_id: Optional[str] = None  # Updated from medicine_id
    batch_id: Optional[str] = None  # New field for batch tracking
    product_name: Optional[str] = None  # Updated from medicine_name
    brand: Optional[str] = None
    batch_no: Optional[str] = None  # Updated from batch_number
    expiry_date: Optional[str] = None
    quantity: int
    unit_price: float  # The actual selling price
    mrp: float
    discount: float = 0
    gst_percent: float = 5  # Updated from gst_rate
    line_total: float  # Updated from total

class Bill(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    bill_number: str
    invoice_type: str = "SALE"  # SALE or SALES_RETURN
    ref_invoice_id: Optional[str] = None  # for returns
    status: str = "paid"  # draft, paid, due, refunded, cancelled
    customer_id: Optional[str] = None
    customer_name: Optional[str] = None
    customer_mobile: Optional[str] = None
    doctor_id: Optional[str] = None
    doctor_name: Optional[str] = None
    items: List[Dict[str, Any]]
    subtotal: float
    discount: float = 0
    tax_rate: float
    tax_amount: float
    total_amount: float
    paid_amount: float = 0  # Total amount paid so far
    due_amount: float = 0  # Remaining amount to be paid
    payment_method: Optional[str] = None  # Deprecated: use payments table instead
    cashier_id: str
    cashier_name: str
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

class BillCreate(BaseModel):
    customer_id: Optional[str] = None
    customer_name: Optional[str] = None
    customer_mobile: Optional[str] = None
    doctor_id: Optional[str] = None
    doctor_name: Optional[str] = None
    items: List[Dict[str, Any]]
    discount: float = 0
    tax_rate: float
    payments: Optional[List[Dict[str, Any]]] = None  # List of payment objects
    payment_method: Optional[str] = None  # Legacy support: single payment method
    status: str = "paid"  # draft or paid
    invoice_type: str = "SALE"
    ref_invoice_id: Optional[str] = None
    refund: Optional[Dict[str, Any]] = None  # For returns with refund

# Bill Number Sequence Model
class BillNumberSequence(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    branch_id: Optional[str] = None  # Future-ready for multi-branch
    prefix: str = "INV"
    current_sequence: int = 0
    sequence_length: int = 6
    allow_prefix_change: bool = True
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))
    updated_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

class BillSequenceSettings(BaseModel):
    prefix: str = "INV"
    starting_number: int = 1
    sequence_length: int = 6
    allow_prefix_change: bool = True

# Stock Movement Models (Enhanced Ledger)
class StockMovement(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    product_sku: str  # Phase 0: FK to product SKU
    batch_id: str
    product_name: str  # Denormalized for display
    batch_no: str  # Denormalized for display
    qty_delta_units: int  # Phase 0: positive for IN, negative for OUT (in units)
    movement_type: str  # Phase 0: 'opening_stock', 'adjustment', 'sale', 'sales_return', 'purchase', 'purchase_return'
    ref_type: str  # Phase 0: 'adjustment', 'opening', 'invoice', 'purchase'
    ref_id: str  # bill_id or purchase_id or adjustment_id
    location: Optional[str] = "default"
    reason: Optional[str] = None
    performed_by: str  # Phase 0: performed_by instead of created_by
    performed_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))  # Phase 0: performed_at

class StockMovementCreate(BaseModel):
    product_sku: str
    batch_id: str
    product_name: str
    batch_no: str
    qty_delta_units: int
    movement_type: str
    ref_type: str
    ref_id: str
    location: Optional[str] = "default"
    reason: Optional[str] = None

# Stock Adjustment Model - Phase 0
class StockAdjustment(BaseModel):
    batch_id: str
    adjustment_type: str  # 'add' or 'remove'
    qty_units: int  # Quantity to add/remove in units
    reason: str
    reference_number: Optional[str] = None
    notes: Optional[str] = None

# Schedule H1 Register Entry Model - Drug Schedule Compliance
class ScheduleH1Entry(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    dispensed_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))
    product_sku: str
    product_name: str
    batch_no: str
    quantity_dispensed: int
    prescriber_name: str
    prescriber_address: str
    prescriber_registration_no: str
    patient_name: str
    patient_address: Optional[str] = None
    bill_id: str
    bill_number: str
    dispensed_by: str
    dispensed_by_name: str

# Payment Models (for multiple payment methods support)
class Payment(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    invoice_id: str
    amount: float
    payment_method: str  # cash, card, upi, credit
    reference_number: Optional[str] = None  # for card/UPI transactions
    notes: Optional[str] = None
    created_by: str
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

class PaymentCreate(BaseModel):
    invoice_id: str
    amount: float
    payment_method: str
    reference_number: Optional[str] = None
    notes: Optional[str] = None

# Refund Models (for sales returns)
class Refund(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    return_invoice_id: str  # ID of the return invoice
    original_invoice_id: Optional[str] = None  # Original sale invoice if linked
    amount: float
    refund_method: str  # cash, card, upi, credit_note
    reference_number: Optional[str] = None
    reason: Optional[str] = None  # damaged, expired, wrong_item, customer_request
    notes: Optional[str] = None
    created_by: str
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

class RefundCreate(BaseModel):
    return_invoice_id: str
    original_invoice_id: Optional[str] = None
    amount: float
    refund_method: str
    reference_number: Optional[str] = None
    reason: Optional[str] = None
    notes: Optional[str] = None

# Audit Log Models (for compliance and tracking)
class AuditLog(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    entity_type: str  # 'invoice', 'payment', 'refund', 'product', 'batch'
    entity_id: str
    action: str  # 'create', 'update', 'delete', 'status_change'
    old_value: Optional[dict] = None
    new_value: Optional[dict] = None
    changes: Optional[dict] = None  # Specific fields that changed
    performed_by: str  # User ID
    performed_by_name: str  # User name for easy display
    reason: Optional[str] = None
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

class AuditLogCreate(BaseModel):
    entity_type: str
    entity_id: str
    action: str
    old_value: Optional[dict] = None
    new_value: Optional[dict] = None
    changes: Optional[dict] = None
    reason: Optional[str] = None

# ==================== SUPPLIER MODELS ====================
class SupplierPayment(BaseModel):
    """Record of a payment made to supplier"""
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    amount: float
    payment_date: str  # ISO date string
    payment_method: str = "cash"  # cash, bank_transfer, cheque, upi
    reference_no: Optional[str] = None  # Cheque/transaction reference
    notes: Optional[str] = None
    purchase_ids: List[str] = []  # Purchases this payment applies to
    created_by: Optional[str] = None
    created_at: str = Field(default_factory=lambda: datetime.now(timezone.utc).isoformat())

class Supplier(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    name: str
    contact_name: Optional[str] = None
    contact_person: Optional[str] = None  # Alias for frontend compatibility
    phone: Optional[str] = None
    email: Optional[str] = None
    gstin: Optional[str] = None
    address: Optional[str] = None
    payment_terms_days: int = 30  # Default 30 days credit
    credit_days: Optional[int] = None  # Alias for frontend compatibility
    notes: Optional[str] = None
    is_active: bool = True  # ADDED: For activate/deactivate functionality
    outstanding: float = 0.0  # Outstanding balance owed to supplier
    payment_history: List[SupplierPayment] = []  # History of payments made
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))
    updated_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

class SupplierCreate(BaseModel):
    name: str
    contact_name: Optional[str] = None
    contact_person: Optional[str] = None  # Alias for frontend
    phone: Optional[str] = None
    email: Optional[str] = None
    gstin: Optional[str] = None
    address: Optional[str] = None
    payment_terms_days: int = 30
    credit_days: Optional[int] = None  # Alias for frontend
    notes: Optional[str] = None

class SupplierUpdate(BaseModel):
    name: Optional[str] = None
    contact_name: Optional[str] = None
    contact_person: Optional[str] = None  # Alias for frontend
    phone: Optional[str] = None
    email: Optional[str] = None
    gstin: Optional[str] = None
    address: Optional[str] = None
    payment_terms_days: Optional[int] = None
    credit_days: Optional[int] = None  # Alias for frontend
    notes: Optional[str] = None
    is_active: Optional[bool] = None  # ADDED: For deactivation

# ==================== PURCHASE MODELS ====================
class PurchaseItem(BaseModel):
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    product_sku: str
    product_name: str  # Denormalized for display
    batch_no: Optional[str] = None  # Optional: can be assigned during receive
    expiry_date: Optional[datetime] = None
    qty_packs: Optional[int] = None  # Input convenience
    qty_units: int  # Canonical quantity in units
    free_qty_units: int = 0  # Free quantity from scheme
    cost_price_per_unit: float  # Base cost
    ptr_per_unit: Optional[float] = None  # PTR = Price to Retailer (what we pay)
    mrp_per_unit: float
    gst_percent: float = 5.0
    batch_priority: str = "LIFA"  # LIFA or LILA
    line_total: float
    received_qty_units: int = 0  # Track how many units received so far

class Purchase(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    purchase_number: str  # e.g., PUR-2024-0001
    supplier_id: str
    supplier_name: str  # Denormalized
    purchase_date: datetime
    due_date: Optional[datetime] = None  # Payment due date
    supplier_invoice_no: Optional[str] = None
    supplier_invoice_date: Optional[datetime] = None
    order_type: str = "direct"  # direct, credit, consignment
    with_gst: bool = True
    purchase_on: str = "credit"  # credit, cash
    status: str = "draft"  # draft, confirmed, received, partially_received, closed, cancelled
    payment_status: str = "unpaid"  # unpaid, partial, paid
    items: List[PurchaseItem] = []
    subtotal: float = 0
    tax_value: float = 0
    round_off: float = 0
    total_value: float = 0
    amount_paid: float = 0  # Amount paid so far
    payment_terms_days: int = 30
    note: Optional[str] = None
    created_by: str
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))
    updated_by: Optional[str] = None
    updated_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

class PurchaseItemCreate(BaseModel):
    product_sku: str
    product_name: str
    batch_no: Optional[str] = None
    expiry_date: Optional[str] = None  # ISO string
    qty_packs: Optional[int] = None
    qty_units: int
    free_qty_units: Optional[int] = 0  # Free quantity from scheme
    cost_price_per_unit: float  # Base cost before scheme
    ptr_per_unit: Optional[float] = None  # PTR = Price to Retailer
    mrp_per_unit: float
    gst_percent: float = 5.0
    batch_priority: str = "LIFA"  # LIFA or LILA

class PurchaseCreate(BaseModel):
    supplier_id: str
    purchase_date: str  # ISO date string
    due_date: Optional[str] = None  # Payment due date
    supplier_invoice_no: Optional[str] = None
    supplier_invoice_date: Optional[str] = None
    order_type: str = "direct"  # direct, credit, consignment
    with_gst: bool = True  # GST inclusive or not
    purchase_on: str = "credit"  # credit, cash
    items: List[PurchaseItemCreate]
    note: Optional[str] = None
    status: Optional[str] = "draft"  # draft or confirmed
    payment_status: str = "unpaid"  # unpaid, partial, paid

class PurchaseUpdate(BaseModel):
    supplier_id: Optional[str] = None
    purchase_date: Optional[str] = None
    supplier_invoice_no: Optional[str] = None
    supplier_invoice_date: Optional[str] = None
    items: Optional[List[PurchaseItemCreate]] = None
    note: Optional[str] = None

# Customer Models
class Customer(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    name: str
    phone: str
    email: Optional[str] = None
    address: Optional[str] = None
    customer_type: str = "regular"  # regular, wholesale, institution
    gstin: Optional[str] = None
    credit_limit: float = 0
    notes: Optional[str] = None
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

# ==================== GOODS RECEIPT MODELS ====================
class GoodsReceiptItem(BaseModel):
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    purchase_item_id: str
    product_sku: str
    product_name: str
    batch_id: str  # Stock batch created/updated
    batch_no: str
    expiry_date: Optional[datetime] = None
    qty_units: int  # Quantity received in this receipt
    cost_price_per_unit: float
    mrp_per_unit: float

class GoodsReceipt(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    purchase_id: str
    purchase_number: str  # Denormalized
    receipt_date: datetime
    received_by: str
    received_by_name: str  # Denormalized
    supplier_invoice_no: Optional[str] = None
    note: Optional[str] = None
    items: List[GoodsReceiptItem] = []
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

class ReceiveItemInput(BaseModel):
    purchase_item_id: str
    batch_no: str
    expiry_date: Optional[str] = None  # ISO string
    qty_units: int  # Quantity to receive
    cost_price_per_unit: Optional[float] = None  # Override if different
    mrp_per_unit: Optional[float] = None  # Override if different

class ReceiveGoodsInput(BaseModel):
    receipt_date: str  # ISO date
    items: List[ReceiveItemInput]
    supplier_invoice_no: Optional[str] = None
    note: Optional[str] = None

# ==================== PURCHASE RETURN MODELS ====================
class PurchaseReturnItem(BaseModel):
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    product_sku: str
    product_name: str
    batch_id: str
    batch_no: str
    qty_units: int  # Quantity being returned
    cost_price_per_unit: float
    reason: str  # Expired, Damaged, Overstock, Other
    line_total: float

class PurchaseReturn(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    return_number: str  # e.g., PRET-2024-0001
    supplier_id: str
    supplier_name: str  # Denormalized
    purchase_id: Optional[str] = None  # Optional link to original purchase
    purchase_number: Optional[str] = None
    return_date: datetime
    status: str = "draft"  # draft, confirmed
    items: List[PurchaseReturnItem] = []
    total_value: float = 0
    note: Optional[str] = None
    created_by: str
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))
    confirmed_at: Optional[datetime] = None
    confirmed_by: Optional[str] = None

class PurchaseReturnItemCreate(BaseModel):
    product_sku: str
    product_name: str
    batch_id: Optional[str] = None
    batch_no: Optional[str] = None
    expiry_date: Optional[str] = None
    expiry: Optional[str] = None  # Alias
    mrp: Optional[float] = None
    ptr: Optional[float] = None
    gst_percent: Optional[float] = 5
    qty_units: Optional[int] = None
    return_qty_units: Optional[int] = None  # Alias for qty_units from frontend
    cost_price_per_unit: Optional[float] = None
    reason: Optional[str] = None

class PurchaseReturnCreate(BaseModel):
    supplier_id: str
    purchase_id: Optional[str] = None
    return_date: str  # ISO date
    items: List[PurchaseReturnItemCreate]
    note: Optional[str] = None
    notes: Optional[str] = None  # Alias
    reason: Optional[str] = None
    billed_by: Optional[str] = None
    payment_type: Optional[str] = "credit"

# ==================== SALES RETURNS MODELS ====================
class SalesReturnItem(BaseModel):
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    medicine_id: Optional[str] = None
    medicine_name: str
    product_sku: Optional[str] = None
    batch_id: Optional[str] = None
    batch_no: str
    expiry_date: Optional[str] = None
    mrp: float
    qty: int  # Quantity being returned
    original_qty: int  # Original billed qty for validation
    disc_percent: float = 0
    disc_price: float = 0  # Price after discount
    gst_percent: float = 5
    amount: float  # Line total
    is_damaged: bool = False  # If true, goes to damaged_stock

class SalesReturn(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    return_no: str  # CN-00001 format
    original_bill_id: Optional[str] = None  # null if manual return
    original_bill_no: Optional[str] = None
    return_date: datetime
    entry_date: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))
    patient: Dict[str, Any] = {}  # { id, name, phone }
    billing_for: str = "self"
    doctor: Optional[str] = None
    created_by: Dict[str, Any] = {}  # { id, name }
    items: List[SalesReturnItem] = []
    mrp_total: float = 0
    total_discount: float = 0
    gst_amount: float = 0
    round_off: float = 0
    net_amount: float = 0
    payment_type: Optional[str] = None  # Original payment type
    refund_method: str = "same_as_original"  # cash, upi, credit_to_account, same_as_original
    note: Optional[str] = None
    status: str = "completed"  # completed, cancelled
    credit_note_ref: Optional[str] = None  # Same as return_no for now
    returns: List[str] = []  # For future partial re-returns

class SalesReturnItemCreate(BaseModel):
    medicine_id: Optional[str] = None
    medicine_name: str
    product_sku: Optional[str] = None
    batch_id: Optional[str] = None
    batch_no: str
    expiry_date: Optional[str] = None
    mrp: float
    qty: int
    original_qty: int
    disc_percent: float = 0
    disc_price: Optional[float] = None
    gst_percent: float = 5
    amount: Optional[float] = None
    is_damaged: bool = False

class SalesReturnCreate(BaseModel):
    original_bill_id: Optional[str] = None
    original_bill_no: Optional[str] = None
    return_date: str  # ISO date string
    patient: Optional[Dict[str, Any]] = None
    billing_for: str = "self"
    doctor: Optional[str] = None
    items: List[SalesReturnItemCreate]
    payment_type: Optional[str] = None
    refund_method: str = "same_as_original"
    note: Optional[str] = None

class SalesReturnUpdate(BaseModel):
    billing_for: Optional[str] = None
    doctor: Optional[str] = None
    billed_by: Optional[str] = None
    note: Optional[str] = None
    # Financial edit fields
    items: Optional[List[SalesReturnItemCreate]] = None
    refund_method: Optional[str] = None

# ==================== SUPPLIER CREDIT MODELS ====================
class SupplierCredit(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    supplier_id: str
    supplier_name: str  # Denormalized
    credit_number: str  # e.g., SCRED-2024-0001
    amount: float
    reference: str  # Reference to purchase_return_id or other source
    reference_type: str  # purchase_return, adjustment
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))
    created_by: str


    address: Optional[str] = None
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

class CustomerCreate(BaseModel):
    name: str
    phone: str
    email: Optional[str] = None
    address: Optional[str] = None
    customer_type: str = "regular"
    gstin: Optional[str] = None
    credit_limit: float = 0
    notes: Optional[str] = None

# Doctor Models
class Doctor(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    name: str
    contact: Optional[str] = None
    specialization: Optional[str] = None
    clinic_address: Optional[str] = None
    notes: Optional[str] = None
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

class DoctorCreate(BaseModel):
    name: str
    contact: Optional[str] = None
    specialization: Optional[str] = None
    clinic_address: Optional[str] = None
    notes: Optional[str] = None

# ==================== AUTH HELPERS ====================

def hash_password(password: str) -> str:
    return pwd_context.hash(password)

def verify_password(plain_password: str, hashed_password: str) -> bool:
    return pwd_context.verify(plain_password, hashed_password)

def create_access_token(data: dict):
    to_encode = data.copy()
    expire = datetime.now(timezone.utc) + timedelta(minutes=ACCESS_TOKEN_EXPIRE_MINUTES)
    to_encode.update({"exp": expire})
    encoded_jwt = jwt.encode(to_encode, SECRET_KEY, algorithm=ALGORITHM)
    return encoded_jwt

async def get_current_user(request: Request, session_token: Optional[str] = Cookie(None)):
    # Check cookie first
    token = session_token
    
    # Fallback to Authorization header
    if not token:
        auth_header = request.headers.get("Authorization")
        if auth_header and auth_header.startswith("Bearer "):
            token = auth_header.split(" ")[1]
    
    if not token:
        raise HTTPException(status_code=401, detail="Not authenticated")
    
    try:
        # Check if it's a JWT token
        payload = jwt.decode(token, SECRET_KEY, algorithms=[ALGORITHM])
        user_id = payload.get("sub")
        if user_id is None:
            raise HTTPException(status_code=401, detail="Invalid token")
        
        user = await db.users.find_one({"id": user_id}, {"_id": 0})
        if not user:
            raise HTTPException(status_code=401, detail="User not found")
        return User(**user)
    except jwt.ExpiredSignatureError:
        # Check if it's an Emergent session token
        session = await db.sessions.find_one({"session_token": token}, {"_id": 0})
        if not session:
            raise HTTPException(status_code=401, detail="Token expired")
        
        # Check if session expired
        expires_at = session['expires_at']
        if isinstance(expires_at, str):
            expires_at = datetime.fromisoformat(expires_at)
        
        if datetime.now(timezone.utc) > expires_at:
            raise HTTPException(status_code=401, detail="Session expired")
        
        user = await db.users.find_one({"id": session['user_id']}, {"_id": 0})
        if not user:
            raise HTTPException(status_code=401, detail="User not found")
        return User(**user)
    except jwt.InvalidTokenError:
        # Check if it's an Emergent session token
        session = await db.sessions.find_one({"session_token": token}, {"_id": 0})
        if not session:
            raise HTTPException(status_code=401, detail="Invalid token")
        
        expires_at = session['expires_at']
        if isinstance(expires_at, str):
            expires_at = datetime.fromisoformat(expires_at)
        
        if datetime.now(timezone.utc) > expires_at:
            raise HTTPException(status_code=401, detail="Session expired")
        
        user = await db.users.find_one({"id": session['user_id']}, {"_id": 0})
        if not user:
            raise HTTPException(status_code=401, detail="User not found")
        return User(**user)

# ==================== API UTILITY HELPERS ====================

def parse_fields_param(fields: Optional[str]) -> Optional[dict]:
    """
    Parse comma-separated fields into MongoDB projection.
    Example: "name,sku,price" -> {"_id": 0, "name": 1, "sku": 1, "price": 1}
    """
    if not fields:
        return {"_id": 0}  # Default: exclude _id only
    
    field_list = [f.strip() for f in fields.split(",") if f.strip()]
    if not field_list:
        return {"_id": 0}
    
    projection = {"_id": 0}
    for field in field_list:
        projection[field] = 1
    return projection

def paginate_response(items: list, page: int, page_size: int, total: int) -> dict:
    """
    Create a standardized paginated response.
    """
    return {
        "data": items,
        "pagination": {
            "page": page,
            "page_size": page_size,
            "total_items": total,
            "total_pages": (total + page_size - 1) // page_size,
            "has_next": page * page_size < total,
            "has_prev": page > 1
        }
    }

# ==================== AUTH ROUTES ====================

@api_router.post("/auth/register")
async def register(user_data: UserCreate):
    # Check if user exists
    existing_user = await db.users.find_one({"email": user_data.email})
    if existing_user:
        raise HTTPException(status_code=400, detail="Email already registered")
    
    # Create user
    user = User(
        email=user_data.email,
        name=user_data.name,
        role=user_data.role,
        password_hash=hash_password(user_data.password)
    )
    
    doc = user.model_dump()
    doc['created_at'] = doc['created_at'].isoformat()
    await db.users.insert_one(doc)
    
    # Create token
    token = create_access_token({"sub": user.id, "email": user.email})
    
    return {
        "token": token,
        "user": {
            "id": user.id,
            "email": user.email,
            "name": user.name,
            "role": user.role
        }
    }

@api_router.post("/auth/login")
async def login(credentials: UserLogin):
    user_doc = await db.users.find_one({"email": credentials.email}, {"_id": 0})
    if not user_doc:
        raise HTTPException(status_code=401, detail="Invalid credentials")
    
    if not verify_password(credentials.password, user_doc['password_hash']):
        raise HTTPException(status_code=401, detail="Invalid credentials")
    
    token = create_access_token({"sub": user_doc['id'], "email": user_doc['email']})
    
    return {
        "token": token,
        "user": {
            "id": user_doc['id'],
            "email": user_doc['email'],
            "name": user_doc['name'],
            "role": user_doc['role']
        }
    }

@api_router.post("/auth/session")
async def create_session(request: Request, response: Response):
    session_id = request.headers.get("X-Session-ID")
    if not session_id:
        raise HTTPException(status_code=400, detail="Session ID required")
    
    # Call Emergent auth service
    async with httpx.AsyncClient() as client:
        try:
            auth_response = await client.get(
                "https://demobackend.emergentagent.com/auth/v1/env/oauth/session-data",
                headers={"X-Session-ID": session_id}
            )
            auth_response.raise_for_status()
            session_data = auth_response.json()
        except Exception as e:
            raise HTTPException(status_code=400, detail=f"Failed to validate session: {str(e)}")
    
    # Check if user exists
    user_doc = await db.users.find_one({"email": session_data['email']}, {"_id": 0})
    
    if not user_doc:
        # Create new user with admin role for first user
        user_count = await db.users.count_documents({})
        role = "admin" if user_count == 0 else "cashier"
        
        user = User(
            email=session_data['email'],
            name=session_data['name'],
            role=role
        )
        doc = user.model_dump()
        doc['created_at'] = doc['created_at'].isoformat()
        await db.users.insert_one(doc)
        user_id = user.id
    else:
        user_id = user_doc['id']
        user = User(**user_doc)
    
    # Store session
    expires_at = datetime.now(timezone.utc) + timedelta(days=7)
    session = SessionCreate(
        user_id=user_id,
        session_token=session_data['session_token'],
        email=session_data['email'],
        name=session_data['name'],
        expires_at=expires_at
    )
    
    session_doc = session.model_dump()
    session_doc['expires_at'] = session_doc['expires_at'].isoformat()
    await db.sessions.insert_one(session_doc)
    
    # Set cookie
    response.set_cookie(
        key="session_token",
        value=session_data['session_token'],
        httponly=True,
        secure=True,
        samesite="none",
        path="/",
        max_age=60 * 60 * 24 * 7
    )
    
    return {
        "user": {
            "id": user.id,
            "email": user.email,
            "name": user.name,
            "role": user.role
        }
    }

@api_router.post("/auth/logout")
async def logout(response: Response, current_user: User = Depends(get_current_user)):
    # Delete session from database
    await db.sessions.delete_many({"user_id": current_user.id})
    
    # Clear cookie
    response.delete_cookie(key="session_token", path="/")
    
    return {"message": "Logged out successfully"}

@api_router.get("/auth/me")
async def get_me(current_user: User = Depends(get_current_user)):
    return {
        "id": current_user.id,
        "email": current_user.email,
        "name": current_user.name,
        "role": current_user.role,
        "is_active": current_user.is_active
    }

# ==================== USER MANAGEMENT ROUTES ====================

@api_router.get("/users", response_model=List[User])
async def get_all_users(current_user: User = Depends(get_current_user)):
    """Get all users - Admin only"""
    if current_user.role != "admin":
        raise HTTPException(status_code=403, detail="Admin access required")
    
    users = await db.users.find({}, {"_id": 0, "password_hash": 0}).to_list(1000)
    return users

@api_router.post("/users", response_model=User)
async def create_user(user_data: UserCreate, current_user: User = Depends(get_current_user)):
    """Create new user - Admin only"""
    if current_user.role != "admin":
        raise HTTPException(status_code=403, detail="Admin access required")
    
    # Validate role
    if user_data.role not in ["admin", "manager", "cashier", "inventory_staff"]:
        raise HTTPException(status_code=400, detail="Invalid role")
    
    # Check if email already exists
    existing_user = await db.users.find_one({"email": user_data.email})
    if existing_user:
        raise HTTPException(status_code=400, detail="Email already registered")
    
    # Hash password
    password_hash = pwd_context.hash(user_data.password)
    
    # Create user
    user = User(
        email=user_data.email,
        name=user_data.name,
        role=user_data.role,
        password_hash=password_hash,
        created_by=current_user.id,
        is_active=True
    )
    
    doc = user.model_dump()
    await db.users.insert_one(doc)
    
    # Return without password_hash
    user.password_hash = None
    return user

@api_router.get("/users/{user_id}", response_model=User)
async def get_user(user_id: str, current_user: User = Depends(get_current_user)):
    """Get user by ID - Admin only"""
    if current_user.role != "admin":
        raise HTTPException(status_code=403, detail="Admin access required")
    
    user = await db.users.find_one({"id": user_id}, {"_id": 0, "password_hash": 0})
    if not user:
        raise HTTPException(status_code=404, detail="User not found")
    
    return user

@api_router.put("/users/{user_id}")
async def update_user(
    user_id: str,
    user_update: UserUpdate,
    current_user: User = Depends(get_current_user)
):
    """Update user - Admin only"""
    if current_user.role != "admin":
        raise HTTPException(status_code=403, detail="Admin access required")
    
    # Check if user exists
    existing_user = await db.users.find_one({"id": user_id})
    if not existing_user:
        raise HTTPException(status_code=404, detail="User not found")
    
    # Validate role if provided
    if user_update.role and user_update.role not in ["admin", "manager", "cashier", "inventory_staff"]:
        raise HTTPException(status_code=400, detail="Invalid role")
    
    # Check if email is being changed and if it's already taken
    if user_update.email and user_update.email != existing_user.get("email"):
        email_exists = await db.users.find_one({"email": user_update.email})
        if email_exists:
            raise HTTPException(status_code=400, detail="Email already in use")
    
    # Prepare update data
    update_data = {k: v for k, v in user_update.model_dump().items() if v is not None}
    update_data["updated_at"] = datetime.now(timezone.utc)
    update_data["updated_by"] = current_user.id
    
    # Update user
    await db.users.update_one({"id": user_id}, {"$set": update_data})
    
    # Get updated user
    updated_user = await db.users.find_one({"id": user_id}, {"_id": 0, "password_hash": 0})
    return updated_user

@api_router.delete("/users/{user_id}")
async def deactivate_user(user_id: str, current_user: User = Depends(get_current_user)):
    """Deactivate user - Admin only"""
    if current_user.role != "admin":
        raise HTTPException(status_code=403, detail="Admin access required")
    
    # Cannot deactivate self
    if user_id == current_user.id:
        raise HTTPException(status_code=400, detail="Cannot deactivate yourself")
    
    # Check if user exists
    existing_user = await db.users.find_one({"id": user_id})
    if not existing_user:
        raise HTTPException(status_code=404, detail="User not found")
    
    # Deactivate user
    await db.users.update_one(
        {"id": user_id},
        {"$set": {
            "is_active": False,
            "updated_at": datetime.now(timezone.utc),
            "updated_by": current_user.id
        }}
    )
    
    # Delete all sessions for this user
    await db.sessions.delete_many({"user_id": user_id})
    
    return {"message": "User deactivated successfully"}

@api_router.put("/users/me/change-password")
async def change_password(
    password_data: ChangePassword,
    current_user: User = Depends(get_current_user)
):
    """Change own password - All users"""
    # Get user with password hash
    user = await db.users.find_one({"id": current_user.id})
    if not user:
        raise HTTPException(status_code=404, detail="User not found")
    
    # Verify current password
    if not pwd_context.verify(password_data.current_password, user["password_hash"]):
        raise HTTPException(status_code=400, detail="Current password is incorrect")
    
    # Hash new password
    new_password_hash = pwd_context.hash(password_data.new_password)
    
    # Update password
    await db.users.update_one(
        {"id": current_user.id},
        {"$set": {
            "password_hash": new_password_hash,
            "updated_at": datetime.now(timezone.utc)
        }}
    )
    
    return {"message": "Password changed successfully"}

# ==================== SETTINGS ROUTES ====================

@api_router.get("/settings")
async def get_settings(current_user: User = Depends(get_current_user)):
    """Get application settings"""
    settings = await db.settings.find_one({}, {"_id": 0})
    if not settings:
        # Return default settings
        return {
            "inventory": {
                "near_expiry_days": 30,
                "block_expired_stock": True,
                "allow_near_expiry_sale": True,
                "low_stock_alert_enabled": True
            },
            "billing": {
                "enable_draft_bills": True,
                "auto_print_invoice": False
            },
            "returns": {
                "return_window_days": 7,
                "require_original_bill": False,
                "allow_partial_return": True
            },
            "general": {
                "pharmacy_name": "PharmaCare",
                "currency": "INR",
                "timezone": "Asia/Kolkata"
            }
        }
    return settings

@api_router.put("/settings")
async def update_settings(
    settings_data: dict,
    current_user: User = Depends(get_current_user)
):
    """Update application settings - Admin only"""
    if current_user.role != "admin":
        raise HTTPException(status_code=403, detail="Admin access required")
    
    # Check if settings exist
    existing = await db.settings.find_one({})
    
    if existing:
        # Update existing
        await db.settings.update_one(
            {},
            {
                "$set": {
                    **settings_data,
                    "updated_at": datetime.now(timezone.utc).isoformat(),
                    "updated_by": current_user.id
                }
            }
        )
    else:
        # Create new
        settings_doc = {
            "id": str(uuid.uuid4()),
            **settings_data,
            "created_at": datetime.now(timezone.utc).isoformat(),
            "created_by": current_user.id
        }
        await db.settings.insert_one(settings_doc)
    
    return {"message": "Settings updated successfully"}

# ==================== ROLES & PERMISSIONS ROUTES ====================

@api_router.get("/permissions")
async def get_all_permissions(current_user: User = Depends(get_current_user)):
    """Get all available permissions - Admin only"""
    if current_user.role != "admin":
        raise HTTPException(status_code=403, detail="Admin access required")
    
    return ALL_PERMISSIONS

@api_router.get("/roles", response_model=List[Role])
async def get_all_roles(current_user: User = Depends(get_current_user)):
    """Get all roles - Admin only"""
    if current_user.role != "admin":
        raise HTTPException(status_code=403, detail="Admin access required")
    
    roles = await db.roles.find({}, {"_id": 0}).to_list(1000)
    return roles

@api_router.post("/roles", response_model=Role)
async def create_role(role_data: RoleCreate, current_user: User = Depends(get_current_user)):
    """Create new custom role - Admin only"""
    if current_user.role != "admin":
        raise HTTPException(status_code=403, detail="Admin access required")
    
    # Check if role name already exists
    existing_role = await db.roles.find_one({"name": role_data.name})
    if existing_role:
        raise HTTPException(status_code=400, detail="Role name already exists")
    
    # Create role
    role = Role(
        name=role_data.name,
        display_name=role_data.display_name,
        permissions=role_data.permissions,
        is_default=False,
        is_super_admin=False,
        created_by=current_user.id
    )
    
    doc = role.model_dump()
    await db.roles.insert_one(doc)
    
    return role

@api_router.get("/roles/{role_id}", response_model=Role)
async def get_role(role_id: str, current_user: User = Depends(get_current_user)):
    """Get role by ID - Admin only"""
    if current_user.role != "admin":
        raise HTTPException(status_code=403, detail="Admin access required")
    
    role = await db.roles.find_one({"id": role_id}, {"_id": 0})
    if not role:
        raise HTTPException(status_code=404, detail="Role not found")
    
    return role

@api_router.put("/roles/{role_id}")
async def update_role(
    role_id: str,
    role_update: RoleUpdate,
    current_user: User = Depends(get_current_user)
):
    """Update role - Admin only"""
    if current_user.role != "admin":
        raise HTTPException(status_code=403, detail="Admin access required")
    
    # Check if role exists
    existing_role = await db.roles.find_one({"id": role_id})
    if not existing_role:
        raise HTTPException(status_code=404, detail="Role not found")
    
    # Cannot edit default roles
    if existing_role.get("is_default", False):
        raise HTTPException(status_code=400, detail="Cannot edit default roles")
    
    # Prepare update data
    update_data = {k: v for k, v in role_update.model_dump().items() if v is not None}
    update_data["updated_at"] = datetime.now(timezone.utc)
    
    # Update role
    await db.roles.update_one({"id": role_id}, {"$set": update_data})
    
    # Get updated role
    updated_role = await db.roles.find_one({"id": role_id}, {"_id": 0})
    return updated_role

@api_router.delete("/roles/{role_id}")
async def delete_role(role_id: str, current_user: User = Depends(get_current_user)):
    """Delete custom role - Admin only"""
    if current_user.role != "admin":
        raise HTTPException(status_code=403, detail="Admin access required")
    
    # Check if role exists
    existing_role = await db.roles.find_one({"id": role_id})
    if not existing_role:
        raise HTTPException(status_code=404, detail="Role not found")
    
    # Cannot delete default roles
    if existing_role.get("is_default", False):
        raise HTTPException(status_code=400, detail="Cannot delete default roles")
    
    # Check if any users have this role
    users_with_role = await db.users.count_documents({"role": existing_role["name"]})
    if users_with_role > 0:
        raise HTTPException(
            status_code=400, 
            detail=f"Cannot delete role. {users_with_role} user(s) are assigned this role"
        )
    
    # Delete role
    await db.roles.delete_one({"id": role_id})
    
    return {"message": "Role deleted successfully"}

# ==================== MEDICINE ROUTES ====================

@api_router.post("/medicines", response_model=Medicine)
async def create_medicine(medicine_data: MedicineCreate, current_user: User = Depends(get_current_user)):
    data = medicine_data.model_dump()
    data['expiry_date'] = datetime.fromisoformat(medicine_data.expiry_date)
    medicine = Medicine(**data)
    
    doc = medicine.model_dump()
    doc['expiry_date'] = doc['expiry_date'].isoformat()
    doc['created_at'] = doc['created_at'].isoformat()
    doc['updated_at'] = doc['updated_at'].isoformat()
    
    await db.medicines.insert_one(doc)
    return medicine

@api_router.get("/medicines", response_model=List[Medicine])
async def get_medicines(current_user: User = Depends(get_current_user)):
    medicines = await db.medicines.find({}, {"_id": 0}).to_list(10000)
    for med in medicines:
        if isinstance(med['expiry_date'], str):
            med['expiry_date'] = datetime.fromisoformat(med['expiry_date'])
        if isinstance(med['created_at'], str):
            med['created_at'] = datetime.fromisoformat(med['created_at'])
        if isinstance(med['updated_at'], str):
            med['updated_at'] = datetime.fromisoformat(med['updated_at'])
    return medicines

@api_router.get("/medicines/search")
async def search_medicines(q: str, current_user: User = Depends(get_current_user)):
    medicines = await db.medicines.find(
        {"$or": [
            {"name": {"$regex": q, "$options": "i"}},
            {"batch_number": {"$regex": q, "$options": "i"}}
        ]},
        {"_id": 0}
    ).to_list(100)
    
    for med in medicines:
        if isinstance(med['expiry_date'], str):
            med['expiry_date'] = datetime.fromisoformat(med['expiry_date'])
        if isinstance(med['created_at'], str):
            med['created_at'] = datetime.fromisoformat(med['created_at'])
        if isinstance(med['updated_at'], str):
            med['updated_at'] = datetime.fromisoformat(med['updated_at'])
    
    return medicines

@api_router.get("/medicines/{medicine_id}", response_model=Medicine)
async def get_medicine(medicine_id: str, current_user: User = Depends(get_current_user)):
    medicine = await db.medicines.find_one({"id": medicine_id}, {"_id": 0})
    if not medicine:
        raise HTTPException(status_code=404, detail="Medicine not found")
    
    if isinstance(medicine['expiry_date'], str):
        medicine['expiry_date'] = datetime.fromisoformat(medicine['expiry_date'])
    if isinstance(medicine['created_at'], str):
        medicine['created_at'] = datetime.fromisoformat(medicine['created_at'])
    if isinstance(medicine['updated_at'], str):
        medicine['updated_at'] = datetime.fromisoformat(medicine['updated_at'])
    
    return Medicine(**medicine)

@api_router.put("/medicines/{medicine_id}")
async def update_medicine(
    medicine_id: str,
    medicine_data: MedicineUpdate,
    current_user: User = Depends(get_current_user)
):
    update_data = {k: v for k, v in medicine_data.model_dump().items() if v is not None}
    
    if 'expiry_date' in update_data:
        update_data['expiry_date'] = datetime.fromisoformat(update_data['expiry_date']).isoformat()
    
    update_data['updated_at'] = datetime.now(timezone.utc).isoformat()
    
    result = await db.medicines.update_one(
        {"id": medicine_id},
        {"$set": update_data}
    )
    
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Medicine not found")
    
    return {"message": "Medicine updated successfully"}

@api_router.delete("/medicines/{medicine_id}")
async def delete_medicine(medicine_id: str, current_user: User = Depends(get_current_user)):
    if current_user.role != "admin":
        raise HTTPException(status_code=403, detail="Only admins can delete medicines")
    
    result = await db.medicines.delete_one({"id": medicine_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Medicine not found")
    
    return {"message": "Medicine deleted successfully"}

@api_router.get("/medicines/alerts/low-stock")
async def get_low_stock_alerts(current_user: User = Depends(get_current_user)):
    medicines = await db.medicines.find({"quantity": {"$lt": 10}}, {"_id": 0}).to_list(1000)
    for med in medicines:
        if isinstance(med['expiry_date'], str):
            med['expiry_date'] = datetime.fromisoformat(med['expiry_date'])
    return medicines

@api_router.get("/medicines/alerts/expiring-soon")
async def get_expiring_medicines(current_user: User = Depends(get_current_user)):
    try:
        thirty_days_later = datetime.now(timezone.utc) + timedelta(days=30)
        medicines = await db.medicines.find({}, {"_id": 0}).to_list(10000)
        
        expiring_medicines = []
        for med in medicines:
            try:
                expiry = med.get('expiry_date')
                if not expiry:
                    continue
                    
                if isinstance(expiry, str):
                    expiry = datetime.fromisoformat(expiry)
                
                if expiry.tzinfo is None:
                    expiry = expiry.replace(tzinfo=timezone.utc)
                
                if expiry <= thirty_days_later:
                    med['expiry_date'] = expiry
                    expiring_medicines.append(med)
            except Exception as e:
                logger.warning(f"Error processing medicine expiry: {e}")
                continue
        
        return expiring_medicines
    except Exception as e:
        logger.error(f"Expiring medicines error: {e}")
        return []


# ==================== AUDIT LOG HELPER ====================

async def create_audit_log(
    entity_type: str,
    entity_id: str,
    action: str,
    user: User,
    old_value: Optional[dict] = None,
    new_value: Optional[dict] = None,
    reason: Optional[str] = None
):
    """Helper function to create audit log entries"""
    
    # Calculate changes if both old and new values provided
    changes = None
    if old_value and new_value:
        changes = {}
        for key in new_value:
            if key not in old_value or old_value[key] != new_value[key]:
                changes[key] = {
                    'old': old_value.get(key),
                    'new': new_value[key]
                }
    
    audit_log = AuditLog(
        entity_type=entity_type,
        entity_id=entity_id,
        action=action,
        old_value=old_value,
        new_value=new_value,
        changes=changes,
        performed_by=user.id,
        performed_by_name=user.name,
        reason=reason
    )
    
    audit_doc = audit_log.model_dump()
    audit_doc['created_at'] = audit_doc['created_at'].isoformat()
    
    try:
        await db.audit_logs.insert_one(audit_doc)
    except Exception as e:
        logger.error(f"Failed to create audit log: {e}")
        # Don't fail the main operation if audit log fails
    
    return audit_log



# ==================== PRODUCT ROUTES ====================

@api_router.post("/products", response_model=Product)
async def create_product(product_data: ProductCreate, current_user: User = Depends(get_current_user)):
    # Check if SKU already exists
    existing = await db.products.find_one({"sku": product_data.sku}, {"_id": 0})
    if existing:
        raise HTTPException(status_code=400, detail="Product with this SKU already exists")
    
    product = Product(**product_data.model_dump(), created_by=current_user.id, updated_by=current_user.id)
    doc = product.model_dump()
    doc['created_at'] = doc['created_at'].isoformat()
    doc['updated_at'] = doc['updated_at'].isoformat()
    await db.products.insert_one(doc)
    
    return product

@api_router.get("/products")
async def get_products(
    search: Optional[str] = None,
    category: Optional[str] = None,
    fields: Optional[str] = None,
    page: int = 1,
    page_size: int = 100,
    current_user: User = Depends(get_current_user)
):
    """
    Get products with optional pagination and field selection.
    - fields: comma-separated list of fields to return (e.g., "name,sku,default_mrp")
    """
    query = {}
    if search:
        query["$or"] = [
            {"name": {"$regex": search, "$options": "i"}},
            {"sku": {"$regex": search, "$options": "i"}},
            {"brand": {"$regex": search, "$options": "i"}},
            {"manufacturer": {"$regex": search, "$options": "i"}}
        ]
    if category:
        query["category"] = category
    
    projection = parse_fields_param(fields)
    
    # Get total count for pagination
    total = await db.products.count_documents(query)
    
    skip = (page - 1) * page_size
    products = await db.products.find(query, projection).sort("name", 1).skip(skip).limit(page_size).to_list(page_size)
    for prod in products:
        if 'created_at' in prod and isinstance(prod['created_at'], str):
            prod['created_at'] = datetime.fromisoformat(prod['created_at'])
        if 'updated_at' in prod and isinstance(prod['updated_at'], str):
            prod['updated_at'] = datetime.fromisoformat(prod['updated_at'])
        
        # Backward compatibility: normalize old field names to new
        if 'default_mrp' in prod and 'default_mrp_per_unit' not in prod:
            prod['default_mrp_per_unit'] = prod['default_mrp']
        if 'low_stock_threshold' in prod and 'low_stock_threshold_units' not in prod:
            prod['low_stock_threshold_units'] = prod['low_stock_threshold']
        
        # Ensure required fields have defaults
        if 'default_mrp_per_unit' not in prod:
            prod['default_mrp_per_unit'] = prod.get('default_mrp', 0)
        if 'low_stock_threshold_units' not in prod:
            prod['low_stock_threshold_units'] = prod.get('low_stock_threshold', 10)
    
    # Return paginated response if pagination requested
    if page > 1 or page_size != 100:
        return paginate_response(products, page, page_size, total)
            
    return products

@api_router.get("/inventory")
async def get_inventory_with_health(
    page: int = 1,
    page_size: int = 20,
    search: Optional[str] = None,
    status_filter: Optional[str] = None,
    category_filter: Optional[str] = None,
    brand_filter: Optional[str] = None,
    current_user: User = Depends(get_current_user)
):
    """
    Get inventory with severity-based sorting and pagination
    Priority: 1) Expired/Out-of-stock 2) Near-expiry/Low-stock 3) Healthy
    Filters: status (out_of_stock, expired, near_expiry, low_stock, healthy), category, brand
    """
    # Get settings
    settings = await db.settings.find_one({}) or {}
    near_expiry_days = settings.get('near_expiry_days', 30)
    
    # Build query
    query = {}
    if search:
        query["$or"] = [
            {"name": {"$regex": search, "$options": "i"}},
            {"sku": {"$regex": search, "$options": "i"}},
            {"brand": {"$regex": search, "$options": "i"}}
        ]
    
    if category_filter:
        query["category"] = category_filter
    
    if brand_filter:
        query["brand"] = brand_filter
    
    # Get all products (for severity calculation)
    products = await db.products.find(query, {"_id": 0}).to_list(10000)
    
    # Calculate inventory health for each product
    inventory_items = []
    today = datetime.now(timezone.utc)
    near_expiry_threshold = today + timedelta(days=near_expiry_days)
    
    for product in products:
        # Get all batches for this product
        batches = await db.stock_batches.find(
            {"product_sku": product['sku']}, 
            {"_id": 0}
        ).to_list(1000)
        
        if not batches:
            # Out of stock
            inventory_items.append({
                'product': product,
                'total_qty_units': 0,
                'total_qty_packs': 0,
                'nearest_expiry': None,
                'severity': 1,  # Critical (out of stock)
                'status': 'out_of_stock',
                'batches_count': 0
            })
            continue
        
        # Calculate totals and find nearest expiry
        total_units = 0
        total_packs = 0
        nearest_expiry = None
        has_expired = False
        has_near_expiry = False
        
        for batch in batches:
            qty_packs = batch.get('qty_on_hand', 0)
            units_per_pack = product.get('units_per_pack', 1)
            qty_units = int(qty_packs * units_per_pack)
            
            total_packs += qty_packs
            total_units += qty_units
            
            # Check expiry
            if batch.get('expiry_date'):
                expiry_str = batch['expiry_date']
                try:
                    if isinstance(expiry_str, str):
                        expiry_date = datetime.fromisoformat(expiry_str.replace('Z', '+00:00'))
                    else:
                        expiry_date = expiry_str
                    
                    if expiry_date < today:
                        has_expired = True
                    elif expiry_date < near_expiry_threshold:
                        has_near_expiry = True
                    
                    if nearest_expiry is None or expiry_date < nearest_expiry:
                        nearest_expiry = expiry_date
                except:
                    pass
        
        # Determine severity and status
        low_stock_threshold = product.get('low_stock_threshold_units', 10)
        
        if total_units == 0:
            severity = 1
            status = 'out_of_stock'
        elif has_expired:
            severity = 1
            status = 'expired'
        elif has_near_expiry:
            severity = 2
            status = 'near_expiry'
        elif total_units <= low_stock_threshold:
            severity = 2
            status = 'low_stock'
        else:
            severity = 3
            status = 'healthy'
        
        inventory_items.append({
            'product': product,
            'total_qty_units': total_units,
            'total_qty_packs': round(total_packs, 2),
            'nearest_expiry': nearest_expiry.isoformat() if nearest_expiry else None,
            'severity': severity,
            'status': status,
            'batches_count': len(batches)
        })
    
    # Apply status filter if provided
    if status_filter:
        inventory_items = [item for item in inventory_items if item['status'] == status_filter]
    
    # Sort by severity (1=critical, 2=warning, 3=healthy), then by expiry, then alphabetically
    inventory_items.sort(key=lambda x: (
        x['severity'],
        x['nearest_expiry'] if x['nearest_expiry'] else '9999-12-31',
        x['product']['name'].lower()
    ))
    
    # Pagination
    total_items = len(inventory_items)
    total_pages = (total_items + page_size - 1) // page_size
    start_idx = (page - 1) * page_size
    end_idx = start_idx + page_size
    page_items = inventory_items[start_idx:end_idx]
    
    return {
        'items': page_items,
        'pagination': {
            'current_page': page,
            'page_size': page_size,
            'total_items': total_items,
            'total_pages': total_pages,
            'has_next': page < total_pages,
            'has_prev': page > 1
        },
        'summary': {
            'critical_count': sum(1 for item in inventory_items if item['severity'] == 1),
            'warning_count': sum(1 for item in inventory_items if item['severity'] == 2),
            'healthy_count': sum(1 for item in inventory_items if item['severity'] == 3)
        }
    }

@api_router.get("/inventory/filters")
async def get_inventory_filters(current_user: User = Depends(get_current_user)):
    """Get unique categories and brands for filtering"""
    categories = await db.products.distinct("category")
    brands = await db.products.distinct("brand")
    
    # Filter out None/empty values
    categories = [c for c in categories if c]
    brands = [b for b in brands if b]
    
    return {
        "categories": sorted(categories),
        "brands": sorted(brands),
        "statuses": [
            {"value": "out_of_stock", "label": "Out of Stock"},
            {"value": "expired", "label": "Expired"},
            {"value": "near_expiry", "label": "Near Expiry"},
            {"value": "low_stock", "label": "Low Stock"},
            {"value": "healthy", "label": "Healthy"}
        ]
    }

@api_router.get("/products/barcode/{barcode}")
async def lookup_by_barcode(
    barcode: str,
    location_id: Optional[str] = "default",
    current_user: User = Depends(get_current_user)
):
    """
    Fast barcode lookup endpoint for USB/camera scanners.
    Returns product with available batches if found.
    """
    # Search by barcode OR SKU (some pharmacies use SKU as barcode)
    product = await db.products.find_one(
        {"$or": [
            {"barcode": barcode},
            {"sku": barcode}
        ]},
        {"_id": 0}
    )
    
    if not product:
        return {"found": False, "message": f"No product found with barcode: {barcode}"}
    
    # Get batches for this product (FEFO)
    batches = await db.stock_batches.find(
        {
            "product_sku": product['sku'],
            "location": location_id,
            "qty_on_hand": {"$gt": 0}
        },
        {"_id": 0}
    ).sort("expiry_date", 1).to_list(10)
    
    if not batches:
        return {
            "found": True,
            "product": product,
            "has_stock": False,
            "message": "Product found but no stock available"
        }
    
    # Format batches
    formatted_batches = []
    total_qty = 0
    units_per_pack = product.get('units_per_pack', 1)
    
    for batch in batches:
        expiry = batch.get('expiry_date')
        expiry_display = 'N/A'
        expiry_iso = None
        
        if expiry:
            if isinstance(expiry, str):
                expiry = datetime.fromisoformat(expiry)
            expiry_display = expiry.strftime('%d-%m-%Y')
            expiry_iso = expiry.isoformat()
        
        total_units_in_batch = batch['qty_on_hand'] * units_per_pack
        
        formatted_batches.append({
            "batch_id": batch['id'],
            "batch_no": batch['batch_no'],
            "expiry_date": expiry_display,
            "expiry_iso": expiry_iso,
            "qty_on_hand": batch['qty_on_hand'],
            "total_units": total_units_in_batch,
            "mrp": batch.get('mrp_per_unit', 0) * units_per_pack,
            "mrp_per_unit": batch.get('mrp_per_unit', 0)
        })
        total_qty += batch['qty_on_hand']
    
    return {
        "found": True,
        "has_stock": True,
        "product": {
            "product_id": product['id'],
            "sku": product['sku'],
            "name": product['name'],
            "brand": product.get('brand'),
            "pack_size": product.get('pack_size'),
            "units_per_pack": units_per_pack,
            "gst_percent": product.get('gst_percent', 5),
            "barcode": product.get('barcode'),
            "total_stock": total_qty,
            "total_units": total_qty * units_per_pack
        },
        "batches": formatted_batches,
        "suggested_batch": formatted_batches[0] if formatted_batches else None
    }

@api_router.get("/products/search-with-batches")
async def search_products_with_batches(
    q: str,
    location_id: Optional[str] = "default",
    current_user: User = Depends(get_current_user)
):
    """
    Search products and return with available batches (FEFO sorted)
    Searches by name, SKU, brand, and barcode
    """
    if len(q) < 2:
        return []
    
    # Search products - include barcode in search
    products = await db.products.find(
        {"$or": [
            {"name": {"$regex": q, "$options": "i"}},
            {"sku": {"$regex": q, "$options": "i"}},
            {"brand": {"$regex": q, "$options": "i"}},
            {"barcode": q}  # Exact match for barcode
        ]},
        {"_id": 0}
    ).to_list(50)
    
    results = []
    for product in products:
        # Backward compatibility for field names
        if 'default_mrp' in product and 'default_mrp_per_unit' not in product:
            product['default_mrp_per_unit'] = product['default_mrp']
        
        # Get batches for this product (FEFO - earliest expiry first)
        batches = await db.stock_batches.find(
            {
                "product_sku": product['sku'],
                "location": location_id,
                "qty_on_hand": {"$gt": 0}
            },
            {"_id": 0}
        ).sort("expiry_date", 1).to_list(10)
        
        if batches:
            # Format batch info
            formatted_batches = []
            total_qty = 0
            
            for batch in batches:
                expiry = batch.get('expiry_date')
                if expiry:
                    if isinstance(expiry, str):
                        expiry = datetime.fromisoformat(expiry)
                    expiry_display = expiry.strftime('%d-%m-%Y')
                    expiry_iso = expiry.isoformat()
                else:
                    expiry_display = 'N/A'
                    expiry_iso = None
                
                units_per_pack = product.get('units_per_pack', 1)
                total_units_in_batch = batch['qty_on_hand'] * units_per_pack
                
                formatted_batches.append({
                    "batch_id": batch['id'],
                    "batch_no": batch['batch_no'],
                    "expiry_date": expiry_display,
                    "expiry_iso": expiry_iso,
                    "qty_on_hand": batch['qty_on_hand'],  # Packs/strips
                    "total_units": total_units_in_batch,  # Individual tablets
                    "mrp": batch['mrp_per_unit'] * units_per_pack,  # Per pack (calculated)
                    "mrp_per_unit": batch['mrp_per_unit'],
                    "cost_price": batch['cost_price_per_unit'] * units_per_pack
                })
                total_qty += batch['qty_on_hand']
            
            units_per_pack = product.get('units_per_pack', 1)
            total_units = total_qty * units_per_pack
            
            # Handle both old and new field names for MRP
            default_mrp = product.get('default_mrp') or product.get('default_mrp_per_unit', 0)
            
            results.append({
                "product_id": product['id'],
                "sku": product['sku'],
                "name": product['name'],
                "brand": product.get('brand', ''),
                "manufacturer": product.get('manufacturer', ''),
                "composition": product.get('composition', ''),
                "pack_size": product.get('pack_size', ''),
                "units_per_pack": units_per_pack,
                "default_mrp": default_mrp,
                "gst_percent": product.get('gst_percent', 5),
                "schedule": product.get('schedule'),  # H, H1, X, or OTC
                "scheduleH": product.get('schedule') in ['H', 'H1'],  # Convenience boolean
                "total_qty": total_qty,  # Total packs
                "total_units": total_units,  # Total individual units
                "batches": formatted_batches,
                "suggested_batch": formatted_batches[0] if formatted_batches else None  # FEFO
            })
    
    return results

@api_router.get("/products/{product_id}", response_model=Product)
async def get_product(product_id: str, current_user: User = Depends(get_current_user)):
    product = await db.products.find_one({"id": product_id}, {"_id": 0})
    if not product:
        raise HTTPException(status_code=404, detail="Product not found")
    
    if isinstance(product['created_at'], str):
        product['created_at'] = datetime.fromisoformat(product['created_at'])
    if isinstance(product['updated_at'], str):
        product['updated_at'] = datetime.fromisoformat(product['updated_at'])
    
    return Product(**product)

@api_router.put("/products/{product_id}")
async def update_product(
    product_id: str,
    product_data: ProductUpdate,
    current_user: User = Depends(get_current_user)
):
    if current_user.role != "admin":
        raise HTTPException(status_code=403, detail="Only admins can update products")
    
    update_dict = {k: v for k, v in product_data.model_dump().items() if v is not None}
    if not update_dict:
        raise HTTPException(status_code=400, detail="No fields to update")
    
    update_dict['updated_at'] = datetime.now(timezone.utc).isoformat()
    update_dict['updated_by'] = current_user.id
    
    result = await db.products.update_one(
        {"id": product_id},
        {"$set": update_dict}
    )
    
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Product not found")
    
    return {"message": "Product updated successfully"}

@api_router.delete("/products/{product_id}")
async def delete_product(product_id: str, current_user: User = Depends(get_current_user)):
    if current_user.role != "admin":
        raise HTTPException(status_code=403, detail="Only admins can delete products")
    
    # Check if product has batches
    batches = await db.stock_batches.count_documents({"product_id": product_id})
    if batches > 0:
        raise HTTPException(status_code=400, detail="Cannot delete product with existing batches")
    
    result = await db.products.delete_one({"id": product_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Product not found")
    
    return {"message": "Product deleted successfully"}

@api_router.post("/products/bulk-update")
async def bulk_update_products(
    data: dict,
    current_user: User = Depends(get_current_user)
):
    """Bulk update multiple products with a single field change"""
    if current_user.role not in ["admin", "manager"]:
        raise HTTPException(status_code=403, detail="Only admins and managers can bulk update products")
    
    skus = data.get("skus", [])
    field = data.get("field", "")
    value = data.get("value", "")
    
    if not skus or not field:
        raise HTTPException(status_code=400, detail="SKUs and field are required")
    
    # Allowed fields for bulk update
    allowed_fields = ["location", "discount_percent", "gst_percent", "category", "schedule", "brand"]
    if field not in allowed_fields:
        raise HTTPException(status_code=400, detail=f"Field '{field}' not allowed for bulk update")
    
    # Convert value to appropriate type
    if field in ["discount_percent", "gst_percent"]:
        try:
            value = float(value)
        except ValueError:
            raise HTTPException(status_code=400, detail=f"Invalid numeric value for {field}")
    
    # Perform bulk update
    result = await db.products.update_many(
        {"sku": {"$in": skus}},
        {"$set": {field: value, "updated_at": datetime.now(timezone.utc).isoformat()}}
    )
    
    return {
        "message": f"Updated {result.modified_count} products",
        "modified_count": result.modified_count
    }

@api_router.get("/products/{sku}/transactions")
async def get_product_transactions(
    sku: str,
    transaction_type: str = "all",  # all, sales, purchases, sales_returns, purchase_returns
    page: int = 1,
    page_size: int = 50,
    current_user: User = Depends(get_current_user)
):
    """Get all transactions for a specific product SKU"""
    
    # Verify product exists
    product = await db.products.find_one({"sku": sku}, {"_id": 0})
    if not product:
        raise HTTPException(status_code=404, detail="Product not found")
    
    result = {
        "product_sku": sku,
        "product_name": product.get("name", ""),
        "sales": [],
        "purchases": [],
        "sales_returns": [],
        "purchase_returns": []
    }
    
    # Get Sales (bills with invoice_type SALE that contain this product)
    if transaction_type in ["all", "sales"]:
        sales_bills = await db.bills.find(
            {"invoice_type": "SALE"},
            {"_id": 0}
        ).sort("created_at", -1).to_list(1000)
        
        for bill in sales_bills:
            for item in bill.get("items", []):
                # Check if this item matches our product (by product_id/sku or product_name)
                item_sku = item.get("product_id") or item.get("product_sku")
                if item_sku == sku or item.get("product_name", "").lower() == product.get("name", "").lower():
                    result["sales"].append({
                        "id": bill.get("id"),
                        "bill_number": bill.get("bill_number"),
                        "date": bill.get("created_at"),
                        "customer_name": bill.get("customer_name") or "Walk-in",
                        "batch_no": item.get("batch_no") or item.get("batch_number") or "–",
                        "quantity": item.get("quantity", 0),
                        "unit_price": item.get("unit_price", 0),
                        "discount": item.get("discount", 0),
                        "line_total": item.get("line_total") or item.get("total", 0),
                        "status": bill.get("status", "paid")
                    })
    
    # Get Sales Returns (bills with invoice_type SALES_RETURN)
    if transaction_type in ["all", "sales_returns"]:
        return_bills = await db.bills.find(
            {"invoice_type": "SALES_RETURN"},
            {"_id": 0}
        ).sort("created_at", -1).to_list(1000)
        
        for bill in return_bills:
            for item in bill.get("items", []):
                item_sku = item.get("product_id") or item.get("product_sku")
                if item_sku == sku or item.get("product_name", "").lower() == product.get("name", "").lower():
                    result["sales_returns"].append({
                        "id": bill.get("id"),
                        "return_number": bill.get("bill_number"),
                        "date": bill.get("created_at"),
                        "customer_name": bill.get("customer_name") or "Walk-in",
                        "original_invoice": bill.get("ref_invoice_id"),
                        "batch_no": item.get("batch_no") or item.get("batch_number") or "–",
                        "quantity": item.get("quantity", 0),
                        "refund_amount": item.get("line_total") or item.get("total", 0),
                        "status": bill.get("status", "refunded")
                    })
    
    # Get Purchases
    if transaction_type in ["all", "purchases"]:
        purchases = await db.purchases.find({}, {"_id": 0}).sort("purchase_date", -1).to_list(1000)
        
        for purchase in purchases:
            for item in purchase.get("items", []):
                if item.get("product_sku") == sku:
                    result["purchases"].append({
                        "id": purchase.get("id"),
                        "purchase_number": purchase.get("purchase_number"),
                        "date": purchase.get("purchase_date"),
                        "supplier_name": purchase.get("supplier_name"),
                        "supplier_invoice": purchase.get("supplier_invoice_no") or "–",
                        "batch_no": item.get("batch_no") or "–",
                        "expiry_date": item.get("expiry_date"),
                        "quantity": item.get("qty_units", 0),
                        "cost_price": item.get("cost_price_per_unit", 0),
                        "mrp": item.get("mrp_per_unit", 0),
                        "line_total": item.get("line_total", 0),
                        "status": purchase.get("status", "draft")
                    })
    
    # Get Purchase Returns
    if transaction_type in ["all", "purchase_returns"]:
        purchase_returns = await db.purchase_returns.find({}, {"_id": 0}).sort("return_date", -1).to_list(1000)
        
        for pr in purchase_returns:
            for item in pr.get("items", []):
                if item.get("product_sku") == sku:
                    result["purchase_returns"].append({
                        "id": pr.get("id"),
                        "return_number": pr.get("return_number"),
                        "date": pr.get("return_date"),
                        "supplier_name": pr.get("supplier_name"),
                        "original_purchase": pr.get("purchase_number") or "–",
                        "batch_no": item.get("batch_no") or "–",
                        "quantity": item.get("qty_units", 0),
                        "reason": item.get("reason") or "–",
                        "line_total": item.get("line_total", 0),
                        "status": pr.get("status", "draft")
                    })
    
    return result

# ==================== STOCK BATCH ROUTES ====================

@api_router.post("/stock/batches", response_model=StockBatch)
async def create_stock_batch(batch_data: StockBatchCreate, current_user: User = Depends(get_current_user)):
    # Verify product exists by SKU
    product = await db.products.find_one({"sku": batch_data.product_sku}, {"_id": 0})
    if not product:
        raise HTTPException(status_code=404, detail="Product not found")
    
    # Check if batch already exists
    existing = await db.stock_batches.find_one({
        "product_sku": batch_data.product_sku,
        "batch_no": batch_data.batch_no,
        "location": batch_data.location or "default"
    }, {"_id": 0})
    
    if existing:
        raise HTTPException(status_code=400, detail="Batch with this number already exists for this product at this location")
    
    data = batch_data.model_dump()
    data['expiry_date'] = datetime.fromisoformat(batch_data.expiry_date)
    if batch_data.manufacture_date:
        data['manufacture_date'] = datetime.fromisoformat(batch_data.manufacture_date)
    if batch_data.received_date:
        data['received_date'] = datetime.fromisoformat(batch_data.received_date)
    
    batch = StockBatch(**data, created_by=current_user.id, updated_by=current_user.id)
    doc = batch.model_dump()
    doc['created_at'] = doc['created_at'].isoformat()
    doc['updated_at'] = doc['updated_at'].isoformat()
    doc['expiry_date'] = doc['expiry_date'].isoformat()
    if doc.get('manufacture_date'):
        doc['manufacture_date'] = doc['manufacture_date'].isoformat()
    if doc.get('received_date'):
        doc['received_date'] = doc['received_date'].isoformat()
    await db.stock_batches.insert_one(doc)
    
    # Create opening stock movement (Phase 0 requirement)
    units_per_pack = product.get('units_per_pack', 1)
    qty_units = batch_data.qty_on_hand * units_per_pack
    
    movement = StockMovement(
        product_sku=batch_data.product_sku,
        batch_id=batch.id,
        product_name=product['name'],
        batch_no=batch_data.batch_no,
        qty_delta_units=qty_units,
        movement_type='opening_stock',
        ref_type='opening',
        ref_id=batch.id,
        location=batch_data.location or "default",
        reason="Initial stock entry",
        performed_by=current_user.id
    )
    movement_doc = movement.model_dump()
    movement_doc['performed_at'] = movement_doc['performed_at'].isoformat()
    await db.stock_movements.insert_one(movement_doc)
    
    return batch

@api_router.get("/stock/batches")
async def get_stock_batches(
    product_sku: Optional[str] = None,
    location: Optional[str] = None,
    current_user: User = Depends(get_current_user)
):
    query = {}
    if product_sku:
        query["product_sku"] = product_sku
    if location:
        query["location"] = location
    
    batches = await db.stock_batches.find(query, {"_id": 0}).sort("expiry_date", 1).to_list(10000)
    
    # Enrich with product info and calculate units
    for batch in batches:
        if isinstance(batch.get('created_at'), str):
            batch['created_at'] = datetime.fromisoformat(batch['created_at'])
        if isinstance(batch.get('updated_at'), str):
            batch['updated_at'] = datetime.fromisoformat(batch['updated_at'])
        if isinstance(batch.get('expiry_date'), str):
            batch['expiry_date'] = datetime.fromisoformat(batch['expiry_date'])
        if isinstance(batch.get('manufacture_date'), str):
            batch['manufacture_date'] = datetime.fromisoformat(batch['manufacture_date'])
        if isinstance(batch.get('received_date'), str):
            batch['received_date'] = datetime.fromisoformat(batch['received_date'])
        
        # Add product info - handle legacy batches that may not have product_sku
        product_sku = batch.get('product_sku') or batch.get('product_id')
        if product_sku:
            product = await db.products.find_one({"sku": product_sku}, {"_id": 0, "name": 1, "brand": 1, "sku": 1, "units_per_pack": 1})
            if product:
                batch['product_name'] = product.get('name', '')
                batch['product_brand'] = product.get('brand', '')
                batch['product_sku'] = product.get('sku', '')
                units_per_pack = product.get('units_per_pack', 1)
                batch['total_units'] = batch.get('qty_on_hand', 0) * units_per_pack
            else:
                batch['product_name'] = ''
                batch['product_brand'] = ''
                batch['total_units'] = batch.get('qty_on_hand', 0)
        else:
            batch['product_name'] = ''
            batch['product_brand'] = ''
            batch['total_units'] = batch.get('qty_on_hand', 0)
    
    return batches

@api_router.get("/stock/batches/{batch_id}")
async def get_stock_batch(batch_id: str, current_user: User = Depends(get_current_user)):
    batch = await db.stock_batches.find_one({"id": batch_id}, {"_id": 0})
    if not batch:
        raise HTTPException(status_code=404, detail="Batch not found")
    
    if isinstance(batch['created_at'], str):
        batch['created_at'] = datetime.fromisoformat(batch['created_at'])
    if isinstance(batch['updated_at'], str):
        batch['updated_at'] = datetime.fromisoformat(batch['updated_at'])
    if isinstance(batch['expiry_date'], str):
        batch['expiry_date'] = datetime.fromisoformat(batch['expiry_date'])
    
    return batch

@api_router.put("/stock/batches/{batch_id}")
async def update_stock_batch(
    batch_id: str,
    batch_data: StockBatchUpdate,
    current_user: User = Depends(get_current_user)
):
    if current_user.role != "admin":
        raise HTTPException(status_code=403, detail="Only admins can update stock batches")
    
    update_dict = {k: v for k, v in batch_data.model_dump().items() if v is not None}
    if not update_dict:
        raise HTTPException(status_code=400, detail="No fields to update")
    
    update_dict['updated_at'] = datetime.now(timezone.utc).isoformat()
    
    result = await db.stock_batches.update_one(
        {"id": batch_id},
        {"$set": update_dict}
    )
    
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Batch not found")
    
    return {"message": "Batch updated successfully"}

@api_router.post("/batches/{batch_id}/adjust")
async def adjust_stock(
    batch_id: str,
    adjustment: StockAdjustment,
    current_user: User = Depends(get_current_user)
):
    """
    Adjust stock for a batch - Phase 0 requirement
    Supports adding or removing units with reason tracking
    """
    # Get batch
    batch = await db.stock_batches.find_one({"id": batch_id}, {"_id": 0})
    if not batch:
        raise HTTPException(status_code=404, detail="Batch not found")
    
    # Get product for units_per_pack
    product = await db.products.find_one({"sku": batch['product_sku']}, {"_id": 0})
    if not product:
        raise HTTPException(status_code=404, detail="Product not found")
    
    units_per_pack = product.get('units_per_pack', 1)
    
    # Calculate pack change from unit change
    qty_delta_units = adjustment.qty_units if adjustment.adjustment_type == 'add' else -adjustment.qty_units
    pack_delta = qty_delta_units / units_per_pack
    
    # Calculate new quantity
    new_qty = batch['qty_on_hand'] + pack_delta
    
    # Validation: cannot go negative
    if new_qty < 0:
        raise HTTPException(
            status_code=400,
            detail=f"Cannot remove {adjustment.qty_units} units. Only {int(batch['qty_on_hand'] * units_per_pack)} units available."
        )
    
    # Update batch
    result = await db.stock_batches.update_one(
        {"id": batch_id},
        {
            "$set": {
                "qty_on_hand": new_qty,
                "updated_at": datetime.now(timezone.utc).isoformat(),
                "updated_by": current_user.id
            }
        }
    )
    
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Batch not found")
    
    # Create stock movement record
    movement = StockMovement(
        product_sku=batch['product_sku'],
        batch_id=batch_id,
        product_name=product['name'],
        batch_no=batch['batch_no'],
        qty_delta_units=qty_delta_units,
        movement_type='adjustment',
        ref_type='adjustment',
        ref_id=str(uuid.uuid4()),  # Generate adjustment ID
        location=batch.get('location', 'default'),
        reason=adjustment.reason,
        performed_by=current_user.id
    )
    
    movement_doc = movement.model_dump()
    movement_doc['performed_at'] = movement_doc['performed_at'].isoformat()
    movement_doc['reference_number'] = adjustment.reference_number
    movement_doc['notes'] = adjustment.notes
    await db.stock_movements.insert_one(movement_doc)
    
    return {
        "message": "Stock adjusted successfully",
        "new_qty_packs": new_qty,
        "new_qty_units": int(new_qty * units_per_pack),
        "adjustment_units": qty_delta_units
    }

@api_router.post("/batches/{batch_id}/writeoff-expiry")
async def writeoff_expired_batch(
    batch_id: str,
    writeoff_data: dict,
    current_user: User = Depends(get_current_user)
):
    """Write off expired stock - removes expired inventory"""
    # Get batch
    batch = await db.stock_batches.find_one({"id": batch_id}, {"_id": 0})
    if not batch:
        raise HTTPException(status_code=404, detail="Batch not found")
    
    # Get product
    product = await db.products.find_one({"sku": batch['product_sku']}, {"_id": 0})
    if not product:
        raise HTTPException(status_code=404, detail="Product not found")
    
    # Verify batch is expired
    expiry_str = batch.get('expiry_date')
    if expiry_str:
        try:
            if isinstance(expiry_str, str):
                expiry_date = datetime.fromisoformat(expiry_str.replace('Z', '+00:00'))
            else:
                expiry_date = expiry_str
            
            if expiry_date >= datetime.now(timezone.utc):
                raise HTTPException(status_code=400, detail="Batch is not expired yet")
        except:
            pass
    
    # Get current quantity
    qty_on_hand = batch.get('qty_on_hand', 0)
    units_per_pack = product.get('units_per_pack', 1)
    qty_units = int(qty_on_hand * units_per_pack)
    
    if qty_units <= 0:
        raise HTTPException(status_code=400, detail="No stock to write off")
    
    # Set quantity to 0
    await db.stock_batches.update_one(
        {"id": batch_id},
        {
            "$set": {
                "qty_on_hand": 0,
                "updated_at": datetime.now(timezone.utc).isoformat(),
                "updated_by": current_user.id,
                "status": "written_off"
            }
        }
    )
    
    # Create stock movement for expiry write-off
    movement = StockMovement(
        product_sku=batch['product_sku'],
        batch_id=batch_id,
        product_name=product['name'],
        batch_no=batch['batch_no'],
        qty_delta_units=-qty_units,  # Negative for removal
        movement_type='expiry_writeoff',
        ref_type='writeoff',
        ref_id=str(uuid.uuid4()),
        location=batch.get('location', 'default'),
        reason=writeoff_data.get('reason', 'Expired stock write-off'),
        performed_by=current_user.id
    )
    
    movement_doc = movement.model_dump()
    movement_doc['performed_at'] = movement_doc['performed_at'].isoformat()
    await db.stock_movements.insert_one(movement_doc)
    
    return {
        "message": "Expired stock written off successfully",
        "qty_written_off_units": qty_units,
        "batch_id": batch_id
    }

@api_router.get("/stock/summary")
async def get_stock_summary(
    product_id: Optional[str] = None,
    current_user: User = Depends(get_current_user)
):
    """Get stock summary by product (total qty across all batches)"""
    query = {}
    if product_id:
        query["product_id"] = product_id
    
    batches = await db.stock_batches.find(query, {"_id": 0}).to_list(10000)
    
    # Group by product
    summary = {}
    for batch in batches:
        pid = batch['product_id']
        if pid not in summary:
            product = await db.products.find_one({"id": pid}, {"_id": 0})
            summary[pid] = {
                "product_id": pid,
                "product_name": product['name'] if product else "Unknown",
                "product_sku": product['sku'] if product else "",
                "total_qty": 0,
                "batches_count": 0,
                "earliest_expiry": None
            }
        
        summary[pid]["total_qty"] += batch['qty_on_hand']
        summary[pid]["batches_count"] += 1
        
        expiry = batch['expiry_date']
        if isinstance(expiry, str):
            expiry = datetime.fromisoformat(expiry)
        
        if summary[pid]["earliest_expiry"] is None or expiry < summary[pid]["earliest_expiry"]:
            summary[pid]["earliest_expiry"] = expiry


@api_router.delete("/stock/batches/{batch_id}")
async def delete_stock_batch(batch_id: str, current_user: User = Depends(get_current_user)):
    """Delete a stock batch (only if qty_on_hand = 0)"""
    if current_user.role != "admin":
        raise HTTPException(status_code=403, detail="Only admins can delete stock batches")
    
    # Check if batch exists and has zero quantity
    batch = await db.stock_batches.find_one({"id": batch_id}, {"_id": 0})
    if not batch:
        raise HTTPException(status_code=404, detail="Batch not found")
    
    if batch.get('qty_on_hand', 0) > 0:
        raise HTTPException(status_code=400, detail="Cannot delete batch with stock. Adjust quantity to 0 first.")
    
    result = await db.stock_batches.delete_one({"id": batch_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Batch not found")
    
    return {"message": "Batch deleted successfully"}

# ==================== STOCK MOVEMENT ROUTES ====================

@api_router.post("/stock-movements")
async def create_stock_movement(
    movement_data: StockMovementCreate,
    current_user: User = Depends(get_current_user)
):
    """Create a stock movement record - Phase 0"""
    movement = StockMovement(
        **movement_data.model_dump(),
        performed_by=current_user.id
    )
    
    movement_doc = movement.model_dump()
    movement_doc['performed_at'] = movement_doc['performed_at'].isoformat()
    await db.stock_movements.insert_one(movement_doc)
    
    return {"message": "Stock movement recorded", "id": movement.id}

@api_router.get("/stock-movements")
async def get_stock_movements(
    product_sku: Optional[str] = None,
    batch_id: Optional[str] = None,
    movement_type: Optional[str] = None,
    limit: int = 100,
    current_user: User = Depends(get_current_user)
):
    """Get stock movements with filters - Phase 0 Stock Ledger"""
    query = {}
    if product_sku:
        query["product_sku"] = product_sku
    if batch_id:
        query["batch_id"] = batch_id
    if movement_type:
        query["movement_type"] = movement_type
    
    movements = await db.stock_movements.find(query, {"_id": 0}).sort("performed_at", -1).limit(limit).to_list(limit)
    
    for movement in movements:
        if isinstance(movement.get('performed_at'), str):
            movement['performed_at'] = datetime.fromisoformat(movement['performed_at'])
    
    return movements


    
    return list(summary.values())


# ==================== MIGRATION ROUTE ====================

@api_router.post("/migrate/medicines-to-products")
async def migrate_medicines_to_products(current_user: User = Depends(get_current_user)):
    """
    One-time migration: Convert old Medicine documents to Product + StockBatch model
    """
    if current_user.role != "admin":
        raise HTTPException(status_code=403, detail="Only admins can run migrations")
    
    try:
        medicines = await db.medicines.find({}, {"_id": 0}).to_list(10000)
        
        if not medicines:
            return {
                "message": "No medicines to migrate",
                "products_created": 0,
                "batches_created": 0
            }
        
        products_created = 0
        batches_created = 0
        errors = []
        
        for med in medicines:
            try:
                # Generate SKU from name if not exists
                sku = med.get('hsn_code') or f"MED-{med['id'][:8]}"
                
                # Check if product already exists
                existing_product = await db.products.find_one({"sku": sku}, {"_id": 0})
                
                if not existing_product:
                    # Create Product
                    product = Product(
                        id=med['id'],  # Keep same ID for consistency
                        sku=sku,
                        name=med['name'],
                        brand=None,
                        pack_size=None,
                        category=None,
                        default_mrp=med.get('mrp', 0),
                        gst_percent=5.0,
                        hsn_code=med.get('hsn_code'),
                        description=None
                    )
                    
                    product_doc = product.model_dump()
                    product_doc['created_at'] = med.get('created_at', datetime.now(timezone.utc).isoformat())
                    if isinstance(product_doc['created_at'], datetime):
                        product_doc['created_at'] = product_doc['created_at'].isoformat()
                    product_doc['updated_at'] = med.get('updated_at', datetime.now(timezone.utc).isoformat())
                    if isinstance(product_doc['updated_at'], datetime):
                        product_doc['updated_at'] = product_doc['updated_at'].isoformat()
                    
                    await db.products.insert_one(product_doc)
                    products_created += 1
                    product_id = med['id']
                else:
                    product_id = existing_product['id']
                
                # Create StockBatch
                expiry = med.get('expiry_date')
                if isinstance(expiry, str):
                    expiry = datetime.fromisoformat(expiry)
                elif not isinstance(expiry, datetime):
                    expiry = datetime.now(timezone.utc) + timedelta(days=365)
                
                # Get product SKU for the batch
                product_sku = sku
                
                batch = StockBatch(
                    product_sku=product_sku,
                    batch_no=med.get('batch_number', 'BATCH-001'),
                    expiry_date=expiry,
                    qty_on_hand=med.get('quantity', 0),
                    cost_price_per_unit=med.get('purchase_rate', 0),
                    mrp_per_unit=med.get('mrp', 0),
                    supplier_name=med.get('supplier_name'),
                    location="default"
                )
                
                batch_doc = batch.model_dump()
                batch_doc['created_at'] = batch_doc['created_at'].isoformat()
                batch_doc['updated_at'] = batch_doc['updated_at'].isoformat()
                batch_doc['expiry_date'] = batch_doc['expiry_date'].isoformat()
                
                await db.stock_batches.insert_one(batch_doc)
                batches_created += 1
                
            except Exception as e:
                errors.append(f"Error migrating medicine {med.get('id', 'unknown')}: {str(e)}")
                logger.error(f"Migration error for medicine {med.get('id')}: {e}")
                continue
        
        return {
            "message": "Migration completed",
            "products_created": products_created,
            "batches_created": batches_created,
            "total_medicines_processed": len(medicines),
            "errors": errors
        }
    
    except Exception as e:
        logger.error(f"Migration failed: {e}")
        raise HTTPException(status_code=500, detail=f"Migration failed: {str(e)}")

# ==================== BILLING ROUTES ====================

@api_router.post("/bills", response_model=Bill)
async def create_bill(bill_data: BillCreate, current_user: User = Depends(get_current_user)):
    # For draft bills, use simple "Draft" label instead of sequential number
    if bill_data.status == "draft":
        # Draft bills get a simple "Draft" identifier, not a real bill number
        bill_number = "Draft"
    else:
        # Generate sequential bill number using atomic operation
        # This ensures no duplicates even with concurrent settlements
        bill_number = await generate_bill_number(bill_data.invoice_type)
    
    # Calculate totals
    # Note: line_total from frontend already includes per-item GST
    # So subtotal here is actually total with tax included
    subtotal = sum(item.get('line_total', item.get('total', 0)) for item in bill_data.items)
    
    # Calculate tax amount for records (approximate - items may have different rates)
    # This is informational only since line_total already has tax
    tax_amount = 0
    for item in bill_data.items:
        base_amt = item.get('unit_price', item.get('mrp', 0)) * item.get('quantity', 0) - item.get('discount', 0)
        gst_pct = item.get('gst_percent', bill_data.tax_rate or 5)
        tax_amount += base_amt * (gst_pct / 100)
    
    total_amount = subtotal - (bill_data.discount or 0)  # Don't add tax again, it's in line_total
    
    # Calculate paid and due amounts
    paid_amount = 0
    if bill_data.payments:
        # Multiple payments provided - handle both dict and object formats
        for p in bill_data.payments:
            if isinstance(p, dict):
                paid_amount += p.get('amount', 0)
            else:
                paid_amount += getattr(p, 'amount', 0)
    elif bill_data.refund and bill_data.invoice_type == "SALES_RETURN":
        # For returns, treat refund amount as "paid"
        if isinstance(bill_data.refund, dict):
            paid_amount = bill_data.refund.get('amount', total_amount)
        else:
            paid_amount = getattr(bill_data.refund, 'amount', total_amount)
    elif bill_data.payment_method:
        # Legacy: single payment method, assume full payment
        paid_amount = total_amount if bill_data.status == "paid" else 0
    
    due_amount = max(0, total_amount - paid_amount)
    
    # Determine status based on payments
    # Draft bills are stored with 'due' status but flagged as draft internally
    is_draft = bill_data.status == "draft"
    if is_draft:
        status = "due"  # Drafts show as "due" in the UI
    elif bill_data.invoice_type == "SALES_RETURN" and bill_data.refund:
        # For returns with refund data, mark as paid/refunded
        status = "paid"
    elif abs(due_amount) < 0.01:  # Allow for small rounding differences
        status = "paid"
    elif paid_amount > 0:
        status = "due"  # Partially paid
    else:
        status = "due"  # Not paid at all but not draft
    
    bill = Bill(
        bill_number=bill_number,
        invoice_type=bill_data.invoice_type,
        ref_invoice_id=bill_data.ref_invoice_id,
        status=status,
        customer_id=bill_data.customer_id,
        customer_name=bill_data.customer_name,
        customer_mobile=bill_data.customer_mobile,
        doctor_id=bill_data.doctor_id,
        doctor_name=bill_data.doctor_name,
        items=bill_data.items,
        subtotal=subtotal,
        discount=bill_data.discount,
        tax_rate=bill_data.tax_rate,
        tax_amount=tax_amount,
        total_amount=total_amount,
        paid_amount=paid_amount,
        due_amount=due_amount,
        payment_method=bill_data.payment_method,  # Legacy field
        cashier_id=current_user.id,
        cashier_name=current_user.name
    )
    
    # Only update stock if NOT draft - allow stock deduction for paid, due, etc.
    if bill_data.status != "draft":
        # Pre-check for Schedule H1 drugs - validate prescription before any stock deduction
        for item in bill_data.items:
            product_sku = item.get('product_sku')
            if product_sku:
                product = await db.products.find_one({"sku": product_sku}, {"_id": 0})
                if product and product.get('schedule') == 'H1':
                    product_name = product['name']
                    if not bill_data.doctor_name or not bill_data.doctor_name.strip():
                        raise HTTPException(
                            status_code=400,
                            detail=f"Prescription details required for Schedule H1 drug: {product_name}"
                        )
        
        for item in bill_data.items:
            # Support both old (medicine_id) and new (product_id/batch_id) format
            batch_id = item.get('batch_id')
            product_id = item.get('product_id') or item.get('medicine_id')
            product_sku = item.get('product_sku')
            batch_no = item.get('batch_no') or item.get('batch_number')
            
            # If no batch_id but we have batch_no and product_sku, find batch by those
            if not batch_id and batch_no and product_sku:
                batch_doc = await db.stock_batches.find_one(
                    {"product_sku": product_sku, "batch_no": batch_no},
                    {"_id": 0}
                )
                if batch_doc:
                    batch_id = batch_doc['id']
            
            if not batch_id and product_id:
                # Legacy support: if batch_id not provided, try to find a batch for this product
                batches = await db.stock_batches.find(
                    {"product_id": product_id, "qty_on_hand": {"$gt": 0}},
                    {"_id": 0}
                ).sort("expiry_date", 1).to_list(1)
                
                if batches:
                    batch_id = batches[0]['id']
                else:
                    # No batch found, skip stock update for this item
                    logger.warning(f"No batch found for product {product_id}")
                    continue
            
            if not batch_id:
                continue
            
            # Get product to determine units_per_pack for conversion
            # Support both old (product_id) and new (product_sku) lookups
            product = await db.products.find_one({"id": product_id}, {"_id": 0}) if product_id else None
            if not product:
                # Try finding by SKU directly from item
                item_sku = item.get('product_sku')
                if item_sku:
                    product = await db.products.find_one({"sku": item_sku}, {"_id": 0})
            if not product:
                # Try finding by SKU from batch
                batch_doc = await db.stock_batches.find_one({"id": batch_id}, {"_id": 0})
                if batch_doc and 'product_sku' in batch_doc:
                    product = await db.products.find_one({"sku": batch_doc['product_sku']}, {"_id": 0})
            
            if not product:
                logger.error(f"Product {product_id} not found")
                continue
            
            units_per_pack = product.get('units_per_pack', 1)
            product_sku = product.get('sku')
            
            # item['quantity'] is in UNITS (tablets)
            # Convert to PACKS for stock deduction
            quantity_in_units = item['quantity']
            quantity_in_packs = quantity_in_units / units_per_pack
            
            # Determine quantity change based on invoice type (in packs)
            pack_change = -quantity_in_packs if bill_data.invoice_type == "SALE" else quantity_in_packs
            
            # Update batch stock (qty_on_hand is in packs)
            result = await db.stock_batches.update_one(
                {"id": batch_id},
                {"$inc": {"qty_on_hand": pack_change}}
            )
            
            if result.matched_count == 0:
                logger.error(f"Batch {batch_id} not found")
                continue
            
            # Create stock movement record (Phase 0 compliant)
            batch = await db.stock_batches.find_one({"id": batch_id}, {"_id": 0})
            
            # Convert pack change to units for stock movement
            qty_delta_units = int(pack_change * units_per_pack)
            
            movement = StockMovement(
                product_sku=product_sku,
                batch_id=batch_id,
                product_name=product['name'] if product else item.get('product_name', item.get('medicine_name', 'Unknown')),
                batch_no=batch['batch_no'] if batch else item.get('batch_no', item.get('batch_number', 'N/A')),
                qty_delta_units=qty_delta_units,  # Phase 0: quantity in units
                movement_type="sale" if bill_data.invoice_type == "SALE" else "sales_return",
                ref_type="invoice",  # Phase 0: ref_type instead of ref_entity
                ref_id=bill.id,
                location=batch.get('location', 'default') if batch else "default",
                performed_by=current_user.id
            )
            movement_doc = movement.model_dump()
            movement_doc['performed_at'] = movement_doc['performed_at'].isoformat()
            await db.stock_movements.insert_one(movement_doc)
            
            # Schedule H1 compliance: auto-create register entry for H1 drugs
            if product.get('schedule') == 'H1' and bill_data.invoice_type == "SALE":
                product_name = product['name'] if product else item.get('product_name', item.get('medicine_name', 'Unknown'))
                
                # Prescription (doctor) details required for H1 drugs
                if not bill_data.doctor_name or not bill_data.doctor_name.strip():
                    raise HTTPException(
                        status_code=400,
                        detail=f"Prescription details required for Schedule H1 drug: {product_name}"
                    )
                
                # Look up prescriber details from doctors collection
                prescriber_address = ""
                prescriber_registration_no = ""
                if bill_data.doctor_name:
                    doctor = await db.doctors.find_one(
                        {"name": {"$regex": f"^{bill_data.doctor_name}$", "$options": "i"}},
                        {"_id": 0}
                    )
                    if doctor:
                        prescriber_address = doctor.get('clinic_address', '') or doctor.get('address', '') or ''
                        prescriber_registration_no = doctor.get('registration_no', '') or doctor.get('contact', '') or ''
                
                # Create H1 register entry
                h1_entry = ScheduleH1Entry(
                    product_sku=product_sku,
                    product_name=product_name,
                    batch_no=batch['batch_no'] if batch else item.get('batch_no', item.get('batch_number', 'N/A')),
                    quantity_dispensed=abs(qty_delta_units),
                    prescriber_name=bill_data.doctor_name,
                    prescriber_address=prescriber_address,
                    prescriber_registration_no=prescriber_registration_no,
                    patient_name=bill_data.customer_name or 'Walk-in Customer',
                    patient_address=None,
                    bill_id=bill.id,
                    bill_number=bill_number,
                    dispensed_by=current_user.id,
                    dispensed_by_name=current_user.name
                )
                h1_doc = h1_entry.model_dump()
                h1_doc['dispensed_at'] = h1_doc['dispensed_at'].isoformat()
                await db.schedule_h1_register.insert_one(h1_doc)
    
    # Save bill first
    doc = bill.model_dump()
    doc['created_at'] = doc['created_at'].isoformat()
    await db.bills.insert_one(doc)
    
    # Create payment records if payments provided
    if bill_data.payments and len(bill_data.payments) > 0:
        for payment_data in bill_data.payments:
            payment = Payment(
                invoice_id=bill.id,
                amount=payment_data.get('amount', 0),
                payment_method=payment_data.get('method', 'cash'),
                reference_number=payment_data.get('reference', None),
                notes=payment_data.get('notes', None),
                created_by=current_user.id
            )
            payment_doc = payment.model_dump()
            payment_doc['created_at'] = payment_doc['created_at'].isoformat()
            await db.payments.insert_one(payment_doc)
    
    # Create refund record for sales returns if refund data provided
    if bill_data.invoice_type == "SALES_RETURN" and bill_data.refund:
        refund = Refund(
            return_invoice_id=bill.id,
            original_invoice_id=bill_data.ref_invoice_id,
            amount=bill_data.refund.get('amount', total_amount),
            refund_method=bill_data.refund.get('method', 'cash'),
            reference_number=bill_data.refund.get('reference', None),
            reason=bill_data.refund.get('reason', None),
            notes=bill_data.refund.get('notes', None),
            created_by=current_user.id
        )
        refund_doc = refund.model_dump()
        refund_doc['created_at'] = refund_doc['created_at'].isoformat()
        await db.refunds.insert_one(refund_doc)
        
        # Audit log for refund
        await create_audit_log(
            entity_type='refund',
            entity_id=refund.id,
            action='create',
            user=current_user,
            new_value=refund_doc,
            reason=bill_data.refund.get('reason')
        )
    
    # Audit log for invoice creation
    await create_audit_log(
        entity_type='invoice',
        entity_id=bill.id,
        action='create',
        user=current_user,
        new_value={
            'bill_number': bill.bill_number,
            'invoice_type': bill.invoice_type,
            'status': bill.status,
            'customer_name': bill.customer_name,
            'total_amount': bill.total_amount,
            'paid_amount': bill.paid_amount,
            'due_amount': bill.due_amount
        }
    )
    
    return bill

@api_router.put("/bills/{bill_id}")
async def update_bill(
    bill_id: str,
    bill_data: BillCreate,
    current_user: User = Depends(get_current_user)
):
    """Update a draft bill - only drafts can be modified"""
    existing_bill = await db.bills.find_one({"id": bill_id})
    
    if not existing_bill:
        raise HTTPException(status_code=404, detail="Bill not found")
    
    if existing_bill.get('status') != 'draft':
        raise HTTPException(status_code=400, detail="Only draft bills can be edited")
    
    # Calculate totals
    items = []
    subtotal = 0
    total_discount = bill_data.discount or 0
    total_tax = 0
    
    for item_data in bill_data.items:
        item = BillItem(**item_data.dict() if hasattr(item_data, 'dict') else item_data)
        items.append(item)
        subtotal += item.mrp * item.quantity
        total_discount += item.discount
        total_tax += (item.mrp * item.quantity - item.discount) * (item.gst_percent or 5) / 100
    
    total_amount = subtotal - total_discount + total_tax
    
    # Update bill
    update_data = {
        "customer_name": bill_data.customer_name or "Counter Sale",
        "customer_mobile": bill_data.customer_mobile,
        "doctor_name": bill_data.doctor_name,
        "items": [item.model_dump() for item in items],
        "subtotal": round(subtotal, 2),
        "total_discount": round(total_discount, 2),
        "total_tax": round(total_tax, 2),
        "total_amount": round(total_amount, 2),
        "discount": bill_data.discount or 0,
        "status": bill_data.status or "draft",
        "updated_at": datetime.now(timezone.utc).isoformat(),
        "updated_by": current_user.id
    }
    
    # Handle status change to paid
    if bill_data.status == "paid" and existing_bill.get('status') == 'draft':
        # Process payments
        if bill_data.payments:
            paid_amount = sum(p.get('amount', 0) if isinstance(p, dict) else p.amount for p in bill_data.payments)
            update_data["paid_amount"] = round(paid_amount, 2)
            update_data["due_amount"] = round(max(0, total_amount - paid_amount), 2)
            
            # Save payment records
            for payment_data in bill_data.payments:
                payment = Payment(
                    invoice_id=bill_id,
                    amount=payment_data.get('amount', 0) if isinstance(payment_data, dict) else payment_data.amount,
                    payment_method=payment_data.get('method', 'cash') if isinstance(payment_data, dict) else payment_data.method,
                    reference_number=payment_data.get('reference') if isinstance(payment_data, dict) else payment_data.reference,
                    created_by=current_user.id
                )
                payment_doc = payment.model_dump()
                payment_doc['created_at'] = payment_doc['created_at'].isoformat()
                await db.payments.insert_one(payment_doc)
        
        # Deduct stock for SALE
        if existing_bill.get('invoice_type') == 'SALE':
            for item in items:
                batch_id = item.batch_id
                if batch_id:
                    batch = await db.stock_batches.find_one({"id": batch_id})
                    if batch:
                        new_qty = batch.get('qty_on_hand', 0) - item.quantity
                        await db.stock_batches.update_one(
                            {"id": batch_id},
                            {"$set": {"qty_on_hand": new_qty}}
                        )
                        
                        # Record stock movement
                        movement = StockMovement(
                            product_sku=batch.get('product_sku', ''),
                            batch_id=batch_id,
                            movement_type="sale",
                            qty_delta_units=-item.quantity,
                            reason=f"Bill {existing_bill.get('bill_number')}",
                            ref_id=bill_id,
                            performed_by=current_user.id
                        )
                        movement_doc = movement.model_dump()
                        movement_doc['performed_at'] = movement_doc['performed_at'].isoformat()
                        await db.stock_movements.insert_one(movement_doc)
    
    await db.bills.update_one({"id": bill_id}, {"$set": update_data})
    
    # Return updated bill
    updated_bill = await db.bills.find_one({"id": bill_id}, {"_id": 0})
    return updated_bill

@api_router.get("/bills")
async def get_bills(
    invoice_type: Optional[str] = None,
    status: Optional[str] = None,
    search: Optional[str] = None,
    from_date: Optional[str] = None,
    to_date: Optional[str] = None,
    page: int = 1,
    page_size: int = 50,
    current_user: User = Depends(get_current_user)
):
    """Get bills with pagination and filters"""
    # Validate pagination params
    page_size = min(max(page_size, 1), 100)  # Clamp between 1 and 100
    page = max(page, 1)
    
    query = {}
    if invoice_type:
        query["invoice_type"] = invoice_type
    if status:
        query["status"] = status
    if search:
        query["$or"] = [
            {"bill_number": {"$regex": search, "$options": "i"}},
            {"customer_name": {"$regex": search, "$options": "i"}},
            {"customer_phone": {"$regex": search, "$options": "i"}}
        ]
    if from_date:
        query["created_at"] = {"$gte": from_date}
    if to_date:
        if "created_at" in query:
            query["created_at"]["$lte"] = to_date
        else:
            query["created_at"] = {"$lte": to_date}
    
    # Get total count
    total = await db.bills.count_documents(query)
    
    # Paginate
    skip = (page - 1) * page_size
    bills = await db.bills.find(query, {"_id": 0}).sort("created_at", -1).skip(skip).limit(page_size).to_list(page_size)
    
    for bill in bills:
        if isinstance(bill['created_at'], str):
            bill['created_at'] = datetime.fromisoformat(bill['created_at'])
        if 'invoice_type' not in bill:
            bill['invoice_type'] = 'SALE'
        if 'status' not in bill:
            bill['status'] = 'paid'
    
    return {
        "data": bills,
        "pagination": {
            "page": page,
            "page_size": page_size,
            "total": total,
            "total_pages": (total + page_size - 1) // page_size,
            "has_next": page * page_size < total,
            "has_prev": page > 1
        }
    }

@api_router.get("/bills/{bill_id}", response_model=Bill)
async def get_bill(bill_id: str, current_user: User = Depends(get_current_user)):
    bill = await db.bills.find_one({"id": bill_id}, {"_id": 0})
    if not bill:
        raise HTTPException(status_code=404, detail="Bill not found")
    
    if isinstance(bill['created_at'], str):
        bill['created_at'] = datetime.fromisoformat(bill['created_at'])
    
    return Bill(**bill)


# ==================== PAYMENT ROUTES ====================

@api_router.post("/payments", response_model=Payment)
async def create_payment(payment_data: PaymentCreate, current_user: User = Depends(get_current_user)):
    """Record a payment against an invoice"""
    # Verify invoice exists
    invoice = await db.bills.find_one({"id": payment_data.invoice_id}, {"_id": 0})
    if not invoice:
        raise HTTPException(status_code=404, detail="Invoice not found")
    
    payment = Payment(
        **payment_data.model_dump(),
        created_by=current_user.id
    )
    
    payment_doc = payment.model_dump()
    payment_doc['created_at'] = payment_doc['created_at'].isoformat()
    await db.payments.insert_one(payment_doc)
    
    # Update invoice paid_amount and due_amount
    total_paid = await db.payments.aggregate([
        {"$match": {"invoice_id": payment_data.invoice_id}},
        {"$group": {"_id": None, "total": {"$sum": "$amount"}}}
    ]).to_list(1)
    
    new_paid_amount = total_paid[0]['total'] if total_paid else 0
    new_due_amount = invoice['total_amount'] - new_paid_amount
    new_status = "paid" if new_due_amount <= 0 else "due"
    
    await db.bills.update_one(
        {"id": payment_data.invoice_id},
        {"$set": {
            "paid_amount": new_paid_amount,
            "due_amount": new_due_amount,
            "status": new_status
        }}
    )
    
    # Audit log for payment
    await create_audit_log(
        entity_type='payment',
        entity_id=payment.id,
        action='create',
        user=current_user,
        new_value=payment_doc
    )
    
    # Audit log for invoice status change
    if invoice.get('status') != new_status:
        await create_audit_log(
            entity_type='invoice',
            entity_id=payment_data.invoice_id,
            action='status_change',
            user=current_user,
            old_value={'status': invoice.get('status'), 'due_amount': invoice.get('due_amount', 0)},
            new_value={'status': new_status, 'due_amount': new_due_amount}
        )
    
    return payment

@api_router.get("/payments")
async def get_payments(
    invoice_id: Optional[str] = None,
    current_user: User = Depends(get_current_user)
):
    """Get payments, optionally filtered by invoice"""
    query = {}
    if invoice_id:
        query["invoice_id"] = invoice_id
    
    payments = await db.payments.find(query, {"_id": 0}).sort("created_at", -1).to_list(1000)
    for payment in payments:
        if isinstance(payment['created_at'], str):
            payment['created_at'] = datetime.fromisoformat(payment['created_at'])
    return payments



# ==================== PDF GENERATION ====================

from reportlab.lib.pagesizes import A4
from reportlab.lib.units import inch
from reportlab.pdfgen import canvas
from reportlab.lib import colors
from io import BytesIO

@api_router.get("/bills/{bill_id}/pdf")
async def generate_bill_pdf(bill_id: str, current_user: User = Depends(get_current_user)):
    """Generate PDF invoice"""
    # Fetch bill
    bill = await db.bills.find_one({"id": bill_id}, {"_id": 0})
    if not bill:
        raise HTTPException(status_code=404, detail="Bill not found")
    
    # Create PDF in memory
    buffer = BytesIO()
    pdf = canvas.Canvas(buffer, pagesize=A4)
    width, height = A4
    
    # Title
    pdf.setFont("Helvetica-Bold", 24)
    pdf.drawString(50, height - 50, "PharmaCare")
    pdf.setFont("Helvetica", 10)
    pdf.drawString(50, height - 70, "Pharmacy Management System")
    
    # Invoice Details
    pdf.setFont("Helvetica-Bold", 14)
    pdf.drawString(50, height - 110, bill['invoice_type'])
    pdf.setFont("Helvetica", 10)
    pdf.drawString(50, height - 130, f"Invoice No: {bill['bill_number']}")
    pdf.drawString(50, height - 145, f"Date: {bill.get('created_at', '')[:10]}")
    
    # Customer Details
    pdf.drawString(50, height - 175, f"Customer: {bill.get('customer_name', 'Counter Sale')}")
    if bill.get('customer_mobile'):
        pdf.drawString(50, height - 190, f"Mobile: {bill['customer_mobile']}")
    if bill.get('doctor_name'):
        pdf.drawString(50, height - 205, f"Doctor: {bill['doctor_name']}")
    
    # Items Table Header
    y = height - 250
    pdf.setFont("Helvetica-Bold", 10)
    pdf.drawString(50, y, "Item")
    pdf.drawString(250, y, "Batch")
    pdf.drawString(350, y, "Qty")
    pdf.drawString(400, y, "Price")
    pdf.drawString(480, y, "Total")
    
    # Items
    pdf.setFont("Helvetica", 9)
    y -= 20
    for item in bill['items']:
        name = item.get('product_name', item.get('medicine_name', 'Item'))[:25]
        batch = item.get('batch_no', item.get('batch_number', ''))[:15]
        qty = str(item['quantity'])
        price = f"₹{item.get('unit_price', item.get('mrp', 0))}"
        total = f"₹{item.get('line_total', item.get('total', 0)):.2f}"
        
        pdf.drawString(50, y, name)
        pdf.drawString(250, y, batch)
        pdf.drawString(350, y, qty)
        pdf.drawString(400, y, price)
        pdf.drawString(480, y, total)
        y -= 15
        
        if y < 100:  # New page if needed
            pdf.showPage()
            y = height - 50
    
    # Totals
    y -= 20
    pdf.setFont("Helvetica-Bold", 10)
    pdf.drawString(400, y, f"Subtotal: ₹{bill['subtotal']:.2f}")
    y -= 15
    pdf.drawString(400, y, f"Discount: -₹{bill['discount']:.2f}")
    y -= 15
    pdf.drawString(400, y, f"GST: ₹{bill['tax_amount']:.2f}")
    y -= 15
    pdf.setFont("Helvetica-Bold", 12)
    pdf.drawString(400, y, f"TOTAL: ₹{bill['total_amount']:.2f}")
    
    # Footer
    pdf.setFont("Helvetica", 8)
    pdf.drawString(50, 50, "Thank you for your business!")
    pdf.drawString(50, 35, f"Cashier: {bill.get('cashier_name', '')}")
    
    pdf.save()
    buffer.seek(0)
    
    return StreamingResponse(
        buffer,
        media_type="application/pdf",
        headers={"Content-Disposition": f"attachment; filename={bill['bill_number']}.pdf"}
    )


# ==================== REFUND ROUTES ====================

@api_router.post("/refunds", response_model=Refund)
async def create_refund(refund_data: RefundCreate, current_user: User = Depends(get_current_user)):
    """Record a refund for a sales return"""
    # Verify return invoice exists
    return_invoice = await db.bills.find_one({"id": refund_data.return_invoice_id}, {"_id": 0})
    if not return_invoice:
        raise HTTPException(status_code=404, detail="Return invoice not found")
    
    if return_invoice.get('invoice_type') != 'SALES_RETURN':
        raise HTTPException(status_code=400, detail="Invoice is not a sales return")
    
    refund = Refund(
        **refund_data.model_dump(),
        created_by=current_user.id
    )
    
    refund_doc = refund.model_dump()
    refund_doc['created_at'] = refund_doc['created_at'].isoformat()
    await db.refunds.insert_one(refund_doc)
    
    # Update return invoice status to refunded


# ==================== AUDIT LOG ROUTES ====================

@api_router.get("/audit-logs")
async def get_audit_logs(
    entity_type: Optional[str] = None,
    entity_id: Optional[str] = None,
    action: Optional[str] = None,
    limit: int = 100,
    current_user: User = Depends(get_current_user)
):
    """Get audit logs with optional filters"""
    query = {}
    if entity_type:
        query["entity_type"] = entity_type
    if entity_id:
        query["entity_id"] = entity_id
    if action:
        query["action"] = action
    
    logs = await db.audit_logs.find(query, {"_id": 0}).sort("created_at", -1).limit(limit).to_list(limit)
    
    for log in logs:
        if isinstance(log['created_at'], str):
            log['created_at'] = datetime.fromisoformat(log['created_at'])
    
    return logs

@api_router.get("/audit-logs/entity/{entity_type}/{entity_id}")
async def get_entity_audit_trail(
    entity_type: str,
    entity_id: str,
    current_user: User = Depends(get_current_user)
):
    """Get complete audit trail for a specific entity"""
    logs = await db.audit_logs.find(
        {
            "entity_type": entity_type,
            "entity_id": entity_id
        },
        {"_id": 0}
    ).sort("created_at", 1).to_list(1000)
    
    for log in logs:
        if isinstance(log['created_at'], str):
            log['created_at'] = datetime.fromisoformat(log['created_at'])
    
    return logs


    await db.bills.update_one(
        {"id": refund_data.return_invoice_id},
        {"$set": {"status": "refunded"}}
    )
    
    return refund

@api_router.get("/refunds")
async def get_refunds(
    return_invoice_id: Optional[str] = None,
    original_invoice_id: Optional[str] = None,
    current_user: User = Depends(get_current_user)
):
    """Get refunds, optionally filtered by invoice"""
    query = {}
    if return_invoice_id:
        query["return_invoice_id"] = return_invoice_id
    if original_invoice_id:
        query["original_invoice_id"] = original_invoice_id
    
    refunds = await db.refunds.find(query, {"_id": 0}).sort("created_at", -1).to_list(1000)
    for refund in refunds:
        if isinstance(refund['created_at'], str):
            refund['created_at'] = datetime.fromisoformat(refund['created_at'])
    return refunds


# ==================== PURCHASE ROUTES (MOVED TO END OF FILE) ====================

# ==================== CUSTOMER ROUTES ====================

@api_router.post("/customers", response_model=Customer)
async def create_customer(customer_data: CustomerCreate, current_user: User = Depends(get_current_user)):
    customer = Customer(**customer_data.model_dump())
    
    doc = customer.model_dump()
    doc['created_at'] = doc['created_at'].isoformat()
    await db.customers.insert_one(doc)
    
    return customer

@api_router.get("/customers")
async def get_customers(
    page: int = 1,
    page_size: int = 50,
    search: Optional[str] = None,
    customer_type: Optional[str] = None,
    fields: Optional[str] = None,
    current_user: User = Depends(get_current_user)
):
    """
    Get customers with optional pagination and field selection.
    - fields: comma-separated list of fields to return (e.g., "name,phone,email")
    - page/page_size: pagination controls
    - search: search by name or phone
    - customer_type: filter by type (regular, wholesale, institution)
    """
    query = {}
    if search:
        query["$or"] = [
            {"name": {"$regex": search, "$options": "i"}},
            {"phone": {"$regex": search, "$options": "i"}},
            {"email": {"$regex": search, "$options": "i"}}
        ]
    if customer_type:
        query["customer_type"] = customer_type
    
    # Get total count for pagination
    total = await db.customers.count_documents(query)
    
    # Use field selection if provided
    projection = parse_fields_param(fields)
    
    # Paginate
    skip = (page - 1) * page_size
    customers = await db.customers.find(query, projection).skip(skip).limit(page_size).to_list(page_size)
    
    for customer in customers:
        if 'created_at' in customer and isinstance(customer['created_at'], str):
            customer['created_at'] = datetime.fromisoformat(customer['created_at'])
    
    # Return paginated response if pagination params provided
    if page > 1 or page_size != 50:
        return paginate_response(customers, page, page_size, total)
    
    return customers

@api_router.get("/customers/search")
async def search_customers(q: str, current_user: User = Depends(get_current_user)):
    customers = await db.customers.find(
        {"$or": [
            {"name": {"$regex": q, "$options": "i"}},
            {"phone": {"$regex": q, "$options": "i"}}
        ]},
        {"_id": 0}
    ).to_list(100)
    return customers

@api_router.get("/customers/{customer_id}")
async def get_customer(customer_id: str, current_user: User = Depends(get_current_user)):
    customer = await db.customers.find_one({"id": customer_id}, {"_id": 0})
    if not customer:
        raise HTTPException(status_code=404, detail="Customer not found")
    return customer

@api_router.put("/customers/{customer_id}")
async def update_customer(customer_id: str, customer_data: dict, current_user: User = Depends(get_current_user)):
    # Update customer
    update_data = {k: v for k, v in customer_data.items() if v is not None}
    update_data['updated_at'] = datetime.now(timezone.utc).isoformat()
    
    result = await db.customers.update_one(
        {"id": customer_id},
        {"$set": update_data}
    )
    
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Customer not found")
    
    return {"message": "Customer updated successfully"}

@api_router.delete("/customers/{customer_id}")
async def delete_customer(customer_id: str, current_user: User = Depends(get_current_user)):
    result = await db.customers.delete_one({"id": customer_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Customer not found")
    return {"message": "Customer deleted successfully"}

@api_router.get("/customers/{customer_id}/stats")
async def get_customer_stats(customer_id: str, current_user: User = Depends(get_current_user)):
    """Get customer purchase statistics"""
    customer = await db.customers.find_one({"id": customer_id}, {"_id": 0})
    if not customer:
        raise HTTPException(status_code=404, detail="Customer not found")
    
    # Get bills for this customer
    bills = await db.bills.find(
        {"customer_name": customer['name'], "invoice_type": "SALE", "status": {"$in": ["paid", "due"]}},
        {"_id": 0, "total_amount": 1, "created_at": 1}
    ).to_list(10000)
    
    total_purchases = len(bills)
    total_value = sum(b.get('total_amount', 0) or 0 for b in bills)
    
    last_purchase = None
    if bills:
        dates = [b.get('created_at') for b in bills if b.get('created_at')]
        if dates:
            last_purchase = max(dates)
            if isinstance(last_purchase, str):
                last_purchase = datetime.fromisoformat(last_purchase).strftime('%d/%m/%Y')
            else:
                last_purchase = last_purchase.strftime('%d/%m/%Y')
    
    return {
        "total_purchases": total_purchases,
        "total_value": round(total_value, 2),
        "last_purchase": last_purchase
    }

# ==================== DOCTOR ROUTES ====================

@api_router.post("/doctors", response_model=Doctor)
async def create_doctor(doctor_data: DoctorCreate, current_user: User = Depends(get_current_user)):
    doctor = Doctor(**doctor_data.model_dump())
    
    doc = doctor.model_dump()
    doc['created_at'] = doc['created_at'].isoformat()
    await db.doctors.insert_one(doc)
    
    return doctor

@api_router.get("/doctors")
async def get_doctors(
    search: Optional[str] = None,
    page: int = 1,
    page_size: int = 50,
    current_user: User = Depends(get_current_user)
):
    """Get all doctors with pagination and search"""
    # Validate pagination params
    page_size = min(max(page_size, 1), 100)
    page = max(page, 1)
    
    query = {}
    if search:
        query["$or"] = [
            {"name": {"$regex": search, "$options": "i"}},
            {"specialization": {"$regex": search, "$options": "i"}},
            {"registration_number": {"$regex": search, "$options": "i"}},
            {"phone": {"$regex": search, "$options": "i"}}
        ]
    
    # Get total count
    total = await db.doctors.count_documents(query)
    
    # Paginate
    skip = (page - 1) * page_size
    doctors = await db.doctors.find(query, {"_id": 0}).sort("name", 1).skip(skip).limit(page_size).to_list(page_size)
    
    for doctor in doctors:
        if isinstance(doctor['created_at'], str):
            doctor['created_at'] = datetime.fromisoformat(doctor['created_at'])
    
    return {
        "data": doctors,
        "pagination": {
            "page": page,
            "page_size": page_size,
            "total": total,
            "total_pages": (total + page_size - 1) // page_size,
            "has_next": page * page_size < total,
            "has_prev": page > 1
        }
    }

@api_router.put("/doctors/{doctor_id}")
async def update_doctor(doctor_id: str, doctor_data: dict, current_user: User = Depends(get_current_user)):
    update_data = {k: v for k, v in doctor_data.items() if v is not None}
    update_data['updated_at'] = datetime.now(timezone.utc).isoformat()
    
    result = await db.doctors.update_one(
        {"id": doctor_id},
        {"$set": update_data}
    )
    
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Doctor not found")
    
    return {"message": "Doctor updated successfully"}

@api_router.delete("/doctors/{doctor_id}")
async def delete_doctor(doctor_id: str, current_user: User = Depends(get_current_user)):
    result = await db.doctors.delete_one({"id": doctor_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Doctor not found")
    return {"message": "Doctor deleted successfully"}

# ==================== REPORTS ====================

@api_router.get("/reports/sales-summary")
async def get_sales_summary(
    from_date: Optional[str] = None,
    to_date: Optional[str] = None,
    current_user: User = Depends(get_current_user)
):
    """Sales summary report with date range filter"""
    try:
        query = {"invoice_type": "SALE", "status": {"$in": ["paid", "due"]}}
        
        bills = await db.bills.find(query, {"_id": 0}).to_list(10000)
        
        # Filter by date range
        filtered_bills = []
        for bill in bills:
            try:
                created_at = bill.get('created_at')
                if isinstance(created_at, str):
                    created_at = datetime.fromisoformat(created_at)
                
                bill_date = created_at.strftime('%Y-%m-%d') if created_at else None
                
                if from_date and bill_date and bill_date < from_date:
                    continue
                if to_date and bill_date and bill_date > to_date:
                    continue
                
                filtered_bills.append({
                    "bill_number": bill.get('bill_number'),
                    "date": created_at.strftime('%d/%m/%Y') if created_at else 'N/A',
                    "customer_name": bill.get('customer_name') or 'Walk-in',
                    "items_count": len(bill.get('items', [])),
                    "payment_method": bill.get('payment_method', 'cash'),
                    "total_amount": bill.get('total_amount', 0)
                })
            except Exception as e:
                logger.warning(f"Error processing bill for report: {e}")
                continue
        
        total_sales = sum(b['total_amount'] for b in filtered_bills)
        
        return {
            "summary": {
                "total_bills": len(filtered_bills),
                "total_sales": round(total_sales, 2)
            },
            "data": filtered_bills
        }
    except Exception as e:
        logger.error(f"Sales report error: {e}")
        raise HTTPException(status_code=500, detail=str(e))

@api_router.get("/reports/low-stock")
async def get_low_stock_report(current_user: User = Depends(get_current_user)):
    """Report of items below reorder level"""
    try:
        products = await db.products.find({"status": "active"}, {"_id": 0}).to_list(10000)
        batches = await db.stock_batches.find({}, {"_id": 0}).to_list(10000)
        
        # Calculate total stock per product
        product_stock = {}
        for batch in batches:
            sku = batch.get('product_sku')
            if sku:
                if sku not in product_stock:
                    product_stock[sku] = 0
                product_stock[sku] += batch.get('qty_on_hand', 0)
        
        low_stock_items = []
        for product in products:
            sku = product.get('sku')
            current_stock = product_stock.get(sku, 0)
            reorder_level = product.get('low_stock_threshold_units', 10)
            
            if current_stock <= reorder_level:
                low_stock_items.append({
                    "product_name": product.get('name'),
                    "sku": sku,
                    "current_stock": current_stock,
                    "reorder_level": reorder_level,
                    "shortage": max(0, reorder_level - current_stock)
                })
        
        # Sort by shortage (most critical first)
        low_stock_items.sort(key=lambda x: (-x['shortage'], x['current_stock']))
        
        return {
            "summary": {
                "total_items": len(low_stock_items),
                "out_of_stock": sum(1 for i in low_stock_items if i['current_stock'] == 0)
            },
            "data": low_stock_items
        }
    except Exception as e:
        logger.error(f"Low stock report error: {e}")
        raise HTTPException(status_code=500, detail=str(e))

@api_router.get("/reports/expiry")
async def get_expiry_report(
    days: int = 30,
    current_user: User = Depends(get_current_user)
):
    """Report of items expiring within specified days"""
    try:
        now = datetime.now(timezone.utc)
        threshold_date = now + timedelta(days=days)
        
        batches = await db.stock_batches.find({"qty_on_hand": {"$gt": 0}}, {"_id": 0}).to_list(10000)
        products = await db.products.find({}, {"_id": 0}).to_list(10000)
        
        product_lookup = {p['sku']: p for p in products}
        
        expiring_items = []
        total_value = 0
        
        for batch in batches:
            try:
                expiry = batch.get('expiry_date')
                if not expiry:
                    continue
                
                if isinstance(expiry, str):
                    expiry = datetime.fromisoformat(expiry)
                if expiry.tzinfo is None:
                    expiry = expiry.replace(tzinfo=timezone.utc)
                
                if expiry <= threshold_date:
                    product_sku = batch.get('product_sku')
                    product = product_lookup.get(product_sku, {})
                    qty = batch.get('qty_on_hand', 0)
                    mrp = batch.get('mrp_per_unit', 0)
                    stock_value = qty * mrp
                    days_to_expiry = (expiry - now).days
                    
                    expiring_items.append({
                        "product_name": product.get('name', batch.get('product_name', 'Unknown')),
                        "batch_no": batch.get('batch_no'),
                        "qty": qty,
                        "expiry_date": expiry.strftime('%d/%m/%Y'),
                        "days_to_expiry": days_to_expiry,
                        "stock_value": round(stock_value, 2)
                    })
                    total_value += stock_value
            except Exception as e:
                logger.warning(f"Error processing batch for expiry report: {e}")
                continue
        
        # Sort by days to expiry (soonest first)
        expiring_items.sort(key=lambda x: x['days_to_expiry'])
        
        return {
            "summary": {
                "total_items": len(expiring_items),
                "total_value": round(total_value, 2),
                "expired": sum(1 for i in expiring_items if i['days_to_expiry'] < 0)
            },
            "data": expiring_items
        }
    except Exception as e:
        logger.error(f"Expiry report error: {e}")
        raise HTTPException(status_code=500, detail=str(e))

@api_router.get("/reports/dashboard")
async def get_dashboard_stats(current_user: User = Depends(get_current_user)):
    try:
        # Get today's sales
        today_start = datetime.now(timezone.utc).replace(hour=0, minute=0, second=0, microsecond=0)
        today_bills = await db.bills.find(
            {},
            {"_id": 0}
        ).to_list(10000)
        
        today_sales = 0
        total_sales = 0
        for bill in today_bills:
            try:
                created_at = bill['created_at']
                if isinstance(created_at, str):
                    created_at = datetime.fromisoformat(created_at)
                if created_at.tzinfo is None:
                    created_at = created_at.replace(tzinfo=timezone.utc)
                
                total_sales += bill.get('total_amount', 0)
                if created_at >= today_start:
                    today_sales += bill.get('total_amount', 0)
            except Exception as e:
                logger.warning(f"Error processing bill: {e}")
                continue
        
        # Get stock stats
        medicines = await db.medicines.find({}, {"_id": 0}).to_list(10000)
        total_medicines = len(medicines)
        low_stock_count = len([m for m in medicines if m.get('quantity', 0) < 10])
        
        thirty_days_later = datetime.now(timezone.utc) + timedelta(days=30)
        expiring_count = 0
        total_stock_value = 0
        
        for med in medicines:
            try:
                expiry = med.get('expiry_date')
                if expiry:
                    if isinstance(expiry, str):
                        expiry = datetime.fromisoformat(expiry)
                    if expiry.tzinfo is None:
                        expiry = expiry.replace(tzinfo=timezone.utc)
                    
                    if expiry <= thirty_days_later:
                        expiring_count += 1
                
                total_stock_value += med.get('quantity', 0) * med.get('purchase_rate', 0)
            except Exception as e:
                logger.warning(f"Error processing medicine: {e}")
                continue
        
        return {
            "today_sales": round(today_sales, 2),
            "total_sales": round(total_sales, 2),
            "total_medicines": total_medicines,
            "low_stock_count": low_stock_count,
            "expiring_soon_count": expiring_count,
            "total_stock_value": round(total_stock_value, 2)
        }
    except Exception as e:
        logger.error(f"Dashboard stats error: {e}")
        return {
            "today_sales": 0,
            "total_sales": 0,
            "total_medicines": 0,
            "low_stock_count": 0,
            "expiring_soon_count": 0,
            "total_stock_value": 0
        }

@api_router.get("/reports/sales")
async def get_sales_report(
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    current_user: User = Depends(get_current_user)
):
    query = {}
    bills = await db.bills.find(query, {"_id": 0}).to_list(10000)
    
    # Filter by date if provided
    if start_date:
        start = datetime.fromisoformat(start_date)
        bills = [b for b in bills if datetime.fromisoformat(b['created_at'] if isinstance(b['created_at'], str) else b['created_at'].isoformat()) >= start]
    
    if end_date:
        end = datetime.fromisoformat(end_date)
        bills = [b for b in bills if datetime.fromisoformat(b['created_at'] if isinstance(b['created_at'], str) else b['created_at'].isoformat()) <= end]
    
    total_sales = sum(b['total_amount'] for b in bills)
    total_tax = sum(b['tax_amount'] for b in bills)
    
    return {
        "bills": bills,
        "summary": {
            "total_bills": len(bills),
            "total_sales": round(total_sales, 2),
            "total_tax": round(total_tax, 2)
        }
    }

@api_router.get("/reports/gst")
async def get_gst_report(
    start_date: str,
    end_date: str,
    current_user: User = Depends(get_current_user)
):
    """Generate GST report for given date range"""
    start = datetime.fromisoformat(start_date)
    end = datetime.fromisoformat(end_date)
    
    # Sales GST (Output Tax)
    bills = await db.bills.find({
        "status": "paid",
        "created_at": {
            "$gte": start.isoformat(),
            "$lte": end.isoformat()
        }
    }, {"_id": 0}).to_list(10000)
    
    # Group sales by GST rate
    sales_by_gst = {}
    for bill in bills:
        for item in bill.get('items', []):
            gst_rate = item.get('gst_percent', 0)
            qty = item.get('quantity', 0)
            mrp = item.get('mrp', 0)
            
            taxable_amount = qty * mrp / (1 + gst_rate/100)
            gst_amount = taxable_amount * (gst_rate/100)
            
            if gst_rate not in sales_by_gst:
                sales_by_gst[gst_rate] = {
                    'gst_rate': gst_rate,
                    'taxable_amount': 0,
                    'cgst': 0,
                    'sgst': 0,
                    'igst': 0,
                    'total_gst': 0
                }
            
            sales_by_gst[gst_rate]['taxable_amount'] += taxable_amount
            sales_by_gst[gst_rate]['cgst'] += gst_amount / 2
            sales_by_gst[gst_rate]['sgst'] += gst_amount / 2
            sales_by_gst[gst_rate]['total_gst'] += gst_amount
    
    # Purchases GST (Input Tax Credit)
    purchases = await db.purchases.find({
        "status": "confirmed",
        "purchase_date": {
            "$gte": start.isoformat(),
            "$lte": end.isoformat()
        }
    }, {"_id": 0}).to_list(10000)
    
    # Group purchases by GST rate
    purchases_by_gst = {}
    for purchase in purchases:
        for item in purchase.get('items', []):
            gst_rate = item.get('gst_percent', 0)
            qty = item.get('qty_units', 0)
            cost = item.get('cost_price_per_unit', 0)
            
            taxable_amount = qty * cost / (1 + gst_rate/100)
            gst_amount = taxable_amount * (gst_rate/100)
            
            if gst_rate not in purchases_by_gst:
                purchases_by_gst[gst_rate] = {
                    'gst_rate': gst_rate,
                    'taxable_amount': 0,
                    'cgst': 0,
                    'sgst': 0,
                    'igst': 0,
                    'total_gst': 0
                }
            
            purchases_by_gst[gst_rate]['taxable_amount'] += taxable_amount
            purchases_by_gst[gst_rate]['cgst'] += gst_amount / 2
            purchases_by_gst[gst_rate]['sgst'] += gst_amount / 2
            purchases_by_gst[gst_rate]['total_gst'] += gst_amount
    
    # Calculate summaries
    sales_summary = {
        'total_taxable': sum(v['taxable_amount'] for v in sales_by_gst.values()),
        'total_cgst': sum(v['cgst'] for v in sales_by_gst.values()),
        'total_sgst': sum(v['sgst'] for v in sales_by_gst.values()),
        'total_igst': sum(v['igst'] for v in sales_by_gst.values()),
        'total_gst': sum(v['total_gst'] for v in sales_by_gst.values())
    }
    
    purchases_summary = {
        'total_taxable': sum(v['taxable_amount'] for v in purchases_by_gst.values()),
        'total_cgst': sum(v['cgst'] for v in purchases_by_gst.values()),
        'total_sgst': sum(v['sgst'] for v in purchases_by_gst.values()),
        'total_igst': sum(v['igst'] for v in purchases_by_gst.values()),
        'total_gst': sum(v['total_gst'] for v in purchases_by_gst.values())
    }
    
    net_liability = sales_summary['total_gst'] - purchases_summary['total_gst']
    
    return {
        "sales": list(sales_by_gst.values()),
        "purchases": list(purchases_by_gst.values()),
        "sales_summary": sales_summary,
        "purchases_summary": purchases_summary,
        "net_liability": round(net_liability, 2),
        "period": {
            "start_date": start_date,
            "end_date": end_date
        }
    }

# ==================== COMPLIANCE ====================

@api_router.get("/compliance/schedule-h1-register")
async def get_schedule_h1_register(
    from_date: Optional[str] = None,
    to_date: Optional[str] = None,
    current_user: User = Depends(get_current_user)
):
    """Get Schedule H1 drug dispensing register entries for compliance reporting.
    Restricted to admin and manager roles only."""
    
    # Check role - restrict to admin and manager only
    if current_user.role not in ["admin", "manager"]:
        raise HTTPException(
            status_code=403,
            detail="Access denied. Schedule H1 register is restricted to admin and manager roles."
        )
    
    # Build query with date filters
    query = {}
    if from_date or to_date:
        date_filter = {}
        if from_date:
            date_filter["$gte"] = from_date
        if to_date:
            date_filter["$lte"] = to_date
        if date_filter:
            query["dispensed_at"] = date_filter
    
    # Fetch entries sorted by dispensed_at descending
    entries = await db.schedule_h1_register.find(
        query,
        {"_id": 0}
    ).sort("dispensed_at", -1).to_list(10000)
    
    return {
        "entries": entries,
        "total_count": len(entries),
        "period": {
            "from_date": from_date,
            "to_date": to_date
        }
    }

# ==================== ANALYTICS ====================

@api_router.get("/analytics/summary")
async def get_analytics_summary(current_user: User = Depends(get_current_user)):
    try:
        # Get all bills (sales and returns)
        all_bills = await db.bills.find({}, {"_id": 0}).to_list(10000)
        
        gross_sales = 0
        returns = 0
        pending_amount = 0
        draft_count = 0
        today_sales = 0
        today_start = datetime.now(timezone.utc).replace(hour=0, minute=0, second=0, microsecond=0)
        
        for bill in all_bills:
            try:
                invoice_type = bill.get('invoice_type', 'SALE')
                status = bill.get('status', 'paid')
                amount = bill.get('total_amount', 0)
                
                created_at = bill['created_at']
                if isinstance(created_at, str):
                    created_at = datetime.fromisoformat(created_at)
                if created_at.tzinfo is None:
                    created_at = created_at.replace(tzinfo=timezone.utc)
                
                if invoice_type == 'SALE':
                    if status in ['paid', 'due']:
                        gross_sales += amount
                        if created_at >= today_start and status == 'paid':
                            today_sales += amount
                    if status == 'due':
                        pending_amount += amount
                    elif status == 'draft':
                        draft_count += 1
                elif invoice_type == 'SALES_RETURN':
                    if status in ['paid', 'refunded']:
                        returns += amount
                        
            except Exception as e:
                logger.warning(f"Error processing bill for analytics: {e}")
                continue
        
        net_sales = gross_sales - returns
        return_percentage = (returns / gross_sales * 100) if gross_sales > 0 else 0
        
        return {
            "gross_sales": round(gross_sales, 2),
            "returns": round(returns, 2),
            "net_sales": round(net_sales, 2),
            "return_percentage": round(return_percentage, 2),
            "pending_amount": round(pending_amount, 2),
            "today_sales": round(today_sales, 2),
            "draft_count": draft_count
        }
    except Exception as e:
        logger.error(f"Analytics summary error: {e}")
        return {
            "gross_sales": 0,
            "returns": 0,
            "net_sales": 0,
            "return_percentage": 0,
            "pending_amount": 0,
            "today_sales": 0,
            "draft_count": 0
        }

@api_router.get("/analytics/daily")
async def get_daily_analytics(days: int = 7, current_user: User = Depends(get_current_user)):
    try:
        start_date = datetime.now(timezone.utc) - timedelta(days=days)
        all_bills = await db.bills.find({}, {"_id": 0}).to_list(10000)
        
        # Group by date
        daily_data = {}
        
        for bill in all_bills:
            try:
                created_at = bill['created_at']
                if isinstance(created_at, str):
                    created_at = datetime.fromisoformat(created_at)
                if created_at.tzinfo is None:
                    created_at = created_at.replace(tzinfo=timezone.utc)
                
                if created_at < start_date:
                    continue
                
                date_key = created_at.strftime('%Y-%m-%d')
                if date_key not in daily_data:
                    daily_data[date_key] = {'sales': 0, 'returns': 0, 'net': 0}
                
                invoice_type = bill.get('invoice_type', 'SALE')
                amount = bill.get('total_amount', 0)
                
                if invoice_type == 'SALE' and bill.get('status') in ['paid', 'due']:
                    daily_data[date_key]['sales'] += amount
                elif invoice_type == 'SALES_RETURN' and bill.get('status') in ['paid', 'refunded']:
                    daily_data[date_key]['returns'] += amount
                
                daily_data[date_key]['net'] = daily_data[date_key]['sales'] - daily_data[date_key]['returns']
            except Exception as e:
                logger.warning(f"Error processing bill for daily analytics: {e}")
                continue
        
        # Convert to list
        result = []
        for date_str in sorted(daily_data.keys()):
            result.append({
                "date": date_str,
                "sales": round(daily_data[date_str]['sales'], 2),
                "returns": round(daily_data[date_str]['returns'], 2),
                "net": round(daily_data[date_str]['net'], 2)
            })
        
        return result
    except Exception as e:
        logger.error(f"Daily analytics error: {e}")
        return []

@api_router.get("/analytics/dashboard")
async def get_dashboard_analytics(current_user: User = Depends(get_current_user)):
    """Comprehensive dashboard analytics endpoint"""
    try:
        now = datetime.now(timezone.utc)
        today_start = now.replace(hour=0, minute=0, second=0, microsecond=0)
        week_start = today_start - timedelta(days=today_start.weekday())
        month_start = today_start.replace(day=1)
        yesterday_start = today_start - timedelta(days=1)
        last_week_start = week_start - timedelta(days=7)
        last_month_start = (month_start - timedelta(days=1)).replace(day=1)
        
        # Fetch all data in parallel
        all_bills = await db.bills.find({}, {"_id": 0}).to_list(10000)
        all_products = await db.products.find({}, {"_id": 0}).to_list(10000)
        all_batches = await db.stock_batches.find({}, {"_id": 0}).to_list(10000)
        all_customers = await db.customers.find({}, {"_id": 0}).to_list(10000)
        
        # Initialize metrics
        today_sales = 0
        yesterday_sales = 0
        week_sales = 0
        last_week_sales = 0
        month_sales = 0
        last_month_sales = 0
        total_sales = 0
        pending_payments = 0
        draft_bills = 0
        month_returns = 0
        
        # Sales by category and product
        category_sales = {}
        product_sales = {}
        customer_sales = {}
        
        # Daily trend (last 30 days)
        daily_sales = {}
        for i in range(30):
            date_key = (today_start - timedelta(days=i)).strftime('%Y-%m-%d')
            daily_sales[date_key] = {'sales': 0, 'returns': 0, 'bills': 0}
        
        # Recent bills
        recent_bills = []
        
        for bill in all_bills:
            try:
                created_at = bill.get('created_at')
                if not created_at:
                    continue
                if isinstance(created_at, str):
                    created_at = datetime.fromisoformat(created_at)
                if created_at.tzinfo is None:
                    created_at = created_at.replace(tzinfo=timezone.utc)
                
                invoice_type = bill.get('invoice_type', 'SALE')
                status = bill.get('status', 'paid')
                amount = bill.get('total_amount', 0) or 0
                
                # Count drafts and pending
                if status == 'draft':
                    draft_bills += 1
                    continue
                elif status == 'due':
                    pending_payments += amount
                
                # Process sales
                if invoice_type == 'SALE' and status in ['paid', 'due']:
                    total_sales += amount
                    
                    # Time-based sales
                    if created_at >= today_start:
                        today_sales += amount
                    if created_at >= yesterday_start and created_at < today_start:
                        yesterday_sales += amount
                    if created_at >= week_start:
                        week_sales += amount
                    if created_at >= last_week_start and created_at < week_start:
                        last_week_sales += amount
                    if created_at >= month_start:
                        month_sales += amount
                    if created_at >= last_month_start and created_at < month_start:
                        last_month_sales += amount
                    
                    # Daily trend
                    date_key = created_at.strftime('%Y-%m-%d')
                    if date_key in daily_sales:
                        daily_sales[date_key]['sales'] += amount
                        daily_sales[date_key]['bills'] += 1
                    
                    # Product and category sales
                    for item in bill.get('items', []):
                        product_sku = item.get('product_sku') or item.get('product_id', '')
                        product_name = item.get('product_name', 'Unknown')
                        line_total = item.get('line_total') or item.get('total', 0) or 0
                        
                        if product_sku not in product_sales:
                            product_sales[product_sku] = {'name': product_name, 'revenue': 0, 'qty': 0}
                        product_sales[product_sku]['revenue'] += line_total
                        product_sales[product_sku]['qty'] += item.get('quantity', 0)
                    
                    # Customer sales
                    customer_name = bill.get('customer_name', 'Walk-in')
                    if customer_name:
                        if customer_name not in customer_sales:
                            customer_sales[customer_name] = {'revenue': 0, 'bills': 0}
                        customer_sales[customer_name]['revenue'] += amount
                        customer_sales[customer_name]['bills'] += 1
                    
                    # Recent bills (last 10)
                    if len(recent_bills) < 10:
                        recent_bills.append({
                            'id': bill.get('id'),
                            'bill_number': bill.get('bill_number'),
                            'customer_name': bill.get('customer_name', 'Walk-in'),
                            'amount': amount,
                            'status': status,
                            'created_at': created_at.isoformat()
                        })
                
                # Process returns
                elif invoice_type == 'SALES_RETURN' and status in ['paid', 'refunded']:
                    if created_at >= month_start:
                        month_returns += amount
                    
                    date_key = created_at.strftime('%Y-%m-%d')
                    if date_key in daily_sales:
                        daily_sales[date_key]['returns'] += amount
                        
            except Exception as e:
                logger.warning(f"Dashboard analytics bill error: {e}")
                continue
        
        # Sort recent bills by date (newest first)
        recent_bills.sort(key=lambda x: x['created_at'], reverse=True)
        
        # Calculate percentage changes
        def calc_change(current, previous):
            if previous == 0:
                return 100 if current > 0 else 0
            return round((current - previous) / previous * 100, 1)
        
        today_change = calc_change(today_sales, yesterday_sales)
        week_change = calc_change(week_sales, last_week_sales)
        month_change = calc_change(month_sales, last_month_sales)
        
        # Top 5 products
        top_products = sorted(product_sales.items(), key=lambda x: x[1]['revenue'], reverse=True)[:5]
        top_products_list = [
            {'sku': sku, 'name': data['name'], 'revenue': round(data['revenue'], 2), 'qty': data['qty']}
            for sku, data in top_products
        ]
        
        # Top 5 customers
        top_customers = sorted(customer_sales.items(), key=lambda x: x[1]['revenue'], reverse=True)[:5]
        top_customers_list = [
            {'name': name, 'revenue': round(data['revenue'], 2), 'bills': data['bills']}
            for name, data in top_customers
        ]
        
        # Category-wise sales (from products)
        for product in all_products:
            category = product.get('category', 'Uncategorized') or 'Uncategorized'
            sku = product.get('sku', '')
            if sku in product_sales:
                if category not in category_sales:
                    category_sales[category] = 0
                category_sales[category] += product_sales[sku]['revenue']
        
        category_sales_list = [
            {'category': cat, 'revenue': round(rev, 2)}
            for cat, rev in sorted(category_sales.items(), key=lambda x: x[1], reverse=True)[:6]
        ]
        
        # Daily trend (sorted by date)
        daily_trend = [
            {'date': date, 'sales': round(data['sales'], 2), 'returns': round(data['returns'], 2), 'bills': data['bills']}
            for date, data in sorted(daily_sales.items())
        ]
        
        # Inventory stats
        total_products = len(all_products)
        total_stock_value = 0
        low_stock_items = []
        expiring_items = []
        thirty_days = now + timedelta(days=30)
        
        # Build product lookup
        product_lookup = {p['sku']: p for p in all_products}
        
        for batch in all_batches:
            try:
                qty = batch.get('qty_on_hand', 0)
                cost = batch.get('cost_price_per_unit', 0)
                total_stock_value += qty * cost
                
                product_sku = batch.get('product_sku', '')
                product = product_lookup.get(product_sku, {})
                product_name = product.get('name', batch.get('product_name', 'Unknown'))
                
                # Low stock (less than 10 units)
                if qty > 0 and qty < 10:
                    low_stock_items.append({
                        'product_name': product_name,
                        'batch_no': batch.get('batch_no', 'N/A'),
                        'qty': qty
                    })
                
                # Expiring soon
                expiry = batch.get('expiry_date')
                if expiry and qty > 0:
                    if isinstance(expiry, str):
                        expiry = datetime.fromisoformat(expiry)
                    if expiry.tzinfo is None:
                        expiry = expiry.replace(tzinfo=timezone.utc)
                    
                    if expiry <= thirty_days:
                        expiring_items.append({
                            'product_name': product_name,
                            'batch_no': batch.get('batch_no', 'N/A'),
                            'expiry_date': expiry.strftime('%Y-%m-%d'),
                            'qty': qty
                        })
            except Exception as e:
                logger.warning(f"Dashboard batch error: {e}")
                continue
        
        # Sort alerts
        low_stock_items.sort(key=lambda x: x['qty'])
        expiring_items.sort(key=lambda x: x['expiry_date'])
        
        return {
            # Key metrics
            "metrics": {
                "today_sales": round(today_sales, 2),
                "today_change": today_change,
                "week_sales": round(week_sales, 2),
                "week_change": week_change,
                "month_sales": round(month_sales, 2),
                "month_change": month_change,
                "total_sales": round(total_sales, 2)
            },
            # Charts data
            "daily_trend": daily_trend[-14:],  # Last 14 days
            "category_sales": category_sales_list,
            # Business insights
            "top_products": top_products_list,
            "top_customers": top_customers_list,
            # Alerts
            "low_stock": low_stock_items[:5],
            "expiring_soon": expiring_items[:5],
            "recent_bills": recent_bills[:5],
            # Quick stats
            "quick_stats": {
                "pending_payments": round(pending_payments, 2),
                "draft_bills": draft_bills,
                "month_returns": round(month_returns, 2),
                "total_products": total_products,
                "stock_value": round(total_stock_value, 2),
                "low_stock_count": len(low_stock_items),
                "expiring_count": len(expiring_items)
            }
        }
    except Exception as e:
        logger.error(f"Dashboard analytics error: {e}")
        raise HTTPException(status_code=500, detail=str(e))

# ==================== USER MANAGEMENT ====================

@api_router.get("/users", response_model=List[User])
async def get_users(current_user: User = Depends(get_current_user)):
    if current_user.role != "admin":
        raise HTTPException(status_code=403, detail="Only admins can view users")
    
    users = await db.users.find({}, {"_id": 0, "password_hash": 0}).to_list(1000)
    for user in users:
        if isinstance(user['created_at'], str):
            user['created_at'] = datetime.fromisoformat(user['created_at'])
    return users

@api_router.delete("/users/{user_id}")
async def delete_user(user_id: str, current_user: User = Depends(get_current_user)):
    if current_user.role != "admin":
        raise HTTPException(status_code=403, detail="Only admins can delete users")
    
    if user_id == current_user.id:
        raise HTTPException(status_code=400, detail="Cannot delete yourself")
    
    result = await db.users.delete_one({"id": user_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="User not found")
    
    return {"message": "User deleted successfully"}

# ==================== BACKUP ====================

@api_router.get("/backup/export")
async def export_data(current_user: User = Depends(get_current_user)):
    if current_user.role != "admin":
        raise HTTPException(status_code=403, detail="Only admins can export data")
    
    medicines = await db.medicines.find({}, {"_id": 0}).to_list(10000)
    bills = await db.bills.find({}, {"_id": 0}).to_list(10000)
    purchases = await db.purchases.find({}, {"_id": 0}).to_list(10000)
    customers = await db.customers.find({}, {"_id": 0}).to_list(10000)
    doctors = await db.doctors.find({}, {"_id": 0}).to_list(10000)
    suppliers = await db.suppliers.find({}, {"_id": 0}).to_list(10000)
    
    return {
        "export_date": datetime.now(timezone.utc).isoformat(),
        "medicines": medicines,
        "bills": bills,
        "purchases": purchases,
        "customers": customers,
        "doctors": doctors,
        "suppliers": suppliers
    }


# ==================== SUPPLIER ROUTES ====================

@api_router.get("/suppliers")
async def get_suppliers(
    search: Optional[str] = None,
    active_only: Optional[bool] = None,
    page: int = 1,
    page_size: int = 50,
    current_user: User = Depends(get_current_user)
):
    """Get all suppliers with pagination, search and active filter"""
    # Validate pagination params
    page_size = min(max(page_size, 1), 100)
    page = max(page, 1)
    
    query = {}
    if search:
        query["$or"] = [
            {"name": {"$regex": search, "$options": "i"}},
            {"contact_name": {"$regex": search, "$options": "i"}},
            {"phone": {"$regex": search, "$options": "i"}},
            {"gstin": {"$regex": search, "$options": "i"}}
        ]
    
    if active_only:
        query["is_active"] = {"$ne": False}
    
    # Get total count
    total = await db.suppliers.count_documents(query)
    
    # Paginate
    skip = (page - 1) * page_size
    suppliers = await db.suppliers.find(query, {"_id": 0}).sort("name", 1).skip(skip).limit(page_size).to_list(page_size)
    
    for supplier in suppliers:
        if isinstance(supplier.get('created_at'), str):
            supplier['created_at'] = datetime.fromisoformat(supplier['created_at'])
        if isinstance(supplier.get('updated_at'), str):
            supplier['updated_at'] = datetime.fromisoformat(supplier['updated_at'])
        if 'is_active' not in supplier:
            supplier['is_active'] = True
    
    return {
        "data": suppliers,
        "pagination": {
            "page": page,
            "page_size": page_size,
            "total": total,
            "total_pages": (total + page_size - 1) // page_size,
            "has_next": page * page_size < total,
            "has_prev": page > 1
        }
    }

@api_router.post("/suppliers", response_model=Supplier)
async def create_supplier(
    supplier_data: SupplierCreate,
    current_user: User = Depends(get_current_user)
):
    """Create a new supplier"""
    # Check if supplier with same name exists
    existing = await db.suppliers.find_one({"name": supplier_data.name}, {"_id": 0})
    if existing:
        raise HTTPException(status_code=400, detail="Supplier with this name already exists")
    
    supplier = Supplier(**supplier_data.model_dump())
    doc = supplier.model_dump()
    doc['created_at'] = doc['created_at'].isoformat()
    doc['updated_at'] = doc['updated_at'].isoformat()
    await db.suppliers.insert_one(doc)
    
    return supplier

@api_router.get("/suppliers/{supplier_id}", response_model=Supplier)
async def get_supplier(
    supplier_id: str,
    current_user: User = Depends(get_current_user)
):
    """Get supplier by ID"""
    supplier = await db.suppliers.find_one({"id": supplier_id}, {"_id": 0})
    if not supplier:
        raise HTTPException(status_code=404, detail="Supplier not found")
    
    if isinstance(supplier['created_at'], str):
        supplier['created_at'] = datetime.fromisoformat(supplier['created_at'])
    if isinstance(supplier['updated_at'], str):
        supplier['updated_at'] = datetime.fromisoformat(supplier['updated_at'])
    
    return supplier

@api_router.put("/suppliers/{supplier_id}")
async def update_supplier(
    supplier_id: str,
    supplier_data: SupplierUpdate,
    current_user: User = Depends(get_current_user)
):
    """Update supplier"""
    update_dict = {k: v for k, v in supplier_data.model_dump().items() if v is not None}
    if not update_dict:
        raise HTTPException(status_code=400, detail="No fields to update")
    
    update_dict['updated_at'] = datetime.now(timezone.utc).isoformat()
    
    result = await db.suppliers.update_one(
        {"id": supplier_id},
        {"$set": update_dict}
    )
    
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Supplier not found")
    
    return {"message": "Supplier updated successfully"}

@api_router.delete("/suppliers/{supplier_id}")
async def delete_supplier(
    supplier_id: str,
    current_user: User = Depends(get_current_user)
):
    """Delete supplier - BLOCKED if any purchases exist"""
    # VERIFIED – Data integrity check: cannot delete if transactions exist
    purchase_count = await db.purchases.count_documents({"supplier_id": supplier_id})
    if purchase_count > 0:
        raise HTTPException(
            status_code=400, 
            detail=f"Cannot delete supplier: {purchase_count} purchase(s) exist. Deactivate instead."
        )
    
    result = await db.suppliers.delete_one({"id": supplier_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Supplier not found")
    
    return {"message": "Supplier deleted successfully"}

@api_router.patch("/suppliers/{supplier_id}/toggle-status")
async def toggle_supplier_status(
    supplier_id: str,
    current_user: User = Depends(get_current_user)
):
    """Toggle supplier active/inactive status"""
    supplier = await db.suppliers.find_one({"id": supplier_id}, {"_id": 0})
    if not supplier:
        raise HTTPException(status_code=404, detail="Supplier not found")
    
    new_status = not supplier.get('is_active', True)
    
    await db.suppliers.update_one(
        {"id": supplier_id},
        {"$set": {"is_active": new_status, "updated_at": datetime.now(timezone.utc).isoformat()}}
    )
    
    status_text = "activated" if new_status else "deactivated"
    return {"message": f"Supplier {status_text} successfully", "is_active": new_status}

@api_router.get("/suppliers/{supplier_id}/summary")
async def get_supplier_summary(
    supplier_id: str,
    current_user: User = Depends(get_current_user)
):
    """Get supplier summary with purchase insights (read-only)"""
    supplier = await db.suppliers.find_one({"id": supplier_id}, {"_id": 0})
    if not supplier:
        raise HTTPException(status_code=404, detail="Supplier not found")
    
    # Get purchase statistics - VERIFIED: read-only, no inventory/ledger data
    purchases = await db.purchases.find(
        {"supplier_id": supplier_id, "status": {"$in": ["confirmed", "draft"]}},
        {"_id": 0, "purchase_date": 1, "total_value": 1, "status": 1}
    ).to_list(10000)
    
    total_purchases = len(purchases)
    total_value = sum(p.get('total_value', 0) or 0 for p in purchases)
    confirmed_purchases = [p for p in purchases if p.get('status') == 'confirmed']
    
    last_purchase_date = None
    if confirmed_purchases:
        dates = []
        for p in confirmed_purchases:
            pd = p.get('purchase_date')
            if pd:
                if isinstance(pd, str):
                    dates.append(pd)
                else:
                    dates.append(pd.isoformat())
        if dates:
            last_purchase_date = max(dates)
    
    return {
        "supplier": supplier,
        "total_purchases": total_purchases,
        "total_purchase_value": round(total_value, 2),
        "last_purchase_date": last_purchase_date
    }


# ==================== PURCHASE ROUTES ====================

async def generate_purchase_number():
    """Generate unique purchase number: PUR-2024-0001"""
    current_year = datetime.now(timezone.utc).year
    prefix = f"PUR-{current_year}-"
    
    # Find the last purchase number for this year
    last_purchase = await db.purchases.find_one(
        {"purchase_number": {"$regex": f"^{prefix}"}},
        {"_id": 0, "purchase_number": 1},
        sort=[("purchase_number", -1)]
    )
    
    if last_purchase:
        last_num = int(last_purchase['purchase_number'].split('-')[-1])
        new_num = last_num + 1
    else:
        new_num = 1
    
    return f"{prefix}{new_num:04d}"

@api_router.get("/purchases")
async def get_purchases(
    from_date: Optional[str] = None,
    to_date: Optional[str] = None,
    supplier_id: Optional[str] = None,
    status: Optional[str] = None,
    search: Optional[str] = None,
    page: int = 1,
    page_size: int = 50,
    current_user: User = Depends(get_current_user)
):
    """Get purchases with pagination and filters"""
    # Validate pagination params
    page_size = min(max(page_size, 1), 100)
    page = max(page, 1)
    
    query = {}
    
    if from_date:
        query["purchase_date"] = {"$gte": from_date}
    if to_date:
        if "purchase_date" in query:
            query["purchase_date"]["$lte"] = to_date
        else:
            query["purchase_date"] = {"$lte": to_date}
    
    if supplier_id:
        query["supplier_id"] = supplier_id
    
    if status:
        query["status"] = status
    
    if search:
        query["$or"] = [
            {"purchase_number": {"$regex": search, "$options": "i"}},
            {"supplier_name": {"$regex": search, "$options": "i"}},
            {"supplier_invoice_no": {"$regex": search, "$options": "i"}}
        ]
    
    # Get total count
    total = await db.purchases.count_documents(query)
    
    # Paginate
    skip = (page - 1) * page_size
    purchases = await db.purchases.find(query, {"_id": 0}).sort("purchase_date", -1).skip(skip).limit(page_size).to_list(page_size)
    
    return {
        "data": purchases,
        "pagination": {
            "page": page,
            "page_size": page_size,
            "total": total,
            "total_pages": (total + page_size - 1) // page_size,
            "has_next": page * page_size < total,
            "has_prev": page > 1
        }
    }

@api_router.post("/purchases")
async def create_purchase(
    purchase_data: PurchaseCreate,
    current_user: User = Depends(get_current_user)
):
    """Create new purchase draft or confirm purchase with stock updates"""
    # Get supplier
    supplier = await db.suppliers.find_one({"id": purchase_data.supplier_id}, {"_id": 0})
    if not supplier:
        raise HTTPException(status_code=404, detail="Supplier not found")
    
    # Generate purchase number
    purchase_number = await generate_purchase_number()
    
    # Process items and calculate totals
    items = []
    subtotal = 0
    tax_value = 0
    
    for item_data in purchase_data.items:
        # Get product to check units_per_pack
        product = await db.products.find_one({"sku": item_data.product_sku}, {"_id": 0})
        if not product:
            raise HTTPException(status_code=404, detail=f"Product {item_data.product_sku} not found")
        
        # Calculate line total (PTR is used if provided, else cost_price)
        ptr = item_data.ptr_per_unit if item_data.ptr_per_unit else item_data.cost_price_per_unit
        line_total = item_data.qty_units * ptr
        tax_amount = line_total * (item_data.gst_percent / 100) if purchase_data.with_gst else 0
        
        item_dict = {
            "id": str(uuid.uuid4()),
            "product_sku": item_data.product_sku,
            "product_name": item_data.product_name,
            "batch_no": item_data.batch_no,
            "expiry_date": item_data.expiry_date,  # Keep as string
            "qty_packs": item_data.qty_packs,
            "qty_units": item_data.qty_units,
            "free_qty_units": item_data.free_qty_units or 0,
            "cost_price_per_unit": item_data.cost_price_per_unit,
            "ptr_per_unit": ptr,
            "mrp_per_unit": item_data.mrp_per_unit,
            "gst_percent": item_data.gst_percent,
            "batch_priority": item_data.batch_priority or "LIFA",
            "line_total": line_total + tax_amount,
            "received_qty_units": 0
        }
        
        items.append(item_dict)
        subtotal += line_total
        tax_value += tax_amount
    
    total_value = subtotal + tax_value
    round_off = round(total_value) - total_value
    total_value = round(total_value)
    
    # Determine status and payment status
    status = purchase_data.status if purchase_data.status else "draft"
    payment_status = purchase_data.payment_status if purchase_data.payment_status else "unpaid"
    
    # If cash purchase and confirmed, mark as paid
    if purchase_data.purchase_on == "cash" and status == "confirmed":
        payment_status = "paid"
    
    # Calculate due date
    due_date = None
    if purchase_data.due_date:
        due_date = purchase_data.due_date
    elif purchase_data.purchase_on == "credit":
        # Auto-calculate due date based on payment terms
        from datetime import timedelta
        purchase_dt = datetime.fromisoformat(purchase_data.purchase_date)
        due_dt = purchase_dt + timedelta(days=supplier.get('payment_terms_days', 30))
        due_date = due_dt.isoformat()
    
    # Create purchase document
    purchase_doc = {
        "id": str(uuid.uuid4()),
        "purchase_number": purchase_number,
        "supplier_id": purchase_data.supplier_id,
        "supplier_name": supplier['name'],
        "purchase_date": purchase_data.purchase_date,
        "due_date": due_date,
        "supplier_invoice_no": purchase_data.supplier_invoice_no,
        "supplier_invoice_date": purchase_data.supplier_invoice_date,
        "order_type": purchase_data.order_type or "direct",
        "with_gst": purchase_data.with_gst,
        "purchase_on": purchase_data.purchase_on or "credit",
        "status": status,
        "payment_status": payment_status,
        "items": items,
        "subtotal": subtotal,
        "tax_value": tax_value,
        "round_off": round_off,
        "total_value": total_value,
        "amount_paid": total_value if payment_status == "paid" else 0,
        "payment_terms_days": supplier.get('payment_terms_days', 30),
        "note": purchase_data.note,
        "created_by": current_user.id,
        "created_at": datetime.now(timezone.utc).isoformat(),
        "updated_at": datetime.now(timezone.utc).isoformat()
    }
    
    await db.purchases.insert_one(purchase_doc)
    
    # If status is confirmed, create stock batches and update product LP
    if status == 'confirmed':
        for item in items:
            # Create stock batch with PTR and batch priority
            batch_doc = {
                "id": str(uuid.uuid4()),
                "product_sku": item['product_sku'],
                "batch_no": item['batch_no'] or f"PUR-{purchase_number[:8]}",
                "expiry_date": item['expiry_date'],
                "qty_on_hand": item['qty_units'] + item.get('free_qty_units', 0),
                "cost_price_per_unit": item['cost_price_per_unit'],
                "ptr_per_unit": item['ptr_per_unit'],
                "lp_per_unit": item['ptr_per_unit'],  # LP = PTR for v1
                "mrp_per_unit": item['mrp_per_unit'],
                "free_qty_units": item.get('free_qty_units', 0),
                "batch_priority": item.get('batch_priority', 'LIFA'),
                "supplier_name": supplier['name'],
                "supplier_invoice_no": purchase_data.supplier_invoice_no,
                "location": "default",
                "purchase_id": purchase_doc['id'],
                "created_by": current_user.id,
                "created_at": datetime.now(timezone.utc).isoformat()
            }
            await db.stock_batches.insert_one(batch_doc)
            
            # Update product LP (Landing Price) to PTR from this purchase
            await db.products.update_one(
                {"sku": item['product_sku']},
                {
                    "$set": {
                        "landing_price_per_unit": item['ptr_per_unit'],
                        "default_ptr_per_unit": item['ptr_per_unit'],
                        "updated_at": datetime.now(timezone.utc).isoformat()
                    }
                }
            )
            
            # Record stock movement
            movement_doc = {
                "id": str(uuid.uuid4()),
                "product_sku": item['product_sku'],
                "batch_id": batch_doc['id'],
                "movement_type": "purchase",
                "qty_delta_units": item['qty_units'] + item.get('free_qty_units', 0),
                "reason": f"Purchase {purchase_number}",
                "ref_id": purchase_doc['id'],
                "performed_by": current_user.id,
                "performed_at": datetime.now(timezone.utc).isoformat()
            }
            await db.stock_movements.insert_one(movement_doc)
        
        # If credit purchase, add to supplier outstanding
        if purchase_data.purchase_on == "credit" and payment_status != "paid":
            await db.suppliers.update_one(
                {"id": purchase_data.supplier_id},
                {
                    "$inc": {"outstanding": total_value},
                    "$set": {"updated_at": datetime.now(timezone.utc).isoformat()}
                }
            )
    
    # Create audit log
    await db.audit_logs.insert_one({
        "id": str(uuid.uuid4()),
        "entity_type": "purchase",
        "entity_id": purchase_doc['id'],
        "action": "create",
        "new_value": {"purchase_number": purchase_number, "status": status, "total_value": total_value, "payment_status": payment_status},
        "performed_by": current_user.id,
        "performed_by_name": current_user.name,
        "created_at": datetime.now(timezone.utc).isoformat()
    })
    
    # Return without _id
    purchase_doc.pop('_id', None)
    return purchase_doc

@api_router.put("/purchases/{purchase_id}")
async def update_purchase(
    purchase_id: str,
    purchase_data: PurchaseCreate,
    current_user: User = Depends(get_current_user)
):
    """Update a draft purchase - only drafts can be modified"""
    existing = await db.purchases.find_one({"id": purchase_id})
    
    if not existing:
        raise HTTPException(status_code=404, detail="Purchase not found")
    
    if existing.get('status') != 'draft':
        raise HTTPException(status_code=400, detail="Only draft purchases can be edited")
    
    # Get supplier
    supplier = await db.suppliers.find_one({"id": purchase_data.supplier_id}, {"_id": 0})
    if not supplier:
        raise HTTPException(status_code=404, detail="Supplier not found")
    
    # Process items and calculate totals
    items = []
    subtotal = 0
    tax_value = 0
    
    for item_data in purchase_data.items:
        product = await db.products.find_one({"sku": item_data.product_sku}, {"_id": 0})
        if not product:
            raise HTTPException(status_code=404, detail=f"Product {item_data.product_sku} not found")
        
        line_total = item_data.qty_units * item_data.cost_price_per_unit
        tax_amount = line_total * (item_data.gst_percent / 100)
        
        item = PurchaseItem(
            **item_data.model_dump(),
            line_total=line_total + tax_amount
        )
        
        if item_data.expiry_date:
            item.expiry_date = datetime.fromisoformat(item_data.expiry_date)
        
        items.append(item)
        subtotal += line_total
        tax_value += tax_amount
    
    total_value = subtotal + tax_value
    round_off = round(total_value) - total_value
    total_value = round(total_value)
    
    # Determine status
    status = purchase_data.status if purchase_data.status else "draft"
    
    # Update purchase
    update_data = {
        "supplier_id": purchase_data.supplier_id,
        "supplier_name": supplier['name'],
        "purchase_date": datetime.fromisoformat(purchase_data.purchase_date).isoformat(),
        "supplier_invoice_no": purchase_data.supplier_invoice_no,
        "supplier_invoice_date": datetime.fromisoformat(purchase_data.supplier_invoice_date).isoformat() if purchase_data.supplier_invoice_date else None,
        "items": [item.model_dump() for item in items],
        "subtotal": subtotal,
        "tax_value": tax_value,
        "round_off": round_off,
        "total_value": total_value,
        "status": status,
        "note": purchase_data.note,
        "updated_at": datetime.now(timezone.utc).isoformat(),
        "updated_by": current_user.id
    }
    
    # Convert item expiry dates
    for item in update_data['items']:
        if item.get('expiry_date'):
            item['expiry_date'] = item['expiry_date'].isoformat() if isinstance(item['expiry_date'], datetime) else item['expiry_date']
    
    await db.purchases.update_one({"id": purchase_id}, {"$set": update_data})
    
    # If status changed to confirmed, create stock batches
    if status == 'confirmed' and existing.get('status') == 'draft':
        for item in items:
            # Create stock batch
            batch = StockBatch(
                product_sku=item.product_sku,
                batch_no=item.batch_no or f"PUR-{existing.get('purchase_number', purchase_id)[:8]}",
                expiry_date=item.expiry_date.isoformat() if item.expiry_date else None,
                qty_on_hand=item.qty_units,
                cost_price_per_unit=item.cost_price_per_unit,
                mrp_per_unit=item.mrp_per_unit,
                location="default",
                purchase_id=purchase_id
            )
            batch_doc = batch.model_dump()
            await db.stock_batches.insert_one(batch_doc)
            
            # Record stock movement
            movement = StockMovement(
                product_sku=item.product_sku,
                batch_id=batch.id,
                movement_type="purchase",
                qty_delta_units=item.qty_units,
                reason=f"Purchase {existing.get('purchase_number', '')}",
                ref_id=purchase_id,
                performed_by=current_user.id
            )
            movement_doc = movement.model_dump()
            movement_doc['performed_at'] = movement_doc['performed_at'].isoformat()
            await db.stock_movements.insert_one(movement_doc)
    
    # Create audit log
    await db.audit_logs.insert_one({
        "id": str(uuid.uuid4()),
        "entity_type": "purchase",
        "entity_id": purchase_id,
        "action": "update",
        "new_value": {"status": status, "total_value": total_value},
        "performed_by": current_user.id,
        "performed_by_name": current_user.name,
        "created_at": datetime.now(timezone.utc).isoformat()
    })
    
    updated = await db.purchases.find_one({"id": purchase_id}, {"_id": 0})
    return updated

@api_router.get("/purchases/{purchase_id}")
async def get_purchase(
    purchase_id: str,
    current_user: User = Depends(get_current_user)
):
    """Get purchase by ID"""
    purchase = await db.purchases.find_one({"id": purchase_id}, {"_id": 0})
    if not purchase:
        raise HTTPException(status_code=404, detail="Purchase not found")
    
    # Return as-is without date conversion (dates are stored as ISO strings)
    return purchase


class PurchasePaymentRequest(BaseModel):
    amount: float
    payment_method: str = "cash"  # cash, bank_transfer, cheque, upi
    reference_no: Optional[str] = None
    notes: Optional[str] = None

@api_router.post("/purchases/{purchase_id}/pay")
async def mark_purchase_paid(
    purchase_id: str,
    payment: PurchasePaymentRequest,
    current_user: User = Depends(get_current_user)
):
    """Record payment for a purchase and update supplier outstanding"""
    purchase = await db.purchases.find_one({"id": purchase_id})
    if not purchase:
        raise HTTPException(status_code=404, detail="Purchase not found")
    
    if purchase.get('payment_status') == 'paid':
        raise HTTPException(status_code=400, detail="Purchase is already fully paid")
    
    total_value = purchase.get('total_value', 0)
    amount_paid = purchase.get('amount_paid', 0) + payment.amount
    
    # Determine new payment status
    if amount_paid >= total_value:
        payment_status = "paid"
        amount_paid = total_value  # Cap at total
    else:
        payment_status = "partial"
    
    # Update purchase
    await db.purchases.update_one(
        {"id": purchase_id},
        {
            "$set": {
                "payment_status": payment_status,
                "amount_paid": amount_paid,
                "updated_at": datetime.now(timezone.utc).isoformat()
            }
        }
    )
    
    # Update supplier outstanding (reduce by payment amount)
    supplier_id = purchase.get('supplier_id')
    if supplier_id:
        # Add payment to history
        payment_record = {
            "id": str(uuid.uuid4()),
            "amount": payment.amount,
            "payment_date": datetime.now(timezone.utc).isoformat(),
            "payment_method": payment.payment_method,
            "reference_no": payment.reference_no,
            "notes": payment.notes,
            "purchase_ids": [purchase_id],
            "created_by": current_user.id,
            "created_at": datetime.now(timezone.utc).isoformat()
        }
        
        await db.suppliers.update_one(
            {"id": supplier_id},
            {
                "$inc": {"outstanding": -payment.amount},
                "$push": {"payment_history": payment_record},
                "$set": {"updated_at": datetime.now(timezone.utc).isoformat()}
            }
        )
    
    # Create audit log
    await db.audit_logs.insert_one({
        "id": str(uuid.uuid4()),
        "entity_type": "purchase",
        "entity_id": purchase_id,
        "action": "payment",
        "new_value": {"amount": payment.amount, "payment_method": payment.payment_method, "payment_status": payment_status},
        "performed_by": current_user.id,
        "performed_by_name": current_user.name,
        "created_at": datetime.now(timezone.utc).isoformat()
    })
    
    updated = await db.purchases.find_one({"id": purchase_id}, {"_id": 0})
    return updated


# ==================== BILL SEQUENCE SETTINGS ROUTES ====================

@api_router.get("/settings/bill-sequence")
async def get_bill_sequence_settings_route(
    prefix: str = "INV",
    current_user: User = Depends(get_current_user)
):
    """Get bill sequence settings for a prefix"""
    if current_user.role not in ["admin", "manager"]:
        raise HTTPException(status_code=403, detail="Only admins and managers can view settings")
    
    settings = await get_bill_sequence_settings(prefix)
    return settings


@api_router.get("/settings/bill-sequences")
async def get_all_bill_sequences(
    current_user: User = Depends(get_current_user)
):
    """Get all bill sequence configurations"""
    if current_user.role not in ["admin", "manager"]:
        raise HTTPException(status_code=403, detail="Only admins and managers can view settings")
    
    sequences = await db.bill_number_sequences.find({}, {"_id": 0}).to_list(100)
    
    # If no sequences exist, return defaults
    if not sequences:
        return {
            "sequences": [
                {
                    "prefix": "INV",
                    "current_sequence": 0,
                    "sequence_length": 6,
                    "allow_prefix_change": True,
                    "next_number": 1,
                    "document_type": "Sales Invoice"
                },
                {
                    "prefix": "RTN",
                    "current_sequence": 0,
                    "sequence_length": 6,
                    "allow_prefix_change": True,
                    "next_number": 1,
                    "document_type": "Sales Return"
                }
            ]
        }
    
    # Add next_number and document_type to each sequence
    for seq in sequences:
        seq["next_number"] = seq.get("current_sequence", 0) + 1
        if seq.get("prefix") == "INV":
            seq["document_type"] = "Sales Invoice"
        elif seq.get("prefix") == "RTN":
            seq["document_type"] = "Sales Return"
        else:
            seq["document_type"] = "Custom"
    
    return {"sequences": sequences}


@api_router.put("/settings/bill-sequence")
async def update_bill_sequence_settings_route(
    settings: BillSequenceSettings,
    current_user: User = Depends(get_current_user)
):
    """Update bill sequence settings"""
    if current_user.role != "admin":
        raise HTTPException(status_code=403, detail="Only admins can modify bill sequence settings")
    
    # Validate prefix format
    if not settings.prefix or len(settings.prefix) > 10:
        raise HTTPException(status_code=400, detail="Prefix must be 1-10 characters")
    
    if not settings.prefix.isalnum():
        raise HTTPException(status_code=400, detail="Prefix must be alphanumeric")
    
    # Validate sequence length
    if settings.sequence_length < 3 or settings.sequence_length > 10:
        raise HTTPException(status_code=400, detail="Sequence length must be between 3 and 10")
    
    # Validate starting number
    if settings.starting_number < 1:
        raise HTTPException(status_code=400, detail="Starting number must be at least 1")
    
    # Update settings with validation
    result = await validate_and_update_sequence_settings(
        prefix=settings.prefix.upper(),
        starting_number=settings.starting_number,
        sequence_length=settings.sequence_length
    )
    
    return {
        "message": "Bill sequence settings updated successfully",
        "settings": result
    }


@api_router.post("/settings/bill-sequence/preview")
async def preview_bill_number(
    settings: BillSequenceSettings,
    current_user: User = Depends(get_current_user)
):
    """Preview what the next bill number would look like with given settings"""
    if current_user.role not in ["admin", "manager"]:
        raise HTTPException(status_code=403, detail="Only admins and managers can preview")
    
    # Format sample bill number
    sample = f"{settings.prefix.upper()}-{str(settings.starting_number).zfill(settings.sequence_length)}"
    
    return {
        "preview": sample,
        "format": f"{settings.prefix.upper()}-{'0' * settings.sequence_length}"
    }

app.add_middleware(
    CORSMiddleware,
    allow_credentials=True,
    allow_origins=os.environ.get('CORS_ORIGINS', '*').split(','),
    allow_methods=["*"],
    allow_headers=["*"],
)

logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s'
)


@app.on_event("startup")
async def startup_db():
    """Initialize database with default roles and create indexes"""
    # Check if roles collection is empty
    roles_count = await db.roles.count_documents({})
    
    if roles_count == 0:
        # Insert default roles
        for role_data in DEFAULT_ROLES:
            role = Role(**role_data)
            doc = role.model_dump()
            await db.roles.insert_one(doc)
        print("✅ Default roles initialized")
    
    # Create database indexes for optimization
    try:
        # Products collection indexes
        await db.products.create_index("name")
        await db.products.create_index("sku", unique=True)
        await db.products.create_index("barcode", sparse=True)  # For barcode scanning
        await db.products.create_index("brand")
        await db.products.create_index("category")
        await db.products.create_index("status")
        await db.products.create_index([("name", 1), ("brand", 1), ("sku", 1)])  # Composite for search
        
        # Stock batches collection indexes
        await db.stock_batches.create_index("product_sku")
        await db.stock_batches.create_index("expiry_date")
        await db.stock_batches.create_index("batch_no")
        await db.stock_batches.create_index([("product_sku", 1), ("expiry_date", 1)])
        await db.stock_batches.create_index([("qty_on_hand", 1)], partialFilterExpression={"qty_on_hand": {"$gt": 0}})
        
        # Bills collection indexes
        await db.bills.create_index("bill_number", unique=True)  # Unique constraint for bill numbers
        await db.bills.create_index("created_at")
        await db.bills.create_index("customer_name")
        await db.bills.create_index("invoice_type")
        await db.bills.create_index("status")
        await db.bills.create_index([("invoice_type", 1), ("status", 1), ("created_at", -1)])
        
        # Bill number sequences collection indexes
        await db.bill_number_sequences.create_index(
            [("prefix", 1), ("branch_id", 1)], 
            unique=True
        )  # Unique constraint for prefix + branch combination
        
        # Purchases collection indexes
        await db.purchases.create_index("purchase_number")
        await db.purchases.create_index("supplier_id")
        await db.purchases.create_index("purchase_date")
        await db.purchases.create_index("status")
        
        # Customers collection indexes
        await db.customers.create_index("name")
        await db.customers.create_index("phone", unique=True, sparse=True)
        await db.customers.create_index("email", sparse=True)
        
        # Suppliers collection indexes
        await db.suppliers.create_index("name")
        await db.suppliers.create_index("phone", sparse=True)
        await db.suppliers.create_index("is_active")
        
        # Audit logs index
        await db.audit_logs.create_index([("created_at", -1)])
        await db.audit_logs.create_index("entity_type")
        
        # Stock movements index
        await db.stock_movements.create_index([("performed_at", -1)])
        await db.stock_movements.create_index("product_sku")
        
        # Schedule H1 Register indexes for compliance queries
        await db.schedule_h1_register.create_index([("dispensed_at", -1)])
        await db.schedule_h1_register.create_index("product_sku")
        
        print("✅ Database indexes created successfully")
    except Exception as e:
        print(f"⚠️ Index creation warning (may already exist): {e}")

@app.on_event("shutdown")
async def shutdown_db_client():
    client.close()


# ==================== PURCHASE RETURN ROUTES ====================

async def generate_return_number():
    """Generate unique return number: PRET-2024-0001"""
    current_year = datetime.now(timezone.utc).year
    prefix = f"PRET-{current_year}-"
    
    last_return = await db.purchase_returns.find_one(
        {"return_number": {"$regex": f"^{prefix}"}},
        {"_id": 0, "return_number": 1},
        sort=[("return_number", -1)]
    )
    
    if last_return:
        last_num = int(last_return['return_number'].split('-')[-1])
        new_num = last_num + 1
    else:
        new_num = 1
    
    return f"{prefix}{new_num:04d}"

async def generate_credit_number():
    """Generate unique credit number: SCRED-2024-0001"""
    current_year = datetime.now(timezone.utc).year
    prefix = f"SCRED-{current_year}-"
    
    last_credit = await db.supplier_credits.find_one(
        {"credit_number": {"$regex": f"^{prefix}"}},
        {"_id": 0, "credit_number": 1},
        sort=[("credit_number", -1)]
    )
    
    if last_credit:
        last_num = int(last_credit['credit_number'].split('-')[-1])
        new_num = last_num + 1
    else:
        new_num = 1
    
    return f"{prefix}{new_num:04d}"

# Fix 1: Endpoint to load original purchase items for return
@api_router.get("/purchases/{purchase_id}/items-for-return")
async def get_purchase_items_for_return(
    purchase_id: str,
    current_user: User = Depends(get_current_user)
):
    """Get purchase items with already-returned quantities for creating a return"""
    purchase = await db.purchases.find_one({"id": purchase_id}, {"_id": 0})
    if not purchase:
        raise HTTPException(status_code=404, detail="Purchase not found")
    
    # Get all existing returns for this purchase
    existing_returns = await db.purchase_returns.find(
        {"purchase_id": purchase_id, "status": "confirmed"},
        {"_id": 0}
    ).to_list(100)
    
    # Calculate already returned qty per item (by product_sku + batch_no)
    returned_qtys = {}
    for ret in existing_returns:
        for item in ret.get('items', []):
            key = f"{item.get('product_sku')}_{item.get('batch_no')}"
            returned_qtys[key] = returned_qtys.get(key, 0) + (item.get('qty_units') or 0)
    
    # Build return items array
    items_for_return = []
    for item in purchase.get('items', []):
        key = f"{item.get('product_sku')}_{item.get('batch_no')}"
        already_returned = returned_qtys.get(key, 0)
        original_qty = item.get('qty_units') or item.get('quantity') or 0
        
        items_for_return.append({
            "medicine_id": item.get('product_id') or item.get('medicine_id'),
            "medicine_name": item.get('product_name') or item.get('medicine_name'),
            "product_sku": item.get('product_sku'),
            "batch_id": item.get('batch_id'),
            "batch_no": item.get('batch_no'),
            "expiry_date": item.get('expiry_date') or item.get('expiry_mmyy'),
            "mrp": item.get('mrp_per_unit') or item.get('mrp') or 0,
            "ptr": item.get('ptr_per_unit') or item.get('ptr') or 0,
            "gst_percent": item.get('gst_percent') or 5,
            "original_qty": original_qty,
            "already_returned_qty": already_returned,
            "max_returnable_qty": max(0, original_qty - already_returned)
        })
    
    return {
        "purchase_id": purchase_id,
        "purchase_number": purchase.get('purchase_number'),
        "supplier_id": purchase.get('supplier_id'),
        "supplier_name": purchase.get('supplier_name'),
        "purchase_date": purchase.get('purchase_date'),
        "invoice_no": purchase.get('supplier_invoice_no'),
        "items": items_for_return
    }


@api_router.post("/purchase-returns")
async def create_purchase_return(
    return_data: PurchaseReturnCreate,
    current_user: User = Depends(get_current_user)
):
    """Create purchase return - atomically confirms, deducts stock, updates supplier outstanding"""
    supplier = await db.suppliers.find_one({"id": return_data.supplier_id}, {"_id": 0})
    if not supplier:
        raise HTTPException(status_code=404, detail="Supplier not found")
    
    purchase_number = None
    original_purchase = None
    if return_data.purchase_id:
        original_purchase = await db.purchases.find_one({"id": return_data.purchase_id}, {"_id": 0})
        if original_purchase:
            purchase_number = original_purchase.get('purchase_number')
    
    # Fix 5: Validate qty against original purchase
    if original_purchase:
        # Get existing returns for this purchase
        existing_returns = await db.purchase_returns.find(
            {"purchase_id": return_data.purchase_id, "status": "confirmed"},
            {"_id": 0}
        ).to_list(100)
        
        # Calculate already returned qty per item
        returned_qtys = {}
        for ret in existing_returns:
            for item in ret.get('items', []):
                key = f"{item.get('product_sku')}_{item.get('batch_no')}"
                returned_qtys[key] = returned_qtys.get(key, 0) + (item.get('qty_units') or 0)
        
        # Build original qty map
        original_qtys = {}
        for item in original_purchase.get('items', []):
            key = f"{item.get('product_sku')}_{item.get('batch_no')}"
            original_qtys[key] = item.get('qty_units') or item.get('quantity') or 0
        
        # Validate each return item
        for item_data in return_data.items:
            qty_units = item_data.return_qty_units or item_data.qty_units or 0
            key = f"{item_data.product_sku}_{item_data.batch_no}"
            original_qty = original_qtys.get(key, 0)
            already_returned = returned_qtys.get(key, 0)
            max_returnable = original_qty - already_returned
            
            if qty_units > max_returnable:
                raise HTTPException(
                    status_code=400, 
                    detail=f"Return qty ({qty_units}) exceeds max returnable ({max_returnable}) for {item_data.product_name}"
                )
    
    return_number = await generate_return_number()
    
    items = []
    total_value = 0
    total_gst = 0
    stock_movements = []
    
    for item_data in return_data.items:
        # Get qty - handle both qty_units and return_qty_units
        qty_units = item_data.return_qty_units or item_data.qty_units or 0
        if qty_units <= 0:
            continue
        
        # Get additional fields from item_data
        ptr = getattr(item_data, 'ptr', None) or getattr(item_data, 'cost_price_per_unit', 0)
        mrp = getattr(item_data, 'mrp', None) or 0
        gst_percent = getattr(item_data, 'gst_percent', None) or 5
        expiry = getattr(item_data, 'expiry_date', None) or getattr(item_data, 'expiry', None)
        
        # Calculate line total (PTR * qty)
        line_total = qty_units * ptr
        line_gst = line_total * gst_percent / 100
        
        item_doc = {
            "id": str(uuid.uuid4()),
            "product_sku": item_data.product_sku,
            "product_name": item_data.product_name,
            "batch_id": item_data.batch_id,
            "batch_no": item_data.batch_no,
            "expiry_date": expiry,
            "mrp": mrp,
            "ptr": ptr,
            "gst_percent": gst_percent,
            "qty_units": qty_units,
            "cost_price_per_unit": ptr,
            "reason": item_data.reason or return_data.reason or "return",
            "line_total": line_total,
            "line_gst": line_gst
        }
        
        items.append(item_doc)
        total_value += line_total
        total_gst += line_gst
        
        # Deduct stock immediately (same as sales returns pattern)
        batch = None
        if item_data.batch_id:
            batch = await db.stock_batches.find_one({"id": item_data.batch_id}, {"_id": 0})
        elif item_data.batch_no:
            batch = await db.stock_batches.find_one({
                "product_sku": item_data.product_sku,
                "batch_no": item_data.batch_no
            }, {"_id": 0})
        
        if batch:
            product = await db.products.find_one({"sku": item_data.product_sku}, {"_id": 0})
            units_per_pack = product.get('units_per_pack', 1) if product else 1
            qty_packs = qty_units / units_per_pack
            
            new_qty = max(0, batch.get('qty_on_hand', 0) - qty_packs)
            
            await db.stock_batches.update_one(
                {"id": batch['id']},
                {"$set": {"qty_on_hand": new_qty, "updated_at": datetime.now(timezone.utc).isoformat(), "updated_by": current_user.id}}
            )
            
            movement = {
                "id": str(uuid.uuid4()),
                "product_sku": item_data.product_sku,
                "batch_id": batch['id'],
                "product_name": item_data.product_name,
                "batch_no": item_data.batch_no or batch['batch_no'],
                "qty_delta_units": -qty_units,
                "movement_type": "purchase_return",
                "ref_type": "purchase_return",
                "ref_id": return_number,
                "location": "default",
                "reason": f"Purchase return - {item_data.reason or 'return'}",
                "performed_by": current_user.id,
                "performed_at": datetime.now(timezone.utc).isoformat()
            }
            
            await db.stock_movements.insert_one(movement)
            stock_movements.append(movement)
    
    if not items:
        raise HTTPException(status_code=400, detail="No valid return items")
    
    net_return_amount = round(total_value + total_gst)
    
    # Create return document - status is confirmed immediately
    return_id = str(uuid.uuid4())
    return_doc = {
        "id": return_id,
        "return_number": return_number,
        "supplier_id": return_data.supplier_id,
        "supplier_name": supplier['name'],
        "purchase_id": return_data.purchase_id,
        "purchase_number": purchase_number,
        "return_date": return_data.return_date,
        "status": "confirmed",  # Fix 4: Use confirmed consistently
        "items": items,
        "ptr_total": total_value,
        "gst_amount": total_gst,
        "total_value": net_return_amount,
        "note": return_data.note or return_data.notes,
        "billed_by": getattr(return_data, 'billed_by', None) or current_user.name,
        "payment_type": getattr(return_data, 'payment_type', 'credit'),
        "created_by": current_user.id,
        "created_at": datetime.now(timezone.utc).isoformat(),
        "confirmed_at": datetime.now(timezone.utc).isoformat(),
        "confirmed_by": current_user.id
    }
    
    await db.purchase_returns.insert_one(return_doc)
    
    # Fix 2: Decrement supplier outstanding
    await db.suppliers.update_one(
        {"id": return_data.supplier_id},
        {
            "$inc": {"outstanding": -net_return_amount},
            "$push": {
                "payment_history": {
                    "id": str(uuid.uuid4()),
                    "type": "purchase_return",
                    "return_id": return_id,
                    "return_number": return_number,
                    "date": datetime.now(timezone.utc).isoformat(),
                    "amount": net_return_amount,
                    "note": f"Purchase return {return_number}"
                }
            }
        }
    )
    
    # Update original purchase with return reference
    if return_data.purchase_id:
        await db.purchases.update_one(
            {"id": return_data.purchase_id},
            {"$push": {"returns": return_id}}
        )
    
    # Remove _id before returning
    return_doc.pop('_id', None)
    return return_doc

@api_router.get("/purchase-returns")
async def get_purchase_returns(
    from_date: Optional[str] = None,
    to_date: Optional[str] = None,
    supplier_id: Optional[str] = None,
    status: Optional[str] = None,
    current_user: User = Depends(get_current_user)
):
    """Get purchase returns with filters"""
    query = {}
    
    if from_date:
        query["return_date"] = {"$gte": from_date}
    if to_date:
        if "return_date" in query:
            query["return_date"]["$lte"] = to_date
        else:
            query["return_date"] = {"$lte": to_date}
    
    if supplier_id:
        query["supplier_id"] = supplier_id
    
    if status:
        query["status"] = status
    
    returns = await db.purchase_returns.find(query, {"_id": 0}).sort("return_date", -1).to_list(1000)
    
    # Return as-is without date conversion
    return returns

@api_router.get("/purchase-returns/{return_id}")
async def get_purchase_return(
    return_id: str,
    current_user: User = Depends(get_current_user)
):
    """Get single purchase return by ID"""
    purchase_return = await db.purchase_returns.find_one({"id": return_id}, {"_id": 0})
    if not purchase_return:
        raise HTTPException(status_code=404, detail="Purchase return not found")
    
    # Return as-is without date conversion
    return purchase_return


# Fix 3: PUT endpoint for editing returns
class PurchaseReturnUpdate(BaseModel):
    note: Optional[str] = None
    billed_by: Optional[str] = None
    items: Optional[List[PurchaseReturnItemCreate]] = None
    edit_type: str = "non_financial"  # "non_financial" or "financial"


@api_router.put("/purchase-returns/{return_id}")
async def update_purchase_return(
    return_id: str,
    update_data: PurchaseReturnUpdate,
    current_user: User = Depends(get_current_user)
):
    """Update purchase return - supports both non-financial and financial edits"""
    purchase_return = await db.purchase_returns.find_one({"id": return_id}, {"_id": 0})
    if not purchase_return:
        raise HTTPException(status_code=404, detail="Purchase return not found")
    
    if update_data.edit_type == "non_financial":
        # Non-financial edit: only update note and billed_by
        update_fields = {"updated_at": datetime.now(timezone.utc).isoformat(), "updated_by": current_user.id}
        if update_data.note is not None:
            update_fields["note"] = update_data.note
        if update_data.billed_by is not None:
            update_fields["billed_by"] = update_data.billed_by
        
        await db.purchase_returns.update_one(
            {"id": return_id},
            {"$set": update_fields}
        )
        
        updated = await db.purchase_returns.find_one({"id": return_id}, {"_id": 0})
        return updated
    
    else:
        # Financial edit: recalculate stock differences
        if not update_data.items:
            raise HTTPException(status_code=400, detail="Items required for financial edit")
        
        old_items = purchase_return.get('items', [])
        old_total = purchase_return.get('total_value', 0)
        supplier_id = purchase_return.get('supplier_id')
        
        # Build old qty map
        old_qty_map = {}
        for item in old_items:
            key = f"{item.get('product_sku')}_{item.get('batch_no')}"
            old_qty_map[key] = item.get('qty_units', 0)
        
        # Process new items and calculate stock differences
        new_items = []
        new_total_value = 0
        new_total_gst = 0
        
        for item_data in update_data.items:
            qty_units = item_data.return_qty_units or item_data.qty_units or 0
            if qty_units <= 0:
                continue
            
            ptr = item_data.ptr or item_data.cost_price_per_unit or 0
            gst_percent = item_data.gst_percent or 5
            line_total = qty_units * ptr
            line_gst = line_total * gst_percent / 100
            
            item_doc = {
                "id": str(uuid.uuid4()),
                "product_sku": item_data.product_sku,
                "product_name": item_data.product_name,
                "batch_id": item_data.batch_id,
                "batch_no": item_data.batch_no,
                "expiry_date": item_data.expiry_date or item_data.expiry,
                "mrp": item_data.mrp or 0,
                "ptr": ptr,
                "gst_percent": gst_percent,
                "qty_units": qty_units,
                "cost_price_per_unit": ptr,
                "reason": item_data.reason or "return",
                "line_total": line_total,
                "line_gst": line_gst
            }
            new_items.append(item_doc)
            new_total_value += line_total
            new_total_gst += line_gst
            
            # Calculate stock difference
            key = f"{item_data.product_sku}_{item_data.batch_no}"
            old_qty = old_qty_map.get(key, 0)
            qty_diff = qty_units - old_qty  # Positive means more returned, negative means less
            
            if qty_diff != 0:
                # Find batch
                batch = None
                if item_data.batch_id:
                    batch = await db.stock_batches.find_one({"id": item_data.batch_id}, {"_id": 0})
                elif item_data.batch_no:
                    batch = await db.stock_batches.find_one({
                        "product_sku": item_data.product_sku,
                        "batch_no": item_data.batch_no
                    }, {"_id": 0})
                
                if batch:
                    product = await db.products.find_one({"sku": item_data.product_sku}, {"_id": 0})
                    units_per_pack = product.get('units_per_pack', 1) if product else 1
                    qty_packs_diff = qty_diff / units_per_pack
                    
                    # Adjust stock (positive diff = deduct more, negative diff = add back)
                    new_qty = max(0, batch.get('qty_on_hand', 0) - qty_packs_diff)
                    
                    await db.stock_batches.update_one(
                        {"id": batch['id']},
                        {"$set": {"qty_on_hand": new_qty, "updated_at": datetime.now(timezone.utc).isoformat()}}
                    )
                    
                    # Create stock movement for the difference
                    movement = {
                        "id": str(uuid.uuid4()),
                        "product_sku": item_data.product_sku,
                        "batch_id": batch['id'],
                        "product_name": item_data.product_name,
                        "batch_no": item_data.batch_no,
                        "qty_delta_units": -qty_diff,
                        "movement_type": "purchase_return_edit",
                        "ref_type": "purchase_return",
                        "ref_id": return_id,
                        "location": "default",
                        "reason": "Purchase return edit adjustment",
                        "performed_by": current_user.id,
                        "performed_at": datetime.now(timezone.utc).isoformat()
                    }
                    await db.stock_movements.insert_one(movement)
        
        new_net_amount = round(new_total_value + new_total_gst)
        amount_diff = new_net_amount - old_total
        
        # Update supplier outstanding with the difference
        if amount_diff != 0:
            await db.suppliers.update_one(
                {"id": supplier_id},
                {"$inc": {"outstanding": -amount_diff}}
            )
        
        # Update the return document
        await db.purchase_returns.update_one(
            {"id": return_id},
            {"$set": {
                "items": new_items,
                "ptr_total": new_total_value,
                "gst_amount": new_total_gst,
                "total_value": new_net_amount,
                "note": update_data.note or purchase_return.get('note'),
                "billed_by": update_data.billed_by or purchase_return.get('billed_by'),
                "updated_at": datetime.now(timezone.utc).isoformat(),
                "updated_by": current_user.id
            }}
        )
        
        updated = await db.purchase_returns.find_one({"id": return_id}, {"_id": 0})
        return updated


@api_router.post("/purchase-returns/{return_id}/confirm")
async def confirm_purchase_return(
    return_id: str,
    current_user: User = Depends(get_current_user)
):
    """Confirm purchase return: deduct stock, create movements, create supplier credit"""
    purchase_return = await db.purchase_returns.find_one({"id": return_id}, {"_id": 0})
    if not purchase_return:
        raise HTTPException(status_code=404, detail="Purchase return not found")
    
    # Allow confirming from both 'draft' and 'pending' status
    if purchase_return['status'] == 'confirmed':
        raise HTTPException(status_code=400, detail="Return is already confirmed")
    
    stock_movements = []
    
    for item in purchase_return['items']:
        qty_units = item.get('qty_units') or item.get('return_qty_units', 0)
        
        # Try to find batch by batch_id or by batch_no and product
        batch = None
        if item.get('batch_id'):
            batch = await db.stock_batches.find_one({"id": item['batch_id']}, {"_id": 0})
        elif item.get('batch_no'):
            batch = await db.stock_batches.find_one({
                "product_sku": item['product_sku'],
                "batch_no": item['batch_no']
            }, {"_id": 0})
        
        if not batch:
            # Find any batch for this product
            batch = await db.stock_batches.find_one({"product_sku": item['product_sku']}, {"_id": 0})
        
        if batch:
            product = await db.products.find_one({"sku": item['product_sku']}, {"_id": 0})
            units_per_pack = product.get('units_per_pack', 1) if product else 1
            qty_packs = qty_units / units_per_pack
            
            if batch['qty_on_hand'] >= qty_packs:
                new_qty = batch['qty_on_hand'] - qty_packs
                
                await db.stock_batches.update_one(
                    {"id": batch['id']},
                    {"$set": {"qty_on_hand": new_qty, "updated_at": datetime.now(timezone.utc).isoformat(), "updated_by": current_user.id}}
                )
                
                movement = {
                    "id": str(uuid.uuid4()),
                    "product_sku": item['product_sku'],
                    "batch_id": batch['id'],
                    "product_name": item['product_name'],
                    "batch_no": item.get('batch_no') or batch['batch_no'],
                    "qty_delta_units": -qty_units,
                    "movement_type": "purchase_return",
                    "ref_type": "purchase_return",
                    "ref_id": return_id,
                    "location": "default",
                    "reason": f"Purchase return - {item.get('reason', 'return')}",
                    "performed_by": current_user.id,
                    "performed_at": datetime.now(timezone.utc).isoformat()
                }
                
                await db.stock_movements.insert_one(movement)
                stock_movements.append(movement)
            else:
                logger.warning(f"Insufficient stock for return: {item['product_sku']}")
        else:
            logger.warning(f"No batch found for product: {item['product_sku']}")
    
    # Create supplier credit
    credit_number = await generate_credit_number()
    
    credit_doc = {
        "id": str(uuid.uuid4()),
        "supplier_id": purchase_return['supplier_id'],
        "supplier_name": purchase_return['supplier_name'],
        "credit_number": credit_number,
        "amount": purchase_return['total_value'],
        "reference": return_id,
        "reference_type": "purchase_return",
        "status": "active",
        "created_by": current_user.id,
        "created_at": datetime.now(timezone.utc).isoformat()
    }
    
    await db.supplier_credits.insert_one(credit_doc)
    
    await db.purchase_returns.update_one(
        {"id": return_id},
        {"$set": {"status": "confirmed", "confirmed_at": datetime.now(timezone.utc).isoformat(), "confirmed_by": current_user.id}}
    )
    
    return {
        "message": "Purchase return confirmed successfully",
        "credit_number": credit_number,
        "credit_amount": purchase_return['total_value'],
        "stock_movements_created": len(stock_movements)
    }

# ==================== SALES RETURNS ROUTES ====================

async def generate_credit_note_number() -> str:
    """Generate sequential credit note number CN-00001"""
    sequence_doc = await db.credit_note_sequences.find_one_and_update(
        {"type": "sales_return"},
        {"$inc": {"current_sequence": 1}},
        upsert=True,
        return_document=True
    )
    
    if not sequence_doc:
        # First time, create the sequence
        await db.credit_note_sequences.insert_one({
            "type": "sales_return",
            "current_sequence": 1,
            "prefix": "CN",
            "sequence_length": 5
        })
        return "CN-00001"
    
    seq = sequence_doc.get('current_sequence', 1)
    length = sequence_doc.get('sequence_length', 5)
    prefix = sequence_doc.get('prefix', 'CN')
    return f"{prefix}-{str(seq).zfill(length)}"

@api_router.post("/sales-returns")
async def create_sales_return(
    return_data: SalesReturnCreate,
    current_user: User = Depends(get_current_user)
):
    """Create a new sales return and update inventory"""
    
    # Check manual returns permission if no original bill
    if not return_data.original_bill_id:
        # Get user's role permissions
        role = await db.roles.find_one({"name": current_user.role}, {"_id": 0})
        allow_manual = role.get("allow_manual_returns", False) if role else False
        if not allow_manual and current_user.role != "admin":
            raise HTTPException(
                status_code=403, 
                detail="Manual returns require permission. Returns can only be created from an existing bill."
            )
    
    # Validate quantities against original bill if provided
    original_bill = None
    if return_data.original_bill_id:
        original_bill = await db.bills.find_one({"id": return_data.original_bill_id}, {"_id": 0})
        if not original_bill:
            raise HTTPException(status_code=404, detail="Original bill not found")
        
        # Validate return quantities
        original_items = {(item.get('batch_no') or item.get('batch_number')): item for item in original_bill.get('items', [])}
        for item in return_data.items:
            orig_item = original_items.get(item.batch_no)
            if orig_item:
                max_qty = orig_item.get('quantity', 0)
                if item.qty > max_qty:
                    raise HTTPException(
                        status_code=400,
                        detail=f"Return quantity for {item.medicine_name} ({item.qty}) exceeds original billed quantity ({max_qty})"
                    )
    
    # Generate credit note number
    return_no = await generate_credit_note_number()
    
    # Process items and calculate totals
    items = []
    mrp_total = 0
    total_discount = 0
    gst_amount = 0
    
    for item_data in return_data.items:
        # Calculate amounts
        base_amount = item_data.mrp * item_data.qty
        disc_amount = base_amount * (item_data.disc_percent / 100)
        after_disc = base_amount - disc_amount
        gst = after_disc * (item_data.gst_percent / 100)
        line_total = after_disc + gst
        
        item_doc = {
            "id": str(uuid.uuid4()),
            "medicine_id": item_data.medicine_id,
            "medicine_name": item_data.medicine_name,
            "product_sku": item_data.product_sku,
            "batch_id": item_data.batch_id,
            "batch_no": item_data.batch_no,
            "expiry_date": item_data.expiry_date,
            "mrp": item_data.mrp,
            "qty": item_data.qty,
            "original_qty": item_data.original_qty,
            "disc_percent": item_data.disc_percent,
            "disc_price": after_disc / item_data.qty if item_data.qty > 0 else 0,
            "gst_percent": item_data.gst_percent,
            "amount": line_total,
            "is_damaged": item_data.is_damaged
        }
        items.append(item_doc)
        
        mrp_total += base_amount
        total_discount += disc_amount
        gst_amount += gst
    
    net_amount = mrp_total - total_discount + gst_amount
    round_off = round(net_amount) - net_amount
    net_amount = round(net_amount)
    
    # Create return document
    return_date = datetime.fromisoformat(return_data.return_date.replace('Z', '+00:00')) if return_data.return_date else datetime.now(timezone.utc)
    
    return_doc = {
        "id": str(uuid.uuid4()),
        "return_no": return_no,
        "original_bill_id": return_data.original_bill_id,
        "original_bill_no": return_data.original_bill_no or (original_bill.get('bill_number') if original_bill else None),
        "return_date": return_date.isoformat(),
        "entry_date": datetime.now(timezone.utc).isoformat(),
        "patient": return_data.patient or {},
        "billing_for": return_data.billing_for,
        "doctor": return_data.doctor,
        "created_by": {"id": current_user.id, "name": current_user.name},
        "items": items,
        "mrp_total": mrp_total,
        "total_discount": total_discount,
        "gst_amount": gst_amount,
        "round_off": round_off,
        "net_amount": net_amount,
        "payment_type": return_data.payment_type or (original_bill.get('payment_method') if original_bill else None),
        "refund_method": return_data.refund_method,
        "note": return_data.note,
        "status": "completed",
        "credit_note_ref": return_no,
        "returns": []
    }
    
    await db.sales_returns.insert_one(return_doc)
    
    # Update inventory - increment stock for returned items
    for item in items:
        # Find the batch
        batch = None
        if item.get('batch_id'):
            batch = await db.stock_batches.find_one({"id": item['batch_id']}, {"_id": 0})
        elif item.get('product_sku') and item.get('batch_no'):
            batch = await db.stock_batches.find_one({
                "product_sku": item['product_sku'],
                "batch_no": item['batch_no']
            }, {"_id": 0})
        
        if batch:
            # Get product for units_per_pack
            product = await db.products.find_one({"sku": batch.get('product_sku')}, {"_id": 0})
            units_per_pack = product.get('units_per_pack', 1) if product else 1
            
            # Convert units to packs
            qty_packs = item['qty'] / units_per_pack
            
            if item.get('is_damaged'):
                # Increment damaged_stock instead of qty_on_hand
                await db.stock_batches.update_one(
                    {"id": batch['id']},
                    {"$inc": {"damaged_stock": qty_packs}}
                )
            else:
                # Increment regular stock
                await db.stock_batches.update_one(
                    {"id": batch['id']},
                    {"$inc": {"qty_on_hand": qty_packs}}
                )
            
            # Create stock movement record
            movement_doc = {
                "id": str(uuid.uuid4()),
                "product_sku": batch.get('product_sku'),
                "batch_id": batch['id'],
                "product_name": item['medicine_name'],
                "batch_no": item['batch_no'],
                "qty_delta_units": item['qty'],
                "movement_type": "sales_return",
                "ref_type": "sales_return",
                "ref_id": return_doc['id'],
                "location": "default",
                "reason": "Sales return" + (" (damaged)" if item.get('is_damaged') else ""),
                "performed_by": current_user.id,
                "performed_at": datetime.now(timezone.utc).isoformat()
            }
            await db.stock_movements.insert_one(movement_doc)
    
    # Update original bill with return reference
    if return_data.original_bill_id:
        await db.bills.update_one(
            {"id": return_data.original_bill_id},
            {"$push": {"returns": return_doc['id']}}
        )
    
    return_doc.pop('_id', None)
    return return_doc

@api_router.get("/sales-returns")
async def get_sales_returns(
    from_date: Optional[str] = None,
    to_date: Optional[str] = None,
    search: Optional[str] = None,
    payment_type: Optional[str] = None,
    page: int = 1,
    page_size: int = 50,
    current_user: User = Depends(get_current_user)
):
    """Get sales returns with filters"""
    query = {}
    
    if from_date:
        query["return_date"] = {"$gte": from_date}
    if to_date:
        if "return_date" in query:
            query["return_date"]["$lte"] = to_date
        else:
            query["return_date"] = {"$lte": to_date}
    
    if payment_type and payment_type != "all":
        query["refund_method"] = payment_type
    
    if search:
        query["$or"] = [
            {"return_no": {"$regex": search, "$options": "i"}},
            {"original_bill_no": {"$regex": search, "$options": "i"}},
            {"patient.name": {"$regex": search, "$options": "i"}},
            {"patient.phone": {"$regex": search, "$options": "i"}}
        ]
    
    total = await db.sales_returns.count_documents(query)
    skip = (page - 1) * page_size
    
    returns = await db.sales_returns.find(query, {"_id": 0}).sort("entry_date", -1).skip(skip).limit(page_size).to_list(page_size)
    
    # Calculate stats for today
    today_start = datetime.now(timezone.utc).replace(hour=0, minute=0, second=0, microsecond=0).isoformat()
    today_query = {"entry_date": {"$gte": today_start}}
    today_returns = await db.sales_returns.find(today_query, {"_id": 0, "net_amount": 1}).to_list(1000)
    returns_today = len(today_returns)
    total_refunded_today = sum(r.get('net_amount', 0) for r in today_returns)
    
    return {
        "data": returns,
        "total": total,
        "page": page,
        "page_size": page_size,
        "stats": {
            "returns_today": returns_today,
            "total_refunded_today": total_refunded_today
        }
    }

@api_router.get("/sales-returns/{return_id}")
async def get_sales_return(
    return_id: str,
    current_user: User = Depends(get_current_user)
):
    """Get single sales return by ID"""
    sales_return = await db.sales_returns.find_one({"id": return_id}, {"_id": 0})
    if not sales_return:
        # Try by return_no
        sales_return = await db.sales_returns.find_one({"return_no": return_id}, {"_id": 0})
    
    if not sales_return:
        raise HTTPException(status_code=404, detail="Sales return not found")
    
    return sales_return

@api_router.put("/sales-returns/{return_id}")
async def update_sales_return(
    return_id: str,
    update_data: SalesReturnUpdate,
    financial_edit: bool = False,
    current_user: User = Depends(get_current_user)
):
    """Update a sales return. Financial edits require permission."""
    
    existing = await db.sales_returns.find_one({"id": return_id}, {"_id": 0})
    if not existing:
        raise HTTPException(status_code=404, detail="Sales return not found")
    
    if financial_edit and update_data.items:
        # Check financial edit permission
        role = await db.roles.find_one({"name": current_user.role}, {"_id": 0})
        allow_financial = role.get("allow_financial_edit_return", False) if role else False
        if not allow_financial and current_user.role != "admin":
            raise HTTPException(
                status_code=403,
                detail="Financial edit requires permission"
            )
        
        # Reverse old stock changes
        for old_item in existing.get('items', []):
            batch = await db.stock_batches.find_one({"batch_no": old_item['batch_no']}, {"_id": 0})
            if batch:
                product = await db.products.find_one({"sku": batch.get('product_sku')}, {"_id": 0})
                units_per_pack = product.get('units_per_pack', 1) if product else 1
                qty_packs = old_item['qty'] / units_per_pack
                
                if old_item.get('is_damaged'):
                    await db.stock_batches.update_one(
                        {"id": batch['id']},
                        {"$inc": {"damaged_stock": -qty_packs}}
                    )
                else:
                    await db.stock_batches.update_one(
                        {"id": batch['id']},
                        {"$inc": {"qty_on_hand": -qty_packs}}
                    )
        
        # Process new items
        items = []
        mrp_total = 0
        total_discount = 0
        gst_amount = 0
        
        for item_data in update_data.items:
            base_amount = item_data.mrp * item_data.qty
            disc_amount = base_amount * (item_data.disc_percent / 100)
            after_disc = base_amount - disc_amount
            gst = after_disc * (item_data.gst_percent / 100)
            line_total = after_disc + gst
            
            item_doc = {
                "id": str(uuid.uuid4()),
                "medicine_id": item_data.medicine_id,
                "medicine_name": item_data.medicine_name,
                "product_sku": item_data.product_sku,
                "batch_id": item_data.batch_id,
                "batch_no": item_data.batch_no,
                "expiry_date": item_data.expiry_date,
                "mrp": item_data.mrp,
                "qty": item_data.qty,
                "original_qty": item_data.original_qty,
                "disc_percent": item_data.disc_percent,
                "disc_price": after_disc / item_data.qty if item_data.qty > 0 else 0,
                "gst_percent": item_data.gst_percent,
                "amount": line_total,
                "is_damaged": item_data.is_damaged
            }
            items.append(item_doc)
            
            mrp_total += base_amount
            total_discount += disc_amount
            gst_amount += gst
            
            # Apply new stock changes
            batch = None
            if item_data.batch_id:
                batch = await db.stock_batches.find_one({"id": item_data.batch_id}, {"_id": 0})
            elif item_data.product_sku and item_data.batch_no:
                batch = await db.stock_batches.find_one({
                    "product_sku": item_data.product_sku,
                    "batch_no": item_data.batch_no
                }, {"_id": 0})
            
            if batch:
                product = await db.products.find_one({"sku": batch.get('product_sku')}, {"_id": 0})
                units_per_pack = product.get('units_per_pack', 1) if product else 1
                qty_packs = item_data.qty / units_per_pack
                
                if item_data.is_damaged:
                    await db.stock_batches.update_one(
                        {"id": batch['id']},
                        {"$inc": {"damaged_stock": qty_packs}}
                    )
                else:
                    await db.stock_batches.update_one(
                        {"id": batch['id']},
                        {"$inc": {"qty_on_hand": qty_packs}}
                    )
        
        net_amount = mrp_total - total_discount + gst_amount
        round_off = round(net_amount) - net_amount
        net_amount = round(net_amount)
        
        update_dict = {
            "items": items,
            "mrp_total": mrp_total,
            "total_discount": total_discount,
            "gst_amount": gst_amount,
            "round_off": round_off,
            "net_amount": net_amount
        }
        
        if update_data.refund_method:
            update_dict["refund_method"] = update_data.refund_method
    else:
        # Non-financial edit
        update_dict = {}
        if update_data.billing_for is not None:
            update_dict["billing_for"] = update_data.billing_for
        if update_data.doctor is not None:
            update_dict["doctor"] = update_data.doctor
        if update_data.billed_by is not None:
            update_dict["created_by.name"] = update_data.billed_by
        if update_data.note is not None:
            update_dict["note"] = update_data.note
    
    if update_dict:
        await db.sales_returns.update_one({"id": return_id}, {"$set": update_dict})
    
    updated = await db.sales_returns.find_one({"id": return_id}, {"_id": 0})
    return updated

@api_router.get("/roles/{role_name}/permissions/returns")
async def get_role_return_permissions(
    role_name: str,
    current_user: User = Depends(get_current_user)
):
    """Get return-specific permissions for a role"""
    role = await db.roles.find_one({"name": role_name}, {"_id": 0})
    if not role:
        raise HTTPException(status_code=404, detail="Role not found")
    
    return {
        "allow_manual_returns": role.get("allow_manual_returns", False),
        "allow_financial_edit_return": role.get("allow_financial_edit_return", False)
    }

@api_router.put("/roles/{role_id}/permissions/returns")
async def update_role_return_permissions(
    role_id: str,
    allow_manual_returns: bool = False,
    allow_financial_edit_return: bool = False,
    current_user: User = Depends(get_current_user)
):
    """Update return-specific permissions for a role"""
    if current_user.role != "admin":
        raise HTTPException(status_code=403, detail="Only admins can update permissions")
    
    result = await db.roles.update_one(
        {"id": role_id},
        {"$set": {
            "allow_manual_returns": allow_manual_returns,
            "allow_financial_edit_return": allow_financial_edit_return
        }}
    )
    
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Role not found")
    
    return {"message": "Permissions updated successfully"}

@api_router.get("/analytics/purchases")
async def get_purchase_analytics(
    from_date: Optional[str] = None,
    to_date: Optional[str] = None,
    current_user: User = Depends(get_current_user)
):
    """Get purchase analytics"""
    query = {}
    if from_date:
        query["purchase_date"] = {"$gte": from_date}
    if to_date:
        if "purchase_date" in query:
            query["purchase_date"]["$lte"] = to_date
        else:
            query["purchase_date"] = {"$lte": to_date}
    
    query["status"] = {"$nin": ["cancelled", "draft"]}
    purchases = await db.purchases.find(query, {"_id": 0}).to_list(10000)
    total_purchases_value = sum(p.get('total_value', 0) for p in purchases)
    
    return_query = {}
    if from_date:
        return_query["return_date"] = {"$gte": from_date}
    if to_date:
        if "return_date" in return_query:
            return_query["return_date"]["$lte"] = to_date
        else:
            return_query["return_date"] = {"$lte": to_date}
    
    return_query["status"] = "confirmed"
    purchase_returns = await db.purchase_returns.find(return_query, {"_id": 0}).to_list(10000)
    total_purchase_returns_value = sum(r.get('total_value', 0) for r in purchase_returns)
    
    net_purchases = total_purchases_value - total_purchase_returns_value
    
    return {
        "total_purchases_value": total_purchases_value,
        "total_purchase_returns_value": total_purchase_returns_value,
        "net_purchases": net_purchases,
        "total_purchases_count": len(purchases),
        "total_returns_count": len(purchase_returns)
    }

# ==================== EXCEL BULK UPLOAD ====================

# In-memory job storage for bulk upload progress
bulk_upload_jobs: Dict[str, Dict] = {}

# Column mapping keywords for auto-detection
COLUMN_KEYWORDS = {
    "sku": ["sku", "code", "product_code", "item_code", "product code", "item code"],
    "name": ["name", "product_name", "item_name", "medicine", "product name", "medicine name", "item name", "product", "item", "description"],
    "price": ["price", "mrp", "rate", "selling_price", "selling price", "mrp_per_unit", "mrp per unit", "unit_price", "unit price"],
    "quantity": ["quantity", "qty", "stock", "qty_on_hand", "qty on hand", "packs", "units", "opening_qty", "opening qty"],
    "expiry_date": ["expiry", "expiry_date", "exp_date", "exp", "expiry date", "exp date", "expires", "expire"],
    "batch_number": ["batch", "batch_number", "batch_no", "batch no", "lot", "lot_number", "lot number", "batch number"],
    "brand": ["brand", "manufacturer", "company", "mfr"],
    "category": ["category", "type", "group", "class"],
    "cost_price": ["cost", "cost_price", "purchase_price", "purchase price", "cost price", "buy_price", "buy price", "ptr"],
    "gst_percent": ["gst", "gst_percent", "tax", "gst percent", "tax_rate", "tax rate"],
    "hsn_code": ["hsn", "hsn_code", "hsn code"],
    "units_per_pack": ["units_per_pack", "pack_size", "units per pack", "pack size", "strip_qty", "strip qty"]
}

REQUIRED_FIELDS = ["sku", "name", "price", "quantity", "expiry_date", "batch_number"]
OPTIONAL_FIELDS = ["brand", "category", "cost_price", "gst_percent", "hsn_code", "units_per_pack"]

class BulkUploadMapping(BaseModel):
    file_column: str
    system_field: str

class BulkUploadValidateRequest(BaseModel):
    job_id: str
    column_mapping: Dict[str, str]  # system_field -> file_column

class BulkUploadImportRequest(BaseModel):
    job_id: str
    import_valid_only: bool = True

@api_router.get("/inventory/bulk-upload/template")
async def download_bulk_upload_template(current_user: User = Depends(get_current_user)):
    """Download sample Excel template for bulk inventory upload"""
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Inventory Template"
    
    # Header styling
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    required_fill = PatternFill(start_color="FF6B6B", end_color="FF6B6B", fill_type="solid")
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    
    # Define headers
    headers = [
        ("SKU *", True), ("Name *", True), ("Brand", False), ("Category", False),
        ("Batch Number *", True), ("Expiry Date *", True), ("Quantity (Packs) *", True),
        ("MRP per Unit *", True), ("Cost Price per Unit", False), ("GST %", False),
        ("HSN Code", False), ("Units per Pack", False)
    ]
    
    # Write headers
    for col, (header, is_required) in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = required_fill if is_required else header_fill
        cell.alignment = Alignment(horizontal="center")
        cell.border = thin_border
        ws.column_dimensions[ws.cell(row=1, column=col).column_letter].width = 18
    
    # Add sample data rows
    sample_data = [
        ["MED001", "Paracetamol 500mg", "Cipla", "Tablets", "BTN2024001", "2025-12-31", 100, 2.50, 1.80, 12, "30049099", 10],
        ["MED002", "Amoxicillin 250mg", "Sun Pharma", "Antibiotics", "BTN2024002", "2025-06-30", 50, 8.00, 6.50, 12, "30042011", 10],
        ["MED003", "Vitamin D3 Capsules", "Abbott", "Vitamins", "BTN2024003", "2026-03-15", 200, 5.50, 4.20, 5, "21069099", 15],
    ]
    
    for row_num, row_data in enumerate(sample_data, 2):
        for col_num, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_num, column=col_num, value=value)
            cell.border = thin_border
    
    # Add instructions sheet
    ws_instructions = wb.create_sheet("Instructions")
    instructions = [
        ["Excel Bulk Upload Instructions"],
        [""],
        ["Required Fields (marked with *)"],
        ["- SKU: Unique product identifier"],
        ["- Name: Product/medicine name"],
        ["- Batch Number: Batch identifier"],
        ["- Expiry Date: Format YYYY-MM-DD"],
        ["- Quantity: Number of packs"],
        ["- MRP per Unit: Maximum retail price per unit"],
        [""],
        ["Optional Fields"],
        ["- Brand: Manufacturer or brand name"],
        ["- Category: Product category"],
        ["- Cost Price: Purchase price per unit"],
        ["- GST %: Tax percentage (default 5%)"],
        ["- HSN Code: Harmonized System code"],
        ["- Units per Pack: Number of units in a pack (default 1)"],
        [""],
        ["Important Notes:"],
        ["- Dates must be in YYYY-MM-DD format (e.g., 2025-12-31)"],
        ["- Expiry date must be in the future"],
        ["- MRP must be greater than Cost Price"],
        ["- Duplicate SKU+Batch combinations will update existing batches"],
        ["- Maximum 5000 rows per upload"],
    ]
    
    for row_num, row in enumerate(instructions, 1):
        cell = ws_instructions.cell(row=row_num, column=1, value=row[0] if row else "")
        if row_num == 1:
            cell.font = Font(bold=True, size=14)
    
    ws_instructions.column_dimensions['A'].width = 60
    
    # Save to bytes
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=inventory_upload_template.xlsx"}
    )

@api_router.post("/inventory/bulk-upload/parse")
async def parse_bulk_upload_file(
    file: UploadFile = File(...),
    current_user: User = Depends(get_current_user)
):
    """Parse uploaded Excel file and return columns with auto-detected mappings"""
    
    # Validate file type
    if not file.filename.endswith(('.xlsx', '.xls')):
        raise HTTPException(status_code=400, detail="Only Excel files (.xlsx, .xls) are supported")
    
    try:
        contents = await file.read()
        
        # Read Excel file
        if file.filename.endswith('.xlsx'):
            df = pd.read_excel(io.BytesIO(contents), engine='openpyxl')
        else:
            df = pd.read_excel(io.BytesIO(contents), engine='xlrd')
        
        # Validate row count
        if len(df) > 5000:
            raise HTTPException(status_code=400, detail=f"File has {len(df)} rows. Maximum allowed is 5000 rows.")
        
        if len(df) == 0:
            raise HTTPException(status_code=400, detail="File is empty or has no data rows")
        
        # Get column names
        file_columns = list(df.columns.astype(str))
        
        # Auto-detect column mappings
        auto_mappings = {}
        for sys_field, keywords in COLUMN_KEYWORDS.items():
            for col in file_columns:
                col_lower = str(col).lower().strip()
                if col_lower in keywords or any(kw in col_lower for kw in keywords):
                    if sys_field not in auto_mappings:
                        auto_mappings[sys_field] = col
                    break
        
        # Create job ID and store data
        job_id = str(uuid.uuid4())
        
        # Store in memory (for simplicity; in production, use Redis or database)
        bulk_upload_jobs[job_id] = {
            "status": "parsed",
            "filename": file.filename,
            "total_rows": len(df),
            "columns": file_columns,
            "data": df.to_dict(orient='records'),
            "created_at": datetime.now(timezone.utc).isoformat(),
            "created_by": current_user.id
        }
        
        # Get sample data (first 5 rows)
        sample_data = df.head(5).fillna('').to_dict(orient='records')
        
        return {
            "job_id": job_id,
            "filename": file.filename,
            "total_rows": len(df),
            "columns": file_columns,
            "auto_mappings": auto_mappings,
            "sample_data": sample_data,
            "required_fields": REQUIRED_FIELDS,
            "optional_fields": OPTIONAL_FIELDS
        }
        
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Excel parse error: {e}")
        raise HTTPException(status_code=400, detail=f"Failed to parse Excel file: {str(e)}")

@api_router.post("/inventory/bulk-upload/validate")
async def validate_bulk_upload(
    request: BulkUploadValidateRequest,
    current_user: User = Depends(get_current_user)
):
    """Validate mapped data and return validation results"""
    
    job_id = request.job_id
    column_mapping = request.column_mapping
    
    if job_id not in bulk_upload_jobs:
        raise HTTPException(status_code=404, detail="Upload job not found. Please re-upload the file.")
    
    job = bulk_upload_jobs[job_id]
    data = job["data"]
    
    # Check required field mappings
    missing_required = []
    for field in REQUIRED_FIELDS:
        if field not in column_mapping or not column_mapping[field]:
            missing_required.append(field)
    
    if missing_required:
        raise HTTPException(status_code=400, detail=f"Missing required field mappings: {', '.join(missing_required)}")
    
    # Get existing products and batches for duplicate checking
    existing_products = {}
    products_cursor = db.products.find({}, {"_id": 0})
    async for prod in products_cursor:
        if "sku" in prod:
            existing_products[prod["sku"]] = prod
    
    existing_batches = {}
    batches_cursor = db.stock_batches.find({}, {"_id": 0})
    async for batch in batches_cursor:
        if "product_sku" in batch and "batch_no" in batch:
            key = f"{batch['product_sku']}_{batch['batch_no']}"
            existing_batches[key] = batch
    
    # Validate each row
    validation_results = []
    valid_count = 0
    error_count = 0
    warning_count = 0
    today = datetime.now(timezone.utc).date()
    
    for row_idx, row in enumerate(data, start=2):  # Start from 2 to match Excel row numbers
        row_errors = []
        row_warnings = []
        row_data = {}
        
        # Extract and validate each field
        for sys_field, file_col in column_mapping.items():
            value = row.get(file_col, None)
            
            # Handle NaN/empty values
            if pd.isna(value) or value == '' or value is None:
                value = None
            else:
                value = str(value).strip() if not isinstance(value, (int, float)) else value
            
            row_data[sys_field] = value
        
        # Required field validation
        for field in REQUIRED_FIELDS:
            if row_data.get(field) is None or row_data.get(field) == '':
                row_errors.append(f"Missing required field: {field}")
        
        # SKU validation
        sku = row_data.get("sku")
        if sku:
            sku = str(sku).strip().upper()
            row_data["sku"] = sku
        
        # Name validation
        name = row_data.get("name")
        if name and len(str(name)) < 2:
            row_errors.append("Product name must be at least 2 characters")
        
        # Price validation
        try:
            price = float(row_data.get("price", 0) or 0)
            row_data["price"] = price
            if price <= 0:
                row_errors.append("MRP must be greater than 0")
        except (ValueError, TypeError):
            row_errors.append("Invalid MRP value")
            price = 0
        
        # Cost price validation
        try:
            cost_price = float(row_data.get("cost_price", 0) or 0)
            row_data["cost_price"] = cost_price
            if cost_price > 0 and price > 0 and cost_price >= price:
                row_warnings.append("Cost price should be less than MRP")
        except (ValueError, TypeError):
            row_warnings.append("Invalid cost price, will use 0")
            row_data["cost_price"] = 0
        
        # Quantity validation
        try:
            quantity = int(float(row_data.get("quantity", 0) or 0))
            row_data["quantity"] = quantity
            if quantity < 0:
                row_errors.append("Quantity cannot be negative")
        except (ValueError, TypeError):
            row_errors.append("Invalid quantity value")
        
        # Expiry date validation
        expiry_str = row_data.get("expiry_date")
        if expiry_str:
            try:
                if isinstance(expiry_str, datetime):
                    expiry_date = expiry_str.date()
                elif isinstance(expiry_str, str):
                    # Try multiple date formats
                    for fmt in ['%Y-%m-%d', '%d-%m-%Y', '%d/%m/%Y', '%Y/%m/%d', '%m/%d/%Y']:
                        try:
                            expiry_date = datetime.strptime(str(expiry_str).strip(), fmt).date()
                            break
                        except ValueError:
                            continue
                    else:
                        raise ValueError("Unknown date format")
                else:
                    expiry_date = pd.to_datetime(expiry_str).date()
                
                row_data["expiry_date"] = expiry_date.isoformat()
                
                if expiry_date <= today:
                    row_warnings.append("Expiry date is in the past or today")
                elif expiry_date <= today + timedelta(days=30):
                    row_warnings.append("Expiry date is within 30 days")
            except Exception:
                row_errors.append("Invalid expiry date format. Use YYYY-MM-DD")
        
        # Batch number validation
        batch_no = row_data.get("batch_number")
        if batch_no:
            batch_no = str(batch_no).strip().upper()
            row_data["batch_number"] = batch_no
        
        # Check for duplicate SKU+Batch combination in file
        if sku and batch_no:
            batch_key = f"{sku}_{batch_no}"
            if batch_key in existing_batches:
                row_warnings.append("Batch already exists - will update existing batch")
        
        # GST validation
        try:
            gst = float(row_data.get("gst_percent", 5) or 5)
            row_data["gst_percent"] = gst
            if gst < 0 or gst > 100:
                row_warnings.append("GST % should be between 0 and 100")
        except (ValueError, TypeError):
            row_data["gst_percent"] = 5
        
        # Units per pack validation
        try:
            units_per_pack = int(float(row_data.get("units_per_pack", 1) or 1))
            row_data["units_per_pack"] = max(1, units_per_pack)
        except (ValueError, TypeError):
            row_data["units_per_pack"] = 1
        
        # Determine row status
        if row_errors:
            status = "error"
            error_count += 1
        elif row_warnings:
            status = "warning"
            warning_count += 1
            valid_count += 1
        else:
            status = "valid"
            valid_count += 1
        
        validation_results.append({
            "row_number": row_idx,
            "status": status,
            "errors": row_errors,
            "warnings": row_warnings,
            "data": row_data
        })
    
    # Update job with validation results
    bulk_upload_jobs[job_id]["status"] = "validated"
    bulk_upload_jobs[job_id]["validation_results"] = validation_results
    bulk_upload_jobs[job_id]["column_mapping"] = column_mapping
    bulk_upload_jobs[job_id]["valid_count"] = valid_count
    bulk_upload_jobs[job_id]["error_count"] = error_count
    bulk_upload_jobs[job_id]["warning_count"] = warning_count
    
    return {
        "job_id": job_id,
        "total_rows": len(data),
        "valid_count": valid_count,
        "error_count": error_count,
        "warning_count": warning_count,
        "preview_results": validation_results[:10],  # First 10 for preview
        "can_import": valid_count > 0
    }

@api_router.post("/inventory/bulk-upload/import")
async def import_bulk_upload(
    request: BulkUploadImportRequest,
    background_tasks: BackgroundTasks,
    current_user: User = Depends(get_current_user)
):
    """Import validated data into inventory"""
    
    job_id = request.job_id
    import_valid_only = request.import_valid_only
    
    if job_id not in bulk_upload_jobs:
        raise HTTPException(status_code=404, detail="Upload job not found. Please re-upload the file.")
    
    job = bulk_upload_jobs[job_id]
    
    if job.get("status") != "validated":
        raise HTTPException(status_code=400, detail="Data must be validated before import")
    
    validation_results = job.get("validation_results", [])
    
    # Filter rows to import
    rows_to_import = []
    for result in validation_results:
        if import_valid_only:
            if result["status"] in ["valid", "warning"]:
                rows_to_import.append(result)
        else:
            if result["status"] != "error":
                rows_to_import.append(result)
    
    if not rows_to_import:
        raise HTTPException(status_code=400, detail="No valid rows to import")
    
    # Initialize import progress
    bulk_upload_jobs[job_id]["status"] = "importing"
    bulk_upload_jobs[job_id]["import_progress"] = {
        "total": len(rows_to_import),
        "processed": 0,
        "success": 0,
        "failed": 0,
        "errors": []
    }
    
    # Process import in background
    async def process_import():
        progress = bulk_upload_jobs[job_id]["import_progress"]
        
        for idx, result in enumerate(rows_to_import):
            row_data = result["data"]
            row_number = result["row_number"]
            
            try:
                sku = row_data.get("sku")
                name = row_data.get("name")
                batch_no = row_data.get("batch_number")
                
                # Check if product exists
                existing_product = await db.products.find_one({"sku": sku}, {"_id": 0})
                
                if not existing_product:
                    # Create new product
                    product_doc = {
                        "id": str(uuid.uuid4()),
                        "sku": sku,
                        "name": name,
                        "brand": row_data.get("brand"),
                        "category": row_data.get("category"),
                        "units_per_pack": row_data.get("units_per_pack", 1),
                        "default_mrp_per_unit": row_data.get("price", 0),
                        "gst_percent": row_data.get("gst_percent", 5),
                        "hsn_code": row_data.get("hsn_code"),
                        "low_stock_threshold_units": 10,
                        "status": "active",
                        "created_by": current_user.id,
                        "created_at": datetime.now(timezone.utc).isoformat(),
                        "updated_at": datetime.now(timezone.utc).isoformat()
                    }
                    await db.products.insert_one(product_doc)
                    product_id = product_doc["id"]
                else:
                    product_id = existing_product.get("id")
                    # Optionally update product info
                    update_fields = {}
                    if row_data.get("brand") and not existing_product.get("brand"):
                        update_fields["brand"] = row_data["brand"]
                    if row_data.get("category") and not existing_product.get("category"):
                        update_fields["category"] = row_data["category"]
                    if update_fields:
                        update_fields["updated_at"] = datetime.now(timezone.utc).isoformat()
                        await db.products.update_one({"sku": sku}, {"$set": update_fields})
                
                # Check if batch exists
                existing_batch = await db.stock_batches.find_one(
                    {"product_sku": sku, "batch_no": batch_no},
                    {"_id": 0}
                )
                
                if existing_batch:
                    # Update existing batch
                    new_qty = existing_batch.get("qty_on_hand", 0) + row_data.get("quantity", 0)
                    await db.stock_batches.update_one(
                        {"product_sku": sku, "batch_no": batch_no},
                        {"$set": {
                            "qty_on_hand": new_qty,
                            "mrp_per_unit": row_data.get("price", existing_batch.get("mrp_per_unit")),
                            "cost_price_per_unit": row_data.get("cost_price") or existing_batch.get("cost_price_per_unit", 0),
                            "updated_at": datetime.now(timezone.utc).isoformat(),
                            "updated_by": current_user.id
                        }}
                    )
                    batch_id = existing_batch.get("id")
                else:
                    # Create new batch
                    batch_doc = {
                        "id": str(uuid.uuid4()),
                        "product_sku": sku,
                        "batch_no": batch_no,
                        "expiry_date": row_data.get("expiry_date"),
                        "qty_on_hand": row_data.get("quantity", 0),
                        "cost_price_per_unit": row_data.get("cost_price", 0),
                        "mrp_per_unit": row_data.get("price", 0),
                        "location": "default",
                        "created_by": current_user.id,
                        "created_at": datetime.now(timezone.utc).isoformat(),
                        "updated_at": datetime.now(timezone.utc).isoformat()
                    }
                    await db.stock_batches.insert_one(batch_doc)
                    batch_id = batch_doc["id"]
                
                # Create stock movement record
                movement_doc = {
                    "id": str(uuid.uuid4()),
                    "product_sku": sku,
                    "batch_id": batch_id,
                    "product_name": name,
                    "batch_no": batch_no,
                    "qty_delta_units": row_data.get("quantity", 0),
                    "movement_type": "opening_stock",
                    "ref_type": "bulk_upload",
                    "ref_id": job_id,
                    "location": "default",
                    "reason": f"Bulk upload from {job.get('filename', 'Excel file')}",
                    "performed_by": current_user.id,
                    "performed_at": datetime.now(timezone.utc).isoformat()
                }
                await db.stock_movements.insert_one(movement_doc)
                
                progress["success"] += 1
                
            except Exception as e:
                logger.error(f"Import error row {row_number}: {e}")
                progress["failed"] += 1
                progress["errors"].append({
                    "row_number": row_number,
                    "error": str(e)
                })
            
            progress["processed"] = idx + 1
        
        # Mark as complete
        bulk_upload_jobs[job_id]["status"] = "completed"
        bulk_upload_jobs[job_id]["completed_at"] = datetime.now(timezone.utc).isoformat()
    
    # Run import in background
    background_tasks.add_task(process_import)
    
    return {
        "job_id": job_id,
        "message": "Import started",
        "total_rows": len(rows_to_import)
    }

@api_router.get("/inventory/bulk-upload/progress/{job_id}")
async def get_bulk_upload_progress(
    job_id: str,
    current_user: User = Depends(get_current_user)
):
    """Get import progress for a bulk upload job"""
    
    if job_id not in bulk_upload_jobs:
        raise HTTPException(status_code=404, detail="Upload job not found")
    
    job = bulk_upload_jobs[job_id]
    
    return {
        "job_id": job_id,
        "status": job.get("status"),
        "filename": job.get("filename"),
        "total_rows": job.get("total_rows"),
        "valid_count": job.get("valid_count", 0),
        "error_count": job.get("error_count", 0),
        "warning_count": job.get("warning_count", 0),
        "import_progress": job.get("import_progress"),
        "completed_at": job.get("completed_at")
    }

@api_router.get("/inventory/bulk-upload/error-report/{job_id}")
async def download_error_report(
    job_id: str,
    current_user: User = Depends(get_current_user)
):
    """Download error report for a bulk upload job"""
    
    if job_id not in bulk_upload_jobs:
        raise HTTPException(status_code=404, detail="Upload job not found")
    
    job = bulk_upload_jobs[job_id]
    validation_results = job.get("validation_results", [])
    
    if not validation_results:
        raise HTTPException(status_code=400, detail="No validation results available")
    
    # Create workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Validation Results"
    
    # Styling
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    error_fill = PatternFill(start_color="FFCDD2", end_color="FFCDD2", fill_type="solid")
    warning_fill = PatternFill(start_color="FFF9C4", end_color="FFF9C4", fill_type="solid")
    success_fill = PatternFill(start_color="C8E6C9", end_color="C8E6C9", fill_type="solid")
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    
    # Headers
    headers = ["Row #", "Status", "SKU", "Name", "Batch #", "Expiry", "Qty", "MRP", "Errors", "Warnings"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border
        cell.alignment = Alignment(horizontal="center")
    
    # Data rows
    for row_idx, result in enumerate(validation_results, 2):
        data = result.get("data", {})
        status = result.get("status", "unknown")
        
        row_values = [
            result.get("row_number"),
            status.upper(),
            data.get("sku", ""),
            data.get("name", ""),
            data.get("batch_number", ""),
            data.get("expiry_date", ""),
            data.get("quantity", ""),
            data.get("price", ""),
            "; ".join(result.get("errors", [])),
            "; ".join(result.get("warnings", []))
        ]
        
        row_fill = error_fill if status == "error" else (warning_fill if status == "warning" else success_fill)
        
        for col, value in enumerate(row_values, 1):
            cell = ws.cell(row=row_idx, column=col, value=value)
            cell.border = thin_border
            cell.fill = row_fill
    
    # Adjust column widths
    column_widths = [8, 10, 15, 25, 15, 12, 8, 10, 40, 40]
    for col, width in enumerate(column_widths, 1):
        ws.column_dimensions[ws.cell(row=1, column=col).column_letter].width = width
    
    # Add summary sheet
    ws_summary = wb.create_sheet("Summary")
    summary_data = [
        ["Validation Summary"],
        [""],
        ["Total Rows", job.get("total_rows", 0)],
        ["Valid Rows", job.get("valid_count", 0)],
        ["Rows with Warnings", job.get("warning_count", 0)],
        ["Rows with Errors", job.get("error_count", 0)],
        [""],
        ["File Name", job.get("filename", "")],
        ["Uploaded At", job.get("created_at", "")],
    ]
    
    for row_num, row in enumerate(summary_data, 1):
        for col, value in enumerate(row, 1):
            cell = ws_summary.cell(row=row_num, column=col, value=value if len(row) > col - 1 else "")
            if row_num == 1:
                cell.font = Font(bold=True, size=14)
    
    ws_summary.column_dimensions['A'].width = 25
    ws_summary.column_dimensions['B'].width = 30
    
    # Save to bytes
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    filename = f"validation_report_{job_id[:8]}.xlsx"
    
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )

# Include the API router in the app (must be after all routes are defined)
app.include_router(api_router)